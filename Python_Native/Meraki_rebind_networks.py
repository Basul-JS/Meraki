# Created by JS 
# uses the native python library to rebind a meraki network to a new template 
    # allows the claiming and addining of new devices to the network replacing old devices / models
# 20250905 - updated to enable WAN2 on the new MX's
# 20251001 - paginated getter for networks
# 20251001 - update dhcp handling logic to be more robust
# 20251020 - update to to list the number of networks bound to each template and only lists templates that have less than 90 networks bound
# 20251028 - tighthen the template selection logic to more accurately return templates relevant to the current network/template
# 20251028 - removed redundant and unsed functions
# 20251117 - _port_dict_by_number immediately returns {} if ports isn’t a list (e.g., None), so no iteration over None.
    # - compute_port_overrides normalizes None to [] and guards types before comparing.
# 20251117 - adding checkpoints 
# 20251120 - update the template selection logic
# 20251121 - update to remove session state at end of rollback
# 20251201 - using Python's built-in exceptions instead of relying on 'try/except' blocks for error handling
# 20251205 - allow swapping of VLAN 40 and 41, perserving DHCP reservations etc
# 20251209 - updated logic for template handling
    # Update error handling for org networks fetch to save details and abort on unexpected responses
    # adding helpful CLI hints for debugging API issues
    # removal of Dry RUN mode - all actions are live
# 20260114 - Script updated 
    # Centralised retry helper
    # Per-device retry counters
    #  Rollback VLAN restore with backoff
    # Removed the duplicate / conflicting flatten prompt
    # Only allows VLAN swap later when flatten was explicitly requested
    # 
    
    
"""
Meraki Rebind Networks Utility
Version: 2026.01.14_01
"""

SCRIPT_VERSION = "2026.01.14_01"

import requests
import logging
import re
import json
from datetime import datetime
from getpass import getpass
import csv
import os
import time
import signal
import sys
import random
import unicodedata
from typing import Any, Dict, List, Optional, Tuple, Set, Union, Callable, cast
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from dataclasses import dataclass, asdict
from difflib import SequenceMatcher  # still used elsewhere for network matching if needed
from ipaddress import ip_network, ip_address


# =====================
# Config & Constants
# =====================
EXCLUDED_VLANS = {100, 110, 210, 220, 230, 235, 240}
REQUEST_TIMEOUT = 30  # seconds
BASE_URL = "https://api.meraki.com/api/v1"
MAX_RETRIES = 5

# Logging setup
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
logging.basicConfig(
    filename=f"meraki_script_{timestamp}.log",
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
CSV_LOGFILE = f"meraki_techboost25_rebind_{timestamp}.csv"

# =====================
# Utility: CSV audit log
# =====================
def log_change(
    event: str,
    details: str,
    *,
    username: Optional[str] = None,
    device_serial: Optional[str] = None,
    device_name: Optional[str] = None,
    misc: Optional[str] = None,
    org_id: Optional[str] = None,
    org_name: Optional[str] = None,
    network_id: Optional[str] = None,
    network_name: Optional[str] = None,
    
    ) -> None:
    file_exists = os.path.isfile(CSV_LOGFILE)
    with open(CSV_LOGFILE, mode='a', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        if not file_exists:
            writer.writerow([
                'timestamp', 'event', 'details', 'user',
                'device_serial', 'device_name', 'misc',
                'org_id', 'org_name', 'network_id', 'network_name'
            ])
        writer.writerow([
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            event,
            details,
            username or OPERATOR,
            device_serial or '',
            device_name or '',
            misc or '',
            org_id or '',
            org_name or '',
            network_id or '',
            network_name or ''
            
        ])

# -------------------------
# VLAN update tracking (per run)
# -------------------------
VLAN_UPDATES_OK: List[str] = []
VLAN_UPDATES_SKIPPED: List[str] = []   # e.g., VLAN not found (404)
VLAN_UPDATES_FAILED: List[str] = []    # all non-404 unexpected failures

# =====================
# Debug Mode
# =====================
DEBUG_MODE = False  # Set to True to always print JSON debug file paths
JSON_DEBUG_FILES_CREATED: List[str] = []  # Tracks JSON logs created during the run

# ---- Caches to avoid repeated org-wide scans (reduces 429s) ----
_ORG_NETWORKS_CACHE: Dict[str, Dict[str, Any]] = {}
# shape: { org_id: {"expires": float_epoch, "networks": List[Dict[str, Any]]} }

_TEMPLATE_COUNT_CACHE: Dict[str, Dict[str, int]] = {}
# shape: { org_id: { template_id: count, ... } }

def _get_org_networks_cached(org_id: str, ttl_seconds: int = 120) -> List[Dict[str, Any]]:
    """
    Fetch all org networks once (paginated), cache for ttl_seconds to avoid 429s.

    On unexpected API responses (non-list / HTTP errors / exceptions):
      - Save details into logs/last_networks_error.json
      - Print a clear error and abort the script (SystemExit).

    Returning an empty list is reserved for the *legitimate* case where
    the API returns 200 OK with an empty list of networks.
    """
    now = time.time()
    cached = _ORG_NETWORKS_CACHE.get(org_id)
    if cached and cached.get("expires", 0) > now:
        return cast(List[Dict[str, Any]], cached.get("networks", []))

    nets: List[Dict[str, Any]] = []
    per_page = 1000
    starting_after: Optional[str] = None

    # Collect failures for JSON dump if anything weird happens
    failures: List[Dict[str, Any]] = []

    try:
        while True:
            params: Dict[str, Any] = {"perPage": per_page}
            if starting_after:
                params["startingAfter"] = starting_after

            try:
                page_raw: Any = meraki_get(f"/organizations/{org_id}/networks", params=params)
            except requests.exceptions.HTTPError as e:
                failures.append({
                    "stage": "http_error",
                    "status_code": e.response.status_code if e.response is not None else None,
                    "body": e.response.text if e.response is not None else None,
                    "error": str(e),
                    "params": params,
                })
                raise  # handled by outer except
            except Exception as e:
                failures.append({
                    "stage": "exception",
                    "error_type": type(e).__name__,
                    "error": str(e),
                    "params": params,
                })
                raise

            # Must be a list (normal Meraki behavior)
            if not isinstance(page_raw, list):
                failures.append({
                    "stage": "unexpected_response_type",
                    "response_type": type(page_raw).__name__,
                    "response": page_raw,
                    "params": params,
                })
                raise RuntimeError(
                    f"Unexpected response type from /organizations/{org_id}/networks: "
                    f"{type(page_raw).__name__}"
                )

            page: List[Dict[str, Any]] = page_raw
            if not page:
                break

            nets.extend(page)
            if len(page) < per_page:
                break

            last = page[-1]
            starting_after = str(last.get("id") or "")
            if not starting_after:
                break

        # Normal, successful path: cache & return
        _ORG_NETWORKS_CACHE[org_id] = {"expires": now + ttl_seconds, "networks": nets}
        logging.info("Loaded %d networks for org %s", len(nets), org_id)
        return nets

    except Exception as e:
        # Any error / weird response: write JSON error dump and abort.
        try:
            os.makedirs("logs", exist_ok=True)
            error_path = os.path.join("logs", "last_networks_error.json")
            with open(error_path, "w", encoding="utf-8") as f:
                json.dump(
                    {
                        "org_id": org_id,
                        "timestamp": datetime.now().isoformat(),
                        "error_type": type(e).__name__,
                        "error": str(e),
                        "failures": failures,
                    },
                    f,
                    ensure_ascii=False,
                    indent=2,
                )
            JSON_DEBUG_FILES_CREATED.append(os.path.abspath(error_path))            
            logging.error(
                "Saved org networks fetch failure details to %s (org %s)",
                error_path,
                org_id,
            )
            print(f"\n📝 Network listing failure details saved to {error_path}")
        except Exception:
            logging.exception("Failed to write logs/last_networks_error.json")

        print("\n🚨 ERROR: Unable to list networks for this organization via Meraki API.")
        print("    Please check:")
        print("      - API key and org access")
        print("      - Org ID is correct")
        print("      - Meraki dashboard/API health")
        print("    See logs/last_networks_error.json for full details.\n")

        raise SystemExit(1)

# ---------- Resume Checkpoint ----------
CHECKPOINT_DIR = ".meraki_rebind_state"

@dataclass
class Checkpoint:
    org_id: str
    network_id: str
    network_name: str = ""
    step_status: Optional[Dict[str, Union[bool, str]]] = None
    pre_change_template: Optional[str] = None
    pre_change_devices: Optional[List[Dict[str, Any]]] = None
    pre_change_vlans: Optional[List[Dict[str, Any]]] = None
    claimed_serials: Optional[List[str]] = None
    removed_serials: Optional[List[str]] = None
    suggested_template_id: Optional[str] = None
    bound_template_id: Optional[str] = None

    # NEW: resume-safe fields
    primary_mx_serial: Optional[str] = None
    mr_order: Optional[List[str]] = None
    ms_order: Optional[List[str]] = None
    claimed_models: Optional[Dict[str, str]] = None

    def path(self) -> str:
        os.makedirs(CHECKPOINT_DIR, exist_ok=True)
        fname = f"{self.org_id}_{self.network_id}.json"
        return os.path.join(CHECKPOINT_DIR, fname)

    def save(self) -> None:
        payload = asdict(self)
        with open(self.path(), "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

    @staticmethod
    def load(org_id: str, network_id: str) -> Optional["Checkpoint"]:
        os.makedirs(CHECKPOINT_DIR, exist_ok=True)
        path = os.path.join(CHECKPOINT_DIR, f"{org_id}_{network_id}.json")
        if not os.path.isfile(path):
            return None
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return Checkpoint(
            org_id=data["org_id"],
            network_id=data["network_id"],
            network_name=data.get("network_name", ""),
            step_status=data.get("step_status") or {},
            pre_change_template=data.get("pre_change_template"),
            pre_change_devices=data.get("pre_change_devices") or [],
            pre_change_vlans=data.get("pre_change_vlans") or [],
            claimed_serials=data.get("claimed_serials") or [],
            removed_serials=data.get("removed_serials") or [],
            suggested_template_id=data.get("suggested_template_id"),
            bound_template_id=data.get("bound_template_id"),
            primary_mx_serial=data.get("primary_mx_serial"),
            mr_order=data.get("mr_order") or [],
            ms_order=data.get("ms_order") or [],
            claimed_models=data.get("claimed_models") or {},
        )

    def mark(self, key: str, value: Union[bool, str]) -> None:
        if self.step_status is None:
            self.step_status = {}
        self.step_status[key] = value
        self.save()

    def done(self, key: str) -> bool:
        v = (self.step_status or {}).get(key)
        return bool(v is True or (isinstance(v, str) and v.upper() == "NA"))

_current_checkpoint: Optional[Checkpoint] = None

def _template_counts_for_org(org_id: str) -> Dict[str, int]:
    """
    Compute counts for ALL templates by scanning org networks once.
    Cached per org_id; invalidate when needed by clearing the org entry.
    """
    cached = _TEMPLATE_COUNT_CACHE.get(org_id)
    if cached is not None:
        return cached

    nets = _get_org_networks_cached(org_id)
    counts: Dict[str, int] = {}
    for n in nets:
        tpl_id = n.get("configTemplateId")
        if tpl_id:
            counts[tpl_id] = counts.get(tpl_id, 0) + 1

    _TEMPLATE_COUNT_CACHE[org_id] = counts
    return counts



# =====================
# Prompts
# =====================
OPERATOR = input("Enter your name or initials for audit logs: ")
print("Actions will be executed against the live Meraki environment.")

# =====================
# Time-of-day warning (PROMINENT)
# =====================
now = datetime.now()
cutoff_hour = 17
cutoff_minute = 40
if (now.hour < cutoff_hour) or (now.hour == cutoff_hour and now.minute < cutoff_minute):
    print("\n" + "="*80)
    print("⚠️  WARNING: YOU ARE ABOUT TO MAKE LIVE CHANGES TO THE NETWORK ⚠️")
    print("This will bring down the network if applied during business hours.")
    print("Please Ensure the Store is closed before continuing.")
    print(f"Current time: {now.strftime('%H:%M')}")
    print("Recommended run time: AFTER 17:40.")
    print("="*80 + "\n")
    confirm = input("❗ Type 'YES' to proceed, or anything else to abort: ").strip()
    if confirm.upper() != "YES":
        print("❌ Aborting script.")
        raise SystemExit(1)

# =====================
# API auth
# =====================
def validate_api_key(key: str) -> bool:
    return bool(re.fullmatch(r'[A-Fa-f0-9]{40}', key or ''))

MAX_API_KEY_ATTEMPTS = 4
attempts = 0
API_KEY = None
while attempts < MAX_API_KEY_ATTEMPTS:
    API_KEY = getpass("Enter your Meraki API key (hidden): ")
    if validate_api_key(API_KEY):
        break
    attempts += 1
    print(f"❌ Invalid API key. ({MAX_API_KEY_ATTEMPTS - attempts} attempt(s) left)")
else:
    print("❌ Maximum attempts reached. Exiting.")
    raise SystemExit(1)

HEADERS = {
    "X-Cisco-Meraki-API-Key": API_KEY,
    "Content-Type": "application/json",
    "Accept": "application/json",
}

# Graceful abort
_aborted = False
def _handle_sigint(signum, frame):
    global _aborted, _current_checkpoint
    _aborted = True
    print("\nReceived Ctrl+C — attempting graceful checkpoint & shutdown...")
    try:
        log_change('workflow_abort', 'User interrupted with SIGINT')
    finally:
        if _current_checkpoint is not None:
            try:
                _current_checkpoint.save()
                print(f"Saved progress to {_current_checkpoint.path()}")
            except Exception:
                logging.exception("Failed to save checkpoint on SIGINT")
    raise SystemExit(1)

signal.signal(signal.SIGINT, _handle_sigint)

# =====================
# HTTP layer
# =====================

def _request(method: str, path: str, *, params=None, json_data=None) -> Any:
    """
    Low-level HTTP wrapper.

    * Uses requests' built-in HTTPError (via raise_for_status)
    * Retries on generic RequestException and 429 responses
    * Returns parsed JSON when possible, else raw text / None
    """
    url = f"{BASE_URL}{path}"

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            if method == 'GET':
                resp = requests.get(url, headers=HEADERS, params=params, timeout=REQUEST_TIMEOUT)
            elif method == 'POST':
                resp = requests.post(url, headers=HEADERS, json=json_data, timeout=REQUEST_TIMEOUT)
            elif method == 'PUT':
                resp = requests.put(url, headers=HEADERS, json=json_data, timeout=REQUEST_TIMEOUT)
            elif method == 'DELETE':
                resp = requests.delete(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
            else:
                raise ValueError(f"Unknown HTTP method: {method}")

            # Handle rate-limiting with backoff
            if resp.status_code == 429:
                ra = resp.headers.get("Retry-After")
                if ra is not None:
                    try:
                        wait = max(0.5, min(float(ra), 30.0))
                    except ValueError:
                        wait = min(2 ** (attempt - 1), 30)
                else:
                    wait = min(2 ** (attempt - 1), 30)
                logging.warning(f"429 rate limit for {url}. Sleeping {wait}s and retrying...")
                time.sleep(wait)
                continue

            # Raise for HTTP errors (built-in HTTPError from requests)
            try:
                resp.raise_for_status()
            except requests.exceptions.HTTPError as http_err:
                logging.error("%s %s -> %s %s", method, url, resp.status_code, resp.text)
                raise

            if resp.text:
                try:
                    return resp.json()
                except json.JSONDecodeError:
                    return resp.text
            return None

        except requests.exceptions.HTTPError:
            # Don't retry on HTTP errors other than 429 (handled above)
            raise
        except requests.exceptions.RequestException as e:
            # Network / connection / timeout errors
            if attempt == MAX_RETRIES:
                logging.exception(f"HTTP error for {url}: {e}")
                raise
            wait = min(2 ** attempt, 30)
            logging.warning(f"HTTP exception {e} for {url}. Retrying in {wait}s...")
            time.sleep(wait)

def meraki_get(path, params=None):
    return _request('GET', path, params=params)

def meraki_post(path, data=None):
    return _request('POST', path, json_data=data)

def meraki_put(path, data=None):
    return _request('PUT', path, json_data=data)

def meraki_delete(path):
    return _request('DELETE', path)

def do_action(func, *args, **kwargs):
    """
    Wrapper for Meraki API actions (no dry-run mode).
    Always executes the provided function.
    """
    return func(*args, **kwargs)

def put_device_with_transient_404_retry(
    org_id: str,
    serial: str,
    data: Dict[str, Any],
    *,
    attempts: int = 6,
    initial_sleep: float = 1.0,
) -> None:
    """
    Retry PUT /devices/{serial} on 404 because Meraki can be eventually-consistent
    immediately after claim/bind/remove operations.
    """
    sleep_s = initial_sleep

    for attempt in range(1, attempts + 1):
        try:
            do_action(meraki_put, f"/devices/{serial}", data=data)
            return
        except requests.exceptions.HTTPError as e:
            resp = e.response
            status = resp.status_code if resp is not None else None

            # Treat 404 as transient (only for this endpoint) for a few attempts
            if status == 404 and attempt < attempts:
                inv_net = None
                inv_model = None
                try:
                    inv = get_inventory_device(org_id, serial) or {}
                    inv_net = inv.get("networkId")
                    inv_model = inv.get("model")
                except Exception:
                    pass

                logging.warning(
                    "Transient 404 on PUT /devices/%s (attempt %d/%d). inventory.networkId=%r model=%r. Sleep %.1fs then retry.",
                    serial, attempt, attempts, inv_net, inv_model, sleep_s
                )
                time.sleep(sleep_s)
                sleep_s = min(sleep_s * 2, 10.0)
                continue

            raise

# =====================
# Shared API helpers (normalized)
# =====================
def get_inventory_device(org_id: str, serial: str) -> Dict[str, Any]:
    """Single, consistent inventory lookup endpoint."""
    return meraki_get(f"/organizations/{org_id}/inventory/devices/{serial}") or {}

# -------------------------
# Centralised retry helper
# -------------------------

@dataclass
class RetryOutcome:
    ok: bool
    attempts: int
    error_type: Optional[str] = None
    error: Optional[str] = None
    http_status: Optional[int] = None
    last_response_text: Optional[str] = None

def _is_retryable_http_status(status: Optional[int]) -> bool:
    # Meraki transient-ish statuses (besides 429 which your _request already handles)
    return status in {408, 409, 500, 502, 503, 504}

def retry_with_backoff(
    op_name: str,
    fn: Callable[[], Any],
    *,
    max_attempts: int = 6,
    base_sleep: float = 1.0,
    max_sleep: float = 15.0,
    retry_on_http_statuses: Optional[Set[int]] = None,
    retry_on_exceptions: Tuple[type, ...] = (requests.exceptions.RequestException,),
    verbose: bool = False,
) -> RetryOutcome:
    """
    Central retry wrapper used across the script.

    - Retries on retryable HTTP statuses (default: 408/409/5xx)
    - Retries on selected transient exception classes
    """

    if retry_on_http_statuses is None:
        retry_on_http_statuses = {408, 409, 500, 502, 503, 504}

    sleep_s = base_sleep
    last_exc: Optional[Exception] = None

    for attempt in range(1, max_attempts + 1):
        try:
            fn()
            return RetryOutcome(ok=True, attempts=attempt)

        except requests.exceptions.HTTPError as e:
            last_exc = e
            resp = e.response
            status = resp.status_code if resp is not None else None
            text = (resp.text or "")[:500] if resp is not None else None

            if status in retry_on_http_statuses and attempt < max_attempts:
                logging.warning(
                    "[%s] HTTP %s on attempt %d/%d. Backing off %.1fs.",
                    op_name, status, attempt, max_attempts, sleep_s
                )
                if verbose:
                    print(
                        f"⚠️  {op_name}: HTTP {status} "
                        f"(attempt {attempt}/{max_attempts}) retrying in {sleep_s:.1f}s"
                    )
                time.sleep(sleep_s)
                sleep_s = min(sleep_s * 2, max_sleep)
                continue

            return RetryOutcome(
                ok=False,
                attempts=attempt,
                error_type=type(e).__name__,
                error=str(e),
                http_status=status,
                last_response_text=text,
            )

        except Exception as e:
            last_exc = e

            # ✅ runtime filter for retryable exception types
            if isinstance(e, retry_on_exceptions) and attempt < max_attempts:
                logging.warning(
                    "[%s] %s on attempt %d/%d. Backing off %.1fs.",
                    op_name, type(e).__name__, attempt, max_attempts, sleep_s
                )
                if verbose:
                    print(
                        f"⚠️  {op_name}: {type(e).__name__} "
                        f"(attempt {attempt}/{max_attempts}) retrying in {sleep_s:.1f}s"
                    )
                time.sleep(sleep_s)
                sleep_s = min(sleep_s * 2, max_sleep)
                continue

            # Non-retryable OR max attempts reached
            return RetryOutcome(
                ok=False,
                attempts=attempt,
                error_type=type(e).__name__,
                error=str(e),
            )

    # Safety net (should never hit)
    return RetryOutcome(
        ok=False,
        attempts=max_attempts,
        error_type=type(last_exc).__name__ if last_exc else None,
        error=str(last_exc) if last_exc else None,
    )


def _write_json_log(path: str, payload: Dict[str, Any]) -> None:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    JSON_DEBUG_FILES_CREATED.append(os.path.abspath(path))


# # ======================================================
# # ------------- Wireless pre-check helpers -------------
# # ======================================================

WIRELESS_PREFIXES: Tuple[str, ...] = ("MR", "CW")

def _prompt_yes_no(question: str, default_no: bool = True) -> bool:
    prompt = " [y/N] " if default_no else " [Y/n] "
    ans_raw = input(question + prompt).strip().lower()
    if not ans_raw:
        return (not default_no)
    return ans_raw in ("y", "yes")

def _prompt_replacement_mapping(old_serials: List[str], new_serials: List[str]) -> List[Tuple[str, str]]:
    if not old_serials or not new_serials:
        return []
    print("\nEnter replacement pairs as 'OLD:NEW'. Leave blank to finish.")
    print(f"Old (present in network): {', '.join(old_serials)}")
    print(f"New (available to add):   {', '.join(new_serials)}")

    available_new: Set[str] = {s.upper() for s in new_serials}
    old_upper: Set[str] = {s.upper() for s in old_serials}
    mapping: List[Tuple[str, str]] = []

    def _restore_case(target: str, pool: List[str]) -> str:
        for p in pool:
            if p.upper() == target.upper():
                return p
        return target

    while True:
        line = input("Pair (OLD:NEW): ").strip()
        if not line:
            break
        if ":" not in line:
            print("  Format must be OLD:NEW")
            continue
        old_s, new_s = [p.strip() for p in line.split(":", 1)]
        if old_s.upper() not in old_upper:
            print(f"  {old_s} is not in the old-serials list.")
            continue
        if new_s.upper() not in available_new:
            print(f"  {new_s} is not in the new-serials list or already used.")
            continue
        mapping.append((_restore_case(old_s, old_serials), _restore_case(new_s, new_serials)))
        available_new.remove(new_s.upper())
    return mapping

def run_wireless_precheck_and_filter_claims(
    org_id: str,
    network_id: str,
    prevalidated_serials: List[str],
    *,
    block_wireless: bool = False,
) -> Tuple[List[str], List[str], List[str]]:
    mr_removed_serials: List[str] = []
    mr_claimed_serials: List[str] = []

    try:
        prevalidated_serials, mr_removed_serials, mr_claimed_serials = ensure_mr33_and_handle_wireless_replacements(
            org_id, network_id, prevalidated_serials
        )
    except SystemExit:
        raise
    except Exception:
        logging.exception("Wireless pre-check/replacement step failed")

    if block_wireless:
        inv_models = _get_inventory_models_for_serials(org_id, prevalidated_serials)
        wireless_block = {s for s, m in inv_models.items() if _is_wireless_model(m)}
    else:
        wireless_block = set()

    do_not_claim = wireless_block | set(mr_claimed_serials)
    safe_to_claim = [s for s in prevalidated_serials if s not in do_not_claim]

    return safe_to_claim, mr_removed_serials, mr_claimed_serials

def _is_wireless_model(model: Optional[str]) -> bool:
    return bool(model) and model.upper().startswith(WIRELESS_PREFIXES)

def _is_mr33(model: Optional[str]) -> bool:
    return bool(model) and model.upper().startswith("MR33")

def _get_network_wireless_devices(network_id: str) -> List[Dict[str, Any]]:
    try:
        devices: List[Dict[str, Any]] = meraki_get(f"/networks/{network_id}/devices") or []
    except Exception:
        logging.exception("Failed to list devices for wireless check")
        return []
    return [d for d in devices if _is_wireless_model(cast(Optional[str], d.get("model")))]

def _get_inventory_models_for_serials(org_id: str, serials: List[str]) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for s in serials:
        try:
            inv = get_inventory_device(org_id, s) or {}
            mdl = cast(Optional[str], inv.get("model"))
            if mdl:
                out[s] = mdl
        except Exception:
            logging.exception("Inventory lookup failed for %s", s)
    return out

def ensure_mr33_and_handle_wireless_replacements(
    org_id: str,
    network_id: str,
    serials_to_add: List[str],
) -> Tuple[List[str], List[str], List[str]]:
    """
    Returns:
        (serials_safe_to_claim, removed_old_wireless, claimed_new_wireless)
    """
    add_models: Dict[str, str] = _get_inventory_models_for_serials(org_id, serials_to_add)
    incoming_wireless: List[str] = [s for s, m in add_models.items() if _is_wireless_model(m)]
    if not incoming_wireless:
        return serials_to_add, [], []

    wireless_now: List[Dict[str, Any]] = _get_network_wireless_devices(network_id)
    has_mr33_now: bool = any(_is_mr33(cast(Optional[str], d.get("model"))) for d in wireless_now)
    adding_has_mr33: bool = any(_is_mr33(add_models.get(s)) for s in incoming_wireless)

    if not has_mr33_now and not adding_has_mr33:
        proceed = _prompt_yes_no("No MR33 detected in network or incoming. Proceed with wireless changes?", default_no=True)
        if not proceed:
            print("Skipping wireless add/remove per operator choice; continuing with the rest of the workflow.")
            log_change('wireless_skip', "Operator chose to skip wireless changes due to no MR33 present")
            non_wireless = [s for s in serials_to_add if s not in incoming_wireless]
            return non_wireless, [], []

    non_mr33_in_net: List[Dict[str, Any]] = [
        d for d in wireless_now
        if _is_wireless_model(cast(Optional[str], d.get("model"))) and not _is_mr33(cast(Optional[str], d.get("model")))
    ]

    removed_old: List[str] = []
    claimed_new: List[str] = []

    if non_mr33_in_net and _prompt_yes_no("Replace non-MR33 wireless with incoming?", default_no=False):
        mapping = _prompt_replacement_mapping(
            [cast(str, d.get("serial")) for d in non_mr33_in_net if d.get("serial")],
            incoming_wireless
        )
        for old_serial, new_serial in mapping:
            try:
                do_action(meraki_put, f"/devices/{old_serial}", data={"name": "", "address": ""})
                do_action(meraki_post, f"/networks/{network_id}/devices/remove", data={"serial": old_serial})
                log_change('wireless_replace_remove', f"Removed old wireless {old_serial}", device_serial=old_serial)
                removed_old.append(old_serial)
            except Exception:
                logging.exception("Failed to remove %s", old_serial)
            try:
                do_action(meraki_post, f"/networks/{network_id}/devices/claim", data={"serials": [new_serial]})
                log_change('wireless_replace_claim', f"Claimed new wireless {new_serial}", device_serial=new_serial)
                claimed_new.append(new_serial)
            except Exception:
                logging.exception("Failed to claim %s", new_serial)

    claimed_new_set: Set[str] = set(claimed_new)
    serials_out = [s for s in serials_to_add if s not in claimed_new_set]
    return serials_out, removed_old, claimed_new

# =====================
# VLAN error detector (robust)
# =====================
def is_vlans_disabled_error(exc: Exception) -> bool:
    """
    Detects "VLANs are not enabled for this network" on built-in HTTPError
    or any other exception type by inspecting the associated response/text.
    """
    needle = "VLANs are not enabled for this network"

    # Handle requests' HTTPError (built-in library exception)
    if isinstance(exc, requests.exceptions.HTTPError):
        resp = exc.response
        if resp is not None:
            try:
                if resp.status_code == 400:
                    try:
                        body = resp.json()
                    except json.JSONDecodeError:
                        body = None
                    if isinstance(body, dict):
                        errs = body.get("errors")
                        if errs and any(needle in str(e) for e in errs):
                            return True
                if needle in (resp.text or ""):
                    return True
            except Exception:
                # Fall through to generic string check below
                pass

    # Fallback: search the exception’s string representation
    return needle in str(exc)

# =====================
# Switch port helpers (diff + apply)
# =====================
_PORT_FIELDS = [
    "enabled", "name", "tags", "type", "vlan", "voiceVlan", "allowedVlans",
    "poeEnabled", "isolationEnabled", "rstpEnabled", "stpGuard",
    "linkNegotiation", "udld", "accessPolicyType", "accessPolicyNumber",
    "portScheduleId"
]

def _normalize_tags(value):
    if isinstance(value, list):
        return sorted(value)
    if isinstance(value, str):
        return sorted([t for t in value.split() if t])
    return []

def _port_dict_by_number(ports: Optional[List[Dict[str, Any]]]) -> Dict[str, Dict[str, Any]]:
    """
    Build a dict keyed by port number/id. Tolerates None and non-dict entries.
    """
    out: Dict[str, Dict[str, Any]] = {}
    if not isinstance(ports, list):
        return out
    for p in ports:
        if not isinstance(p, dict):
            continue
        pid = p.get("portId") or p.get("number") or p.get("name")
        if pid is None:
            continue
        out[str(pid)] = p
    return out

def compute_port_overrides(
    live_ports: Optional[List[Dict[str, Any]]],
    tmpl_ports: Optional[List[Dict[str, Any]]]
) -> Dict[str, Dict[str, Any]]:
    """
    Compare live vs template ports and return the fields from live that differ.
    Always returns a dict; tolerates None/invalid inputs.
    """
    overrides: Dict[str, Dict[str, Any]] = {}

    live_list = live_ports if isinstance(live_ports, list) else []
    tmpl_list = tmpl_ports if isinstance(tmpl_ports, list) else []

    live = _port_dict_by_number(live_list)
    tmpl = _port_dict_by_number(tmpl_list)

    for pid, lp in live.items():
        tp = tmpl.get(pid)
        if not isinstance(tp, dict):
            continue
        for fld in _PORT_FIELDS:
            lv = lp.get(fld)
            tv = tp.get(fld)
            if fld == "tags":
                lv = _normalize_tags(lv)
                tv = _normalize_tags(tv)
            if lv is not None and lv != tv:
                overrides.setdefault(pid, {})[fld] = lv

    return overrides

def apply_port_overrides(serial: str, overrides: Dict[str, Dict[str, Any]]) -> None:
    for pid, patch in overrides.items():
        try:
            do_action(meraki_put, f"/devices/{serial}/switch/ports/{pid}", data=patch)
            logging.debug(f"Applied port overrides on {serial} port {pid}: {patch}")
            log_change(
                'switch_port_override',
                f"Applied port overrides on port {pid}",
                device_serial=serial,
                misc=json.dumps(patch)
            )
        except Exception:
            logging.exception(f"Failed applying port overrides on {serial} port {pid}")

# =====================
# Domain helpers (raw API)
# =====================

def meraki_list_networks_all(org_id: str) -> List[Dict[str, Any]]:
    # Prefer the cached scan to avoid re-paginating during the same run
    return list(_get_org_networks_cached(org_id))

def _norm(s: Optional[str]) -> str:
    base: str = s or ""
    base = unicodedata.normalize("NFKC", base)
    base = base.replace("–", "-").replace("—", "-")
    base = re.sub(r"\s+", " ", base).strip()
    return base.casefold()

def fetch_matching_networks(org_id: str, partial: str) -> List[Dict[str, Any]]:
    partial_n: str = _norm(partial)
    nets: List[Dict[str, Any]] = meraki_list_networks_all(org_id)
    matches: List[Dict[str, Any]] = []
    for n in nets:
        name = _norm(n.get("name"))
        if partial_n in name:
            matches.append(n)

    logging.debug("Found %d networks matching '%s' (normalized)", len(matches), partial)
    return matches

def fetch_devices(
    org_id: str,
    network_id: str,
    template_id: Optional[str] = None
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Returns (mx_list, ms_list, mr_list). Never returns None on any path.
    """
    # 1) Get devices (robust against None / non-list / exceptions)
    try:
        raw = meraki_get(f"/networks/{network_id}/devices")
    except Exception:
        logging.exception("Failed to fetch devices for network %s", network_id)
        raw = None

    devs: List[Dict[str, Any]] = raw if isinstance(raw, list) else []
    if raw is None:
        logging.warning("devices endpoint returned None for network %s; treating as empty list", network_id)

    # 2) Normalizer
    def _mk(d: Dict[str, Any]) -> Dict[str, Any]:
        model = str(d.get('model', '') or '')
        tags = d.get('tags', [])
        if not isinstance(tags, list):
            tags = (str(tags) or '').split()
        return {
            'serial': d.get('serial', ''),
            'model': model,
            'tags': tags,
            'address': d.get('address', ''),
            'name': d.get('name', ''),
            'switchProfileId': d.get('switchProfileId'),
            'switchProfileName': d.get('switchProfileName'),
        }

    # 3) Split by product type (safe for missing model)
    mx: List[Dict[str, Any]] = [_mk(d) for d in devs if str(d.get('model', '')).startswith('MX')]
    ms: List[Dict[str, Any]] = [_mk(d) for d in devs if str(d.get('model', '')).startswith('MS')]
    mr: List[Dict[str, Any]] = [_mk(d) for d in devs if _is_wireless_model(str(d.get('model', '')))]

    # 4) Compute MS port overrides if we have a template_id
    if template_id:
        for sw in ms:
            profile_id = sw.get('switchProfileId')
            if not profile_id:
                sw['port_overrides'] = {}
                continue
            try:
                live_ports_raw = meraki_get(f"/devices/{sw['serial']}/switch/ports")
                tmpl_ports_raw = meraki_get(
                    f"/organizations/{org_id}/configTemplates/{template_id}/switch/profiles/{profile_id}/ports"
                )
                live_ports: List[Dict[str, Any]] = live_ports_raw if isinstance(live_ports_raw, list) else []
                tmpl_ports: List[Dict[str, Any]] = tmpl_ports_raw if isinstance(tmpl_ports_raw, list) else []
                sw['port_overrides'] = compute_port_overrides(live_ports, tmpl_ports)
                logging.debug("Computed %d port overrides for %s", len(sw['port_overrides']), sw['serial'])
            except Exception:
                logging.exception("Failed computing port overrides for %s", sw.get('serial') or '<unknown>')
                sw['port_overrides'] = {}
    else:
        for sw in ms:
            sw['port_overrides'] = {}

    logging.debug("Fetched devices: MX=%d, MS=%d, MR=%d", len(mx), len(ms), len(mr))
    log_change(
        event='fetch_devices',
        details=f"Fetched devices for network {network_id}",
        network_id=network_id,
        misc=f"mx={json.dumps(mx)}, ms={json.dumps(ms)}, mr={json.dumps(mr)}"
    )

    # 5) ALWAYS return the triple
    return mx, ms, mr

def fetch_vlan_details(network_id: str, *, retries: int = 3, delay: float = 2.0) -> List[Dict[str, Any]]:
    """
    Fetch VLANs for a network with retry logic.
    If no VLANs are returned or response is invalid after all retries:
      - Save failure details into logs/last_vlan_error.json
      - Abort the script (SystemExit), effectively rolling back by not making any changes.
    """
    failures: List[Dict[str, Any]] = []

    for attempt in range(1, retries + 1):
        try:
            vlans_raw = meraki_get(f"/networks/{network_id}/appliance/vlans")

            # Record raw response for debugging
            failures.append({
                "attempt": attempt,
                "stage": "response",
                "response_type": type(vlans_raw).__name__,
                "response": vlans_raw,
            })

            # Must be a list
            if not isinstance(vlans_raw, list):
                msg = f"VLAN API returned invalid response type ({type(vlans_raw).__name__})."
                print(f"❌ {msg} Retrying ({attempt}/{retries})...")
                logging.warning("%s Network=%s, raw=%r", msg, network_id, vlans_raw)
                time.sleep(delay)
                continue

            # Empty list? Possible API or propagation issue
            if not vlans_raw:
                msg = f"No VLANs returned for network {network_id}."
                print(f"❌ {msg} Retrying ({attempt}/{retries})...")
                logging.warning("%s Attempt=%d", msg, attempt)
                time.sleep(delay)
                continue

            # ---------- Filter out excluded VLANs safely ----------
            filtered: List[Dict[str, Any]] = []

            for v in vlans_raw:
                if not isinstance(v, dict):
                    continue

                vid_raw = v.get("id")
                try:
                    vid = int(str(vid_raw))
                except Exception:
                    # Keep non-numeric VLANs rather than failing
                    filtered.append(v)
                    continue

                if vid in EXCLUDED_VLANS:
                    continue

                filtered.append(v)

            logging.debug(
                "Fetched VLANs for %s: %d kept (excluded %d)",
                network_id,
                len(filtered),
                len(vlans_raw) - len(filtered),
            )
            return filtered

        except requests.exceptions.HTTPError as e:
            failures.append({
                "attempt": attempt,
                "stage": "http_error",
                "status_code": e.response.status_code if e.response is not None else None,
                "body": e.response.text if e.response is not None else None,
                "error": str(e),
            })
            if is_vlans_disabled_error(e):
                msg = f"VLANs are disabled on network {network_id}."
                print(f"⚠️ {msg} Aborting.")
                logging.warning("%s HTTP error: %s", msg, e)
                break  # no point retrying this case

            logging.error("HTTP error while fetching VLANs for %s: %s", network_id, e)
            print(f"❌ Error fetching VLANs (HTTPError). Retrying ({attempt}/{retries})...")
            time.sleep(delay)
            continue

        except Exception as e:
            failures.append({
                "attempt": attempt,
                "stage": "exception",
                "error_type": type(e).__name__,
                "error": str(e),
            })
            logging.exception("Unexpected error fetching VLANs for %s", network_id)
            print(f"❌ Unexpected error fetching VLANs: {e}. Retrying ({attempt}/{retries})...")
            time.sleep(delay)
            continue

    # ---------- After all retries failed (or VLANs disabled) ----------
    try:
        os.makedirs("logs", exist_ok=True)
        error_path = os.path.join("logs", "last_vlan_error.json")
        with open(error_path, "w", encoding="utf-8") as f:
            json.dump(
                {
                    "network_id": network_id,
                    "timestamp": datetime.now().isoformat(),
                    "retries": retries,
                    "failures": failures,
                },
                f,
                ensure_ascii=False,
                indent=2,
            )
        
        JSON_DEBUG_FILES_CREATED.append(os.path.abspath(error_path))    
        logging.error("Saved VLAN fetch failure details to %s", error_path)
        print(f"\n📝 VLAN failure details saved to {error_path}")
    except Exception:
        logging.exception("Failed to write logs/last_vlan_error.json")

    print(f"\n🚨 ERROR: Unable to fetch VLANs for network {network_id} after {retries} attempts.")
    print("    No changes have been made. Please check Meraki dashboard/API availability and retry.\n")

    # Abort script – at this stage no changes have been applied, so this is
    # effectively a safe rollback to the original state.
    raise SystemExit(1)
def vlans_enabled(network_id: str, *, retries: int = 3, delay: float = 2.0) -> Optional[bool]:
    """
    Safely determine if VLANs are enabled on the MX, with retry + JSON error logging.

    Returns:
        True  -> VLANs enabled
        False -> VLANs disabled
        None  -> Could not determine state after retries (see logs/last_vlan_settings_error.json)
    """
    failures: List[Dict[str, Any]] = []

    for attempt in range(1, retries + 1):
        try:
            settings = meraki_get(f"/networks/{network_id}/appliance/vlans/settings")

            # Track raw response for debugging
            failures.append({
                "attempt": attempt,
                "stage": "response",
                "response_type": type(settings).__name__,
                "response": settings,
            })

            # Must be a dict for .get() to make sense
            if not isinstance(settings, dict):
                msg = f"VLAN settings API returned invalid response type ({type(settings).__name__})."
                print(f"❌ {msg} Retrying ({attempt}/{retries})...")
                logging.warning("%s Network=%s, raw=%r", msg, network_id, settings)
                time.sleep(delay)
                continue

            # If key is present, we can return a definitive answer
            if "vlansEnabled" in settings:
                return bool(settings.get("vlansEnabled"))

            # Missing key: could be transient or odd MX state – retry
            msg = "VLAN settings payload missing 'vlansEnabled' field."
            print(f"❌ {msg} Retrying ({attempt}/{retries})...")
            logging.warning("%s Network=%s settings=%r", msg, network_id, settings)
            time.sleep(delay)
            continue

        except requests.exceptions.HTTPError as e:
            failures.append({
                "attempt": attempt,
                "stage": "http_error",
                "status_code": e.response.status_code if e.response is not None else None,
                "body": e.response.text if e.response is not None else None,
                "error": str(e),
            })

            # If Meraki explicitly tells us VLANs are disabled, that’s a valid False result
            if is_vlans_disabled_error(e):
                logging.warning("VLAN settings indicate VLANs are disabled on %s (HTTP error: %s)", network_id, e)
                return False

            logging.error("HTTP error while reading VLAN settings for %s: %s", network_id, e)
            print(f"❌ Error reading VLAN settings (HTTPError). Retrying ({attempt}/{retries})...")
            time.sleep(delay)
            continue

        except Exception as e:
            failures.append({
                "attempt": attempt,
                "stage": "exception",
                "error_type": type(e).__name__,
                "error": str(e),
            })
            logging.exception("Unexpected error reading VLAN settings for %s", network_id)
            print(f"❌ Unexpected error reading VLAN settings: {e}. Retrying ({attempt}/{retries})...")
            time.sleep(delay)
            continue

    # ---------- After all retries failed ----------
    try:
        os.makedirs("logs", exist_ok=True)
        error_path = os.path.join("logs", "last_vlan_settings_error.json")
        with open(error_path, "w", encoding="utf-8") as f:
            json.dump(
                {
                    "network_id": network_id,
                    "timestamp": datetime.now().isoformat(),
                    "retries": retries,
                    "failures": failures,
                },
                f,
                ensure_ascii=False,
                indent=2,
            )
        JSON_DEBUG_FILES_CREATED.append(os.path.abspath(error_path))    
        logging.error("Saved VLAN settings failure details to %s", error_path)
        print(f"\n📝 VLAN settings failure details saved to {error_path}")
    except Exception:
        logging.exception("Failed to write logs/last_vlan_settings_error.json")

    print(f"\n🚨 ERROR: Unable to read VLAN settings for network {network_id} after {retries} attempts.")
    print("    VLAN enabled/disabled state is unknown.\n")

    # Let the caller decide whether to rollback/abort based on this.
    return None


def _preserve_dhcp_state_for_vlan(
    original_vlan: Dict[str, Any],
    new_vlan: Dict[str, Any],
    new_fixed: Optional[Dict[str, Any]],
    new_ranges: Optional[List[Dict[str, Any]]],
) -> None:
    """
    Take the DHCP state from original_vlan and impose it on new_vlan:
    - Preserve dhcpHandling string (server/relay/off)
    - For 'server', use new_fixed/new_ranges + original dnsNameservers
    - For 'relay', keep original relay IP(s)
    - For 'off', strip DHCP-specific fields
    """
    orig_dhcp = original_vlan.get("dhcpHandling")
    mode = _dhcp_mode(orig_dhcp)  # uses your existing helper

    # Always preserve the original dhcpHandling string
    if orig_dhcp is not None:
        new_vlan["dhcpHandling"] = orig_dhcp

    # Clear conflicting fields so we don't mix server+relay settings
    for k in ("fixedIpAssignments", "reservedIpRanges", "dnsNameservers",
              "dhcpRelayServerIps", "dhcpRelayServerIp"):
        if k in new_vlan:
            new_vlan.pop(k, None)

    if mode == "server":
        # Re-IP reservations into the new subnet, preserving last octet
        if new_fixed is not None:
            new_vlan["fixedIpAssignments"] = new_fixed
        if new_ranges is not None:
            new_vlan["reservedIpRanges"] = new_ranges
        dns = original_vlan.get("dnsNameservers")
        if dns is not None:
            new_vlan["dnsNameservers"] = dns

    elif mode == "relay":
        relay_ips = (
            original_vlan.get("dhcpRelayServerIps")
            or original_vlan.get("dhcpRelayServerIp")
        )
        if relay_ips is not None:
            new_vlan["dhcpRelayServerIps"] = relay_ips

    # mode == "off": nothing to re-add; we already cleared DHCP fields

def _dhcp_mode(val: Optional[str]) -> str:
    v = (val or "").strip().lower()
    if v in {"run a dhcp server", "run dhcp server", "server", "enabled", "on"}:
        return "server"
    if "relay" in v:
        return "relay"
    if v in {"do not respond", "do not respond to dhcp requests", "off", "disabled", "none"}:
        return "off"
    return "off"

def _nonempty(x: Any) -> bool:
    if x is None:
        return False
    if isinstance(x, (list, dict)) and len(x) == 0:
        return False
    if isinstance(x, str) and x.strip() == "":
        return False
    return True

def update_vlans(network_id: str, network_name: str, vlan_list: List[Dict[str, Any]]) -> None:
    for v in vlan_list:
        vlan_id = str(v.get("id", ""))
        payload: Dict[str, Any] = {}

        if _nonempty(v.get("applianceIp")):
            payload["applianceIp"] = v.get("applianceIp")
        if _nonempty(v.get("subnet")):
            payload["subnet"] = v.get("subnet")

        dhcp_handling_raw = v.get("dhcpHandling")
        if _nonempty(dhcp_handling_raw):
            payload["dhcpHandling"] = dhcp_handling_raw

        mode = _dhcp_mode(dhcp_handling_raw)

        if mode == "server":
            if _nonempty(v.get("fixedIpAssignments")):
                payload["fixedIpAssignments"] = v.get("fixedIpAssignments")
            if _nonempty(v.get("reservedIpRanges")):
                payload["reservedIpRanges"] = v.get("reservedIpRanges")
            if _nonempty(v.get("dnsNameservers")):
                payload["dnsNameservers"] = v.get("dnsNameservers")
        elif mode == "relay":
            relay_ips = v.get("dhcpRelayServerIps") or v.get("dhcpRelayServerIp")
            if _nonempty(relay_ips):
                payload["dhcpRelayServerIps"] = relay_ips

        try:
            do_action(
                meraki_put,
                f"/networks/{network_id}/appliance/vlans/{vlan_id}",
                data=payload,
            )

            VLAN_UPDATES_OK.append(vlan_id)
            logging.debug("Updated VLAN %s with payload: %s", vlan_id, payload)
            log_change(
                "vlan_update",
                f"Updated VLAN {vlan_id}",
                device_name=f"Network: {network_id}",
                network_id=network_id,
                network_name=network_name,
                misc=json.dumps(payload),
            )
            continue

        except requests.exceptions.HTTPError as e:
            resp = e.response
            status = resp.status_code if resp is not None else None

            # VLANs disabled is fatal — allow caller to rollback
            if is_vlans_disabled_error(e):
                VLAN_UPDATES_FAILED.append(vlan_id)
                raise

            # 404 Not Found → VLAN ID removed / doesn't exist → safe skip
            if status == 404:
                VLAN_UPDATES_SKIPPED.append(vlan_id)
                logging.warning(
                    "VLAN %s not found on network %s — skipping update. Payload=%r",
                    vlan_id, network_id, payload,
                )
                print(f"⚠️  VLAN {vlan_id} not present on network — skipped.")
                continue

            # Anything else is a failure, but non-fatal
            VLAN_UPDATES_FAILED.append(vlan_id)
            logging.exception(
                "Failed to update VLAN %s (HTTP %s): %s",
                vlan_id,
                status if status is not None else "?",
                str(e),
            )

        except Exception:
            VLAN_UPDATES_FAILED.append(vlan_id)
            logging.exception("Failed to update VLAN %s", vlan_id)

def print_vlan_update_summary() -> None:
    print("\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print("📘 VLAN Update Summary")
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")

    if VLAN_UPDATES_OK:
        print(f"✅ Updated VLANs ({len(VLAN_UPDATES_OK)}): " + ", ".join(sorted(VLAN_UPDATES_OK)))
    else:
        print("❔ No VLANs were updated.")

    if VLAN_UPDATES_SKIPPED:
        print(f"⚠️ Skipped (not present on MX) ({len(VLAN_UPDATES_SKIPPED)}): " +
              ", ".join(sorted(VLAN_UPDATES_SKIPPED)))

    if VLAN_UPDATES_FAILED:
        print(f"❌ Failed to update ({len(VLAN_UPDATES_FAILED)}): " +
              ", ".join(sorted(VLAN_UPDATES_FAILED)))
        print("   Check logs for detailed error traces.")

    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")

def classify_serials_for_binding(org_id: str, net_id: str, serials: List[str]):
    already, elsewhere, avail = [], [], []
    for s in serials:
        try:
            inv = get_inventory_device(org_id, s)
            nid = inv.get('networkId')
            if nid == net_id:
                already.append(s)
            elif nid:
                elsewhere.append((s, inv.get('networkName') or nid))
            else:
                avail.append(s)
        except requests.exceptions.HTTPError as e:
            if e.response is not None and e.response.status_code == 404:
                # Not in inventory yet
                avail.append(s)
            else:
                logging.error(f"Error checking inventory for {s}: {e}")
        except Exception as e:
            logging.error(f"Error checking inventory for {s}: {e}")
    return already, elsewhere, avail

# ---------- Clear & remove by model (org-aware) ----------
def _clear_and_remove_models(org_id: str, network_id: str, models: Tuple[str, ...]) -> bool:
    mx, ms, mr = fetch_devices(org_id, network_id)
    all_devs = mx + ms + mr
    to_remove = [d['serial'] for d in all_devs if d['model'] in models]
    if not to_remove:
        return True
    for serial in to_remove:
        try:
            do_action(meraki_put, f"/devices/{serial}", data={"name": "", "address": ""})
            log_change('device_clear', f"Cleared config for {serial}", device_serial=serial)
        except Exception:
            logging.exception(f"Error clearing {serial}")
    try:
        for serial in to_remove:
            do_action(meraki_post, f"/networks/{network_id}/devices/remove", data={"serial": serial})
            log_change('device_removed', f"Removed device from network", device_serial=serial)
    except Exception:
        logging.exception("Error removing devices")
    return True

def remove_existing_mx64_devices(org_id: str, network_id: str) -> bool:
    return _clear_and_remove_models(org_id, network_id, ("MX64",))

def remove_existing_mr33_devices(org_id: str, network_id: str) -> bool:
    return _clear_and_remove_models(org_id, network_id, ("MR33",))

# ---------- Prompt + claim into ORG (before selecting network) ----------
def prompt_and_validate_serials(org_id: str) -> List[str]:
    MAX_SERIAL_ATTEMPTS = 4
    MAX_BLANK_ATTEMPTS = 4
    serial_pattern = re.compile(r"[A-Z0-9]{4}-[A-Z0-9]{4}-[A-Z0-9]{4}")

    while True:
        count_raw = input("How many devices/serials will you add to this org? (Enter to skip): ").strip()
        if not count_raw:
            return []
        try:
            intended_count = int(count_raw)
            if intended_count <= 0:
                print("ℹ️  Count must be a positive integer.")
                continue
            break
        except ValueError:
            print("ℹ️  Please enter a whole number (e.g., 3).")

    blank_attempts = 0
    while True:
        print("\nEnter serial numbers:")
        print(" - You can paste them all at once (comma-separated),")
        print(" - OR enter one per line and press Enter on a blank line to finish.\n")
        first_line = input("Enter serial(s): ").strip().upper()

        raw_serials: List[str] = []
        if "," in first_line:
            raw_serials = [s.strip().upper() for s in first_line.split(",") if s.strip()]
        else:
            if first_line:
                raw_serials.append(first_line)
            while True:
                nxt = input("Enter next serial (or blank to finish): ").strip().upper()
                if not nxt:
                    break
                raw_serials.append(nxt)

        if not raw_serials:
            blank_attempts += 1
            remaining = MAX_BLANK_ATTEMPTS - blank_attempts
            if remaining <= 0:
                print("\n❌ No serial number(s) entered after 4 attempts -----------")
                print("   Please retry when serial(s) are known *******")
                sys.exit(1)
            print(f"ℹ️  No serials provided. Try again. (attempt {blank_attempts}/{MAX_BLANK_ATTEMPTS})")
            continue

        seen: Set[str] = set()
        serial_list: List[str] = []
        for s in raw_serials:
            if s in seen:
                print(f"ℹ️  Duplicate serial '{s}' removed from input.")
                continue
            seen.add(s)
            serial_list.append(s)

        entered_count = len(serial_list)
        if entered_count != intended_count:
            print(f"⚠️  You said {intended_count} device(s) but entered {entered_count}.")
            choice = input("Proceed anyway? (yes to proceed / no to re-enter): ").strip().lower()
            if choice not in {"y", "yes"}:
                blank_attempts = 0
                continue

        collected: List[str] = []
        for idx, original_serial in enumerate(serial_list, start=1):
            attempts = 0
            serial = original_serial
            while attempts < MAX_SERIAL_ATTEMPTS:
                if not serial_pattern.fullmatch(serial or ""):
                    attempts += 1
                    if attempts >= MAX_SERIAL_ATTEMPTS:
                        print(f"❌ Maximum attempts reached for serial #{idx} ({original_serial}). Skipping.")
                        break
                    serial = input(
                        f"Serial #{idx} '{serial}' is invalid. Re-enter (attempt {attempts+1}/{MAX_SERIAL_ATTEMPTS}): "
                    ).strip().upper()
                    continue

                try:
                    get_inventory_device(org_id, serial)
                    print(f"✅ {serial} found in org inventory.")
                    collected.append(serial)
                    break
                except requests.exceptions.HTTPError as e:
                    resp = e.response
                    status_code = resp.status_code if resp is not None else None
                    if status_code == 404:
                        try:
                            do_action(
                                meraki_post,
                                f"/organizations/{org_id}/inventory/claim",
                                data={"serials": [serial]},
                            )
                            print(f"✅ Serial '{serial}' successfully claimed into org inventory.")
                            log_change('device_claimed_inventory', "Claimed serial into org inventory", device_serial=serial)
                            collected.append(serial)
                            break
                        except Exception as claim_ex:
                            attempts += 1
                            print(f"❌ Error claiming '{serial}' into org inventory: {claim_ex}")
                            if attempts >= MAX_SERIAL_ATTEMPTS:
                                print(f"❌ Maximum attempts reached for serial #{idx}. Skipping.")
                                break
                            serial = input(
                                f"Re-enter serial #{idx} (attempt {attempts+1}/{MAX_SERIAL_ATTEMPTS}): "
                            ).strip().upper()
                            continue
                    else:
                        print(f"API Error for serial '{serial}': {e}")
                        break
                except Exception as e:
                    print(f"API Error for serial '{serial}': {e}")
                    break

        if len(collected) != intended_count:
            print(f"⚠️  Intended: {intended_count}, Entered: {entered_count}, Validated: {len(collected)}.")
            choice = input("Proceed with validated devices anyway? (yes to proceed / no to re-enter all): ").strip().lower()
            if choice in {"y", "yes"}:
                return collected
            else:
                blank_attempts = 0
                print("Okay, let's re-enter the serial list.")
                continue

        return collected

def summarize_devices_in_org(org_id: str, serials: List[str]) -> Set[str]:
    detected_mx_models: Set[str] = set()
    if not serials:
        print("No serials to summarize.")
        return detected_mx_models

    print("\nValidated / added to organization:")
    for s in serials:
        try:
            inv = get_inventory_device(org_id, s)
            model = inv.get('model') or 'Unknown'
            ptypes = inv.get('productTypes') or []
            ptype = ptypes[0] if isinstance(ptypes, list) and ptypes else inv.get('productType') or 'Unknown'
            name = inv.get('name') or ''
            print(f" - {s}: {model} ({ptype}){f' — {name}' if name else ''}")

            if model.startswith('MX67'):
                detected_mx_models.add('MX67')
            elif model.startswith('MX75'):
                detected_mx_models.add('MX75')
        except Exception as e:
            print(f" - {s}: (lookup failed: {e})")

    return detected_mx_models

# ---------- Claim into network using prevalidated serials ----------
def claim_devices(org_id: str, network_id: str, prevalidated_serials: Optional[List[str]] = None) -> List[str]:
    if prevalidated_serials is not None:
        valids = prevalidated_serials
    else:
        valids = prompt_and_validate_serials(org_id)

    if not valids:
        print("❌ No valid serials.")
        return []

    already, elsewhere, avail = classify_serials_for_binding(org_id, network_id, valids)
    if elsewhere:
        print("⚠️ In use elsewhere:")
        for s, name in elsewhere:
            print(f" - {s} in {name}")

    mx_models: List[str] = []
    for s in avail:
        try:
            inv = get_inventory_device(org_id, s)
            if (inv.get('model') or '').startswith('MX'):
                mx_models.append(inv['model'])
        except Exception:
            pass
    if len(set(mx_models)) > 1:
        print("❌ MX warm spare models mismatch. Aborting.")
        return []
    if not avail:
        print("ℹ️ No newly available devices to claim to the network (perhaps they’re already in this network).")
        return already

    try:
        remove_existing_mx64_devices(org_id, network_id)
        do_action(meraki_post, f"/networks/{network_id}/devices/claim", data={"serials": avail})
        for s in avail:
            log_change('device_claimed', f"Claimed device to network", device_serial=s)
        return avail
    except Exception:
        logging.exception("Failed to claim/bind")
        return []

# ---------- ORDERING HELPERS ----------
def select_primary_mx(org_id: str, serials: List[str]) -> Optional[str]:
    mx_candidates: List[Tuple[str, str]] = []
    for s in serials:
        try:
            inv = get_inventory_device(org_id, s)
            model = (inv.get('model') or '').upper()
            if model.startswith('MX'):
                mx_candidates.append((s, model))
        except Exception:
            logging.exception(f"Unable to read inventory for {s}")

    if len(mx_candidates) == 0:
        return None
    if len(mx_candidates) == 1:
        return mx_candidates[0][0]

    auto_choice = sorted([s for s, _ in mx_candidates])[0]

    print("\nMultiple MX devices detected in the claimed list:")
    for idx, (s, m) in enumerate(mx_candidates, 1):
        print(f" {idx}. {s}  ({m})")

    while True:
        sel = input(
        "Select which MX should be PRIMARY (mx-01). "
                "Enter number, or press Enter / type 'skip'/'cancel' to auto-select: "
        ).strip().lower()

        # Handle skip/cancel/empty
        if not sel or sel in {'skip', 'cancel'}:
            print(f"ℹ️  No explicit choice made. Auto-selecting PRIMARY MX: {auto_choice}")
            return auto_choice

        # Validate numeric input
        if sel.isdigit():
            i = int(sel)
            if 1 <= i <= len(mx_candidates):
                return mx_candidates[i - 1][0]
            else:
                print(f"❌ Invalid number. Please choose between 1 and {len(mx_candidates)}.")
                continue

        # Any other input (e.g., "2,1", "one", etc.)
        print("❌ Invalid input. Please enter a single number corresponding to an MX device, or press Enter to skip.")

def select_device_order(org_id: str, serials: List[str], kind: str) -> List[str]:
    """
    Ask for an order of devices of a given kind (MR/MS).
    - Accepts partial input: e.g., '2,1' when there are 3 items.
    - Orders the selected indices first (in the specified order),
      then appends any remaining devices in their original order.
    - Enter/skip/cancel => auto-order by serial.
    """
    filtered: List[Tuple[str, str]] = []
    for s in serials:
        try:
            inv = get_inventory_device(org_id, s)
            model = (inv.get('model') or '').upper()
            if kind == 'MR' and _is_wireless_model(model):
                filtered.append((s, model))
            elif kind == 'MS' and model.startswith('MS'):
                filtered.append((s, model))
        except Exception:
            logging.exception(f"Unable to read inventory for {s}")

    # If 0 or 1 device, nothing to order
    if len(filtered) <= 1:
        return [s for s, _ in filtered]

    # Default auto order
    auto_order = [s for s, _ in filtered]  # keep original order shown to user

    # Show menu
    print(f"\nSelect ordering for {kind} devices (enter a comma-separated list of indices).")
    for idx, (s, m) in enumerate(filtered, 1):
        print(f" {idx}. {s}  ({m})")

    raw = input(
        f"Desired order for {kind} (e.g. 2,1,3). "
        "Press Enter / type 'skip'/'cancel' to auto-order: "
    ).strip().lower()

    # Auto-order on skip
    if not raw or raw in {'skip', 'cancel'}:
        print(f"ℹ️  Auto-ordering {kind} devices (original order): {', '.join(auto_order)}")
        return auto_order

    # Parse indices; accept partial sets
    parts = [p.strip() for p in raw.split(',') if p.strip()]
    valid_indices: List[int] = []
    seen: set[int] = set()
    for p in parts:
        if not p.isdigit():
            print(f"❌ '{p}' is not a number. Please enter indices like 2,1,3.")
            # re-prompt
            return select_device_order(org_id, serials, kind)
        i = int(p)
        if not (1 <= i <= len(filtered)):
            print(f"❌ {i} is out of range. Valid range is 1..{len(filtered)}.")
            return select_device_order(org_id, serials, kind)
        if i in seen:
            print(f"❌ Duplicate index {i} detected.")
            return select_device_order(org_id, serials, kind)
        seen.add(i)
        valid_indices.append(i)

    # Build final order: chosen first (in specified order), then remaining in original order
    chosen_serials = [filtered[i - 1][0] for i in valid_indices]
    remaining_serials = [s for s, _ in filtered if s not in chosen_serials]
    final_order = chosen_serials + remaining_serials

    print(f"✅ Using {kind} order: {', '.join(final_order)}")
    return final_order

# ---------- Warm spare primary enforcement ----------
def ensure_primary_mx(network_id: str, desired_primary_serial: Optional[str]) -> None:
    if not desired_primary_serial:
        return

    try:
        status = meraki_get(f"/networks/{network_id}/appliance/warmSpare") or {}
        enabled = bool(status.get("enabled"))
        current_primary = status.get("primarySerial")

        if not enabled:
            print("ℹ️  Warm spare is not enabled on this network; cannot swap primary automatically.")
            log_change('mx_warmspare_not_enabled',
                       "Warm spare not enabled; no primary swap performed",
                       network_id=network_id)
            return

        if current_primary and current_primary.upper() == desired_primary_serial.upper():
            print(f"✅ Warm spare already has the correct primary ({desired_primary_serial}).")
            return

        print(f"🔁 Swapping warm spare primary to {desired_primary_serial} ...")
        out = retry_with_backoff(
            "warmspare_swap",
            lambda: do_action(meraki_post, f"/networks/{network_id}/appliance/warmSpare/swap"),
            max_attempts=4,
            base_sleep=2.0,
            max_sleep=15.0,
            verbose=True,
        )
        if not out.ok:
            raise RuntimeError(f"Warm spare swap failed after {out.attempts} attempts: {out.error}")

        log_change('mx_warmspare_swap',
                   f"Swapped warm spare primary to {desired_primary_serial}",
                   device_serial=desired_primary_serial,
                   network_id=network_id)
        print("✅ Warm spare primary swap requested.")

    except Exception as e:
        logging.exception("Failed to ensure warm spare primary")
        print(f"❌ Failed to verify/swap warm spare primary: {e}")

# ---------- Naming & configuration (with ordering) ----------
def name_and_configure_claimed_devices(
    org_id: str,
    network_id: str,
    network_name: str,
    serials: List[str],
    ms_list: List[Dict[str, Any]],
    tpl_profile_map: Dict[str, str],
    old_mx_devices: Optional[List[Dict[str, Any]]] = None,
    old_mr_devices: Optional[List[Dict[str, Any]]] = None,
    primary_mx_serial: Optional[str] = None,
    mr_order: Optional[List[str]] = None,
    ms_order: Optional[List[str]] = None,
) -> None:
    """
    Renames and configures newly-claimed devices using optional ordering.

    Enhancements:
      - Per-device retry counters (central retry helper)
      - Summary of failed devices (console + JSON)
    """
    prefix = "-".join(network_name.split("-")[:2]).lower()
    counts: Dict[str, int] = {"MX": 1, "MR": 1, "MS": 1}

    old_mr33s = sorted(
        [d for d in (old_mr_devices or []) if d.get("model") == "MR33"],
        key=lambda x: str(x.get("name", "")),
    )
    old_mxs_sorted = sorted(
        (old_mx_devices or []),
        key=lambda x: str(x.get("name", "")),
    )

    # ---------------- Inventory cache ----------------
    inv_by_serial: Dict[str, Dict[str, Any]] = {}
    for s in serials:
        try:
            inv_by_serial[s] = get_inventory_device(org_id, s) or {}
        except Exception:
            logging.exception("Failed inventory lookup for %s", s)
            inv_by_serial[s] = {}

    # ---------------- Classify devices ----------------
    mx_serials = [s for s in serials if (inv_by_serial.get(s, {}).get("model") or "").upper().startswith("MX")]
    mr_serials = [s for s in serials if _is_wireless_model((inv_by_serial.get(s, {}).get("model") or "").upper())]
    ms_serials = [s for s in serials if (inv_by_serial.get(s, {}).get("model") or "").upper().startswith("MS")]

    # ---------------- Ordering ----------------
    if primary_mx_serial and primary_mx_serial in mx_serials:
        mx_serials = [primary_mx_serial] + [s for s in mx_serials if s != primary_mx_serial]
    if mr_order:
        mr_serials = [s for s in mr_order if s in mr_serials]
    if ms_order:
        ms_serials = [s for s in ms_order if s in ms_serials]

    # ---------------- Profile selection ----------------
    default_profile_name: Optional[str] = ms_list[0].get("switchProfileName") if ms_list else None
    default_profile_id: Optional[str] = tpl_profile_map.get(default_profile_name) if default_profile_name else None

    # ---------------- Tracking / summaries ----------------
    per_device_attempts: Dict[str, int] = {}
    failed_devices: Dict[str, Dict[str, Any]] = {}  # serial -> {model, role, error...}

    def _record_failure(serial: str, *, role: str, model: str, outcome: RetryOutcome, payload: Dict[str, Any]) -> None:
        failed_devices[serial] = {
            "serial": serial,
            "role": role,
            "model": model,
            "attempts": outcome.attempts,
            "error_type": outcome.error_type,
            "error": outcome.error,
            "http_status": outcome.http_status,
            "last_response_text": outcome.last_response_text,
            "payload": payload,
        }

    def _configure_one(serial: str, payload: Dict[str, Any], *, role: str) -> None:
        model = str(inv_by_serial.get(serial, {}).get("model") or "")
        op_name = f"configure_{role}_{serial}"

        def _op() -> None:
            # Use your eventual-consistency aware function; it may raise.
            put_device_with_transient_404_retry(org_id, serial, payload)

        outcome = retry_with_backoff(op_name, _op, max_attempts=5, base_sleep=1.0, max_sleep=10.0)
        per_device_attempts[serial] = outcome.attempts

        if outcome.ok:
            log_change(
                "device_update",
                f"Renamed and reconfigured device {serial} ({model})",
                device_serial=serial,
                device_name=str(payload.get("name") or ""),
                misc=f"attempts={outcome.attempts} payload={json.dumps(payload)}",
                network_id=network_id,
                network_name=network_name,
            )
        else:
            logging.error("Failed configuring %s (%s) role=%s after %d attempts: %s",
                          serial, model, role, outcome.attempts, outcome.error)
            _record_failure(serial, role=role, model=model, outcome=outcome, payload=payload)

    # ==================================================
    # MX
    # ==================================================
    mx_idx = 0
    for s in mx_serials:
        payload: Dict[str, Any] = {"name": f"{prefix}-mx-{counts['MX']:02}"}

        if mx_idx < len(old_mxs_sorted):
            payload["address"] = old_mxs_sorted[mx_idx].get("address", "") or ""
            payload["tags"] = old_mxs_sorted[mx_idx].get("tags", []) or []
        else:
            payload["address"] = ""
            payload["tags"] = []

        mx_idx += 1
        counts["MX"] += 1
        _configure_one(s, payload, role="MX")

    # ==================================================
    # MR
    # ==================================================
    ap_idx = 0
    for s in mr_serials:
        payload = {"name": f"{prefix}-ap-{counts['MR']:02}"}

        if ap_idx < len(old_mr33s):
            payload["tags"] = old_mr33s[ap_idx].get("tags", []) or []
            payload["address"] = old_mr33s[ap_idx].get("address", "") or ""
        else:
            payload["tags"] = []
            payload["address"] = ""

        ap_idx += 1
        counts["MR"] += 1
        _configure_one(s, payload, role="MR")

    # ==================================================
    # MS
    # ==================================================
    for s in ms_serials:
        payload: Dict[str, Any] = {"name": f"{prefix}-sw-{counts['MS']:02}"}
        counts["MS"] += 1

        if default_profile_id:
            payload["switchProfileId"] = default_profile_id

        _configure_one(s, payload, role="MS")

    # ---------------- Final summary ----------------
    summary = {
        "timestamp": datetime.now().isoformat(),
        "org_id": org_id,
        "network_id": network_id,
        "network_name": network_name,
        "total_devices": len(serials),
        "attempts_by_serial": per_device_attempts,
        "failed_devices": failed_devices,
    }

    try:
        path = os.path.join("logs", f"device_config_summary_{org_id}_{network_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        _write_json_log(path, summary)
        logging.info("Wrote device config summary JSON: %s", path)
        if DEBUG_MODE:
            print(f"📝 Device config summary saved: {path}")
    except Exception:
        logging.exception("Failed writing device config summary JSON")

    if failed_devices:
        print("\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        print("❌ Device Configuration Failures")
        print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        for serial, info in failed_devices.items():
            role = info.get("role")
            model = info.get("model")
            attempts = info.get("attempts")
            status = info.get("http_status")
            err = info.get("error")
            print(f"- {serial} ({model}) role={role} attempts={attempts} http={status} err={err}")
        print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")

def enable_mx_wan2(serial: str) -> bool:
    """
    Enables WAN2 for the given MX by updating /devices/{serial}/appliance/uplinks/settings.
    Preserves existing settings by GET->merge->PUT. Falls back to minimal payload if needed.
    """
    path = f"/devices/{serial}/appliance/uplinks/settings"
    existing: Dict[str, Any] | None = None
    try:
        existing = meraki_get(path)
    except requests.exceptions.HTTPError as e:
        resp = e.response
        if resp is not None and resp.status_code not in (400, 404):
            logging.debug(
                "GET uplink settings for %s returned %s, proceeding with minimal payload",
                serial, resp.status_code
            )
    except Exception:
        logging.exception("Error reading uplink settings for %s; proceeding with minimal payload", serial)

    if isinstance(existing, dict):
        merged = dict(existing)
        wan2 = dict(merged.get("wan2", {}))
        wan2["enabled"] = True
        merged["wan2"] = wan2
        payload: Dict[str, Any] = merged
    else:
        payload = {"wan2": {"enabled": True}}

    try:
        out = retry_with_backoff(
            f"enable_wan2_{serial}",
            lambda: do_action(meraki_put, path, data=payload),
            max_attempts=4,
            base_sleep=1.0,
            max_sleep=10.0,
            verbose=False,
        )
        if not out.ok:
            raise requests.exceptions.HTTPError(out.error)

        log_change(
            "mx_wan2_enable",
            "Enabled WAN2 on MX",
            device_serial=serial,
            misc=json.dumps({"payload": payload})
        )
        logging.info("Enabled WAN2 for %s", serial)
        return True
    except requests.exceptions.HTTPError as e:
        try:
            do_action(meraki_put, path, data={"wan2": {"enabled": True}})
            log_change(
                "mx_wan2_enable",
                "Enabled WAN2 on MX (fallback payload)",
                device_serial=serial,
                misc='{"wan2":{"enabled":true}}'
            )
            logging.info("Enabled WAN2 (fallback) for %s", serial)
            return True
        except Exception:
            resp = e.response
            status = resp.status_code if resp is not None else "?"
            logging.error("Failed enabling WAN2 for %s: %s %s", serial, status, str(e))
            return False
    except Exception:
        logging.exception("Unexpected error enabling WAN2 for %s", serial)
        return False

def enable_wan2_on_claimed_mx(org_id: str, claimed_serials: List[str]) -> None:
    """
    Enable WAN2 only on newly claimed MX67 devices.
    """
    for s in claimed_serials:
        try:
            inv = get_inventory_device(org_id, s)
            model = (inv.get("model") or "").upper()
            if model.startswith("MX67"):
                ok = enable_mx_wan2(s)
                if not ok:
                    logging.warning("WAN2 not enabled for %s (model %s)", s, model)
            else:
                logging.info("Skipping WAN2 enable for %s (model %s is not MX67)", s, model or "unknown")
        except requests.exceptions.HTTPError as e:
            resp = e.response
            status = resp.status_code if resp is not None else "?"
            logging.exception("Inventory check failed for %s: %s %s", s, status, str(e))
        except Exception:
            logging.exception("Could not evaluate/enable WAN2 for %s", s)

def safe_enable_wan2_on_claimed_mx(org_id: str, claimed: List[str]) -> None:
    try:
        if claimed:
            enable_wan2_on_claimed_mx(org_id, claimed)
    except Exception:
        logging.exception("Failed enabling WAN2 on claimed MX devices")

def remove_recently_added_tag(network_id: str):
    devs_raw = meraki_get(f"/networks/{network_id}/devices")
    devs: List[Dict[str, Any]] = devs_raw if isinstance(devs_raw, list) else []
    for d in devs:
        tags = d.get('tags', [])
        if not isinstance(tags, list):
            tags = (tags or '').split()
        if 'recently-added' in tags:
            updated_tags = [t for t in tags if t != 'recently-added']
            print(f"Removing 'recently-added' tag from {d['model']} {d['serial']}")
            try:
                do_action(meraki_put, f"/devices/{d['serial']}", data={"tags": updated_tags})
                log_change(
                    'tag_removed',
                    "Removed 'recently-added' tag",
                    device_serial=d['serial'],
                    device_name=d.get('name', ''),
                    misc=f"old_tags={tags}, new_tags={updated_tags}"
                )
            except Exception:
                logging.exception(f"Failed to remove 'recently-added' from {d['serial']}")

def _pick_template_by_vlan_count(
    templates: List[Dict[str, Any]],
    vlan_count: Optional[int],
    current_template_name: Optional[str] = None,
) -> Optional[Dict[str, Any]]:
    """
    Suggest a template based on VLAN count:
      - 3 VLANs  -> match name like 'NO LEGACY ... MX' or '... MX67/MX75'
      - 5 VLANs  -> match name like '3 X DATA VLAN ... MX75'
      - 4 VLANs  -> if current_template_name looks like a CloudStore variant,
                    try to find a corresponding ...-MX67 / ...-MX75 template
      - other    -> raise ValueError so the caller can warn the user
    """
    if vlan_count is None:
        return None

    # ---------- 3 VLANs ----------
    if vlan_count == 3:
        patterns = [r'NO\s*LEGACY.*MX(?:\d{2})?\b']

    # ---------- 5 VLANs ----------
    elif vlan_count == 5:
        patterns = [r'3\s*X\s*DATA[_\s-]*VLAN.*MX75\b']

    # ---------- 4 VLANs: CloudStore name-based selection ----------
    elif vlan_count == 4:
        if not current_template_name:
            return None

        name = current_template_name.strip()

        # Match either:
        #   GBR-CT-CloudStore-001
        #   GBR-CT-CloudStore-PreSCE-001
        m = re.match(r'^(GBR-CT-CloudStore(?:-PreSCE)?-\d{3})$', name, re.IGNORECASE)
        if not m:
            # Name doesn't match the CloudStore pattern → no suggestion
            return None

        base = m.group(1)  # e.g. GBR-CT-CloudStore-001 or GBR-CT-CloudStore-PreSCE-001
        wanted_prefixes = [
            f"{base}-MX67",
            f"{base}-MX75",
        ]

        for t in templates:
            tname = (t.get("name") or "").strip()
            upper_name = tname.upper()
            if any(upper_name.startswith(p.upper()) for p in wanted_prefixes):
                return t

        # No match found → no suggestion
        return None

    else:
        raise ValueError(
            "Incorrect number of VLANs detected in the current network. "
            "Please double check the selected network."
        )

    # ---------- Default regex-based matching for 3 / 5 VLANs ----------
    for t in templates:
        name = (t.get('name') or '')
        if any(re.search(pat, name, re.IGNORECASE) for pat in patterns):
            return t
    return None

def _get_template_name(org_id: str, template_id: Optional[str]) -> Optional[str]:
    """
    Safely resolve a config template ID to its name.
    Returns None if template_id is empty or lookup fails.
    """
    if not template_id:
        return None
    try:
        tpl = meraki_get(f"/organizations/{org_id}/configTemplates/{template_id}") or {}
        return str(tpl.get("name") or "")
    except Exception:
        logging.exception("Could not fetch template name for %s", template_id)
        return None

# CloudStore / PreSCE template patterns
CLOUDSTORE_BASE_RE = re.compile(r'^GBR-CT-CloudStore-(\d{3})$', re.IGNORECASE)
PRESCE_MX75_RE = re.compile(r'^GBR-CT-CloudStore-PreSCE-(\d{3})-MX75$', re.IGNORECASE)

def _cloudstore_store_code_from_name(name: Optional[str]) -> Optional[str]:
    """
    Returns the 3-digit store code (*** in GBR-CT-CloudStore-***) or None.
    """
    if not name:
        return None
    m = CLOUDSTORE_BASE_RE.match(name.strip())
    return m.group(1) if m else None

def _readdress_ip_preserve_last_octet(old_ip: str, new_subnet: str) -> Optional[str]:
    """
    Take an existing IPv4 address (old_ip) and a new subnet string (e.g. '10.41.41.0/24'),
    and return a new IP in the *new* subnet but with the *same last octet* as old_ip.
    If parsing fails, returns None so the caller can decide what to do.
    """
    try:
        net = ip_network(str(new_subnet), strict=False)
        if net.version != 4:
            return None
        new_net_octets = str(net.network_address).split(".")
        if len(new_net_octets) != 4:
            return None

        old_octets = str(old_ip).split(".")
        if len(old_octets) != 4:
            return None

        last_octet = old_octets[-1]
        return ".".join(new_net_octets[:3] + [last_octet])
    except Exception:
        logging.exception("Failed to readdress IP %s into subnet %s", old_ip, new_subnet)
        return None

def _readdress_fixed_assignments_preserve_last_octet(
    fixed: Optional[Dict[str, Any]],
    new_subnet: Optional[str],
) -> Optional[Dict[str, Any]]:
    """
    Given a fixedIpAssignments dict and a new subnet, re-IP each reservation
    into the new subnet, preserving the last octet of the IP.
    Falls back to the original dict if anything critical fails.
    """
    if not fixed or not isinstance(fixed, dict) or not new_subnet:
        return fixed

    out: Dict[str, Any] = {}
    for key, val in fixed.items():
        if not isinstance(val, dict):
            out[key] = val
            continue
        old_ip = val.get("ip")
        if not old_ip:
            out[key] = val
            continue
        new_ip = _readdress_ip_preserve_last_octet(str(old_ip), str(new_subnet))
        if not new_ip:
            out[key] = val
            continue
        new_val = dict(val)
        new_val["ip"] = new_ip
        out[key] = new_val
    return out

def _readdress_reserved_ranges_preserve_last_octet(
    ranges: Optional[List[Dict[str, Any]]],
    new_subnet: Optional[str],
) -> Optional[List[Dict[str, Any]]]:
    """
    Optional: re-IP reservedIpRanges so that start/end move into the new subnet,
    preserving the last octet of each end-point.
    """
    if not ranges or not isinstance(ranges, list) or not new_subnet:
        return ranges

    out: List[Dict[str, Any]] = []
    for r in ranges:
        if not isinstance(r, dict):
            out.append(r)
            continue
        start = r.get("start")
        end = r.get("end")
        if not start and not end:
            out.append(r)
            continue

        new_r = dict(r)
        if start:
            new_start = _readdress_ip_preserve_last_octet(str(start), str(new_subnet))
            if new_start:
                new_r["start"] = new_start
        if end:
            new_end = _readdress_ip_preserve_last_octet(str(end), str(new_subnet))
            if new_end:
                new_r["end"] = new_end
        out.append(new_r)
    return out

def maybe_swap_vlan_40_41(vlans: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Return a new VLAN list where VLAN 40 and VLAN 41 have their *config*
    swapped, and any DHCP reservations are re-addressed into the new subnet
    (preserving the last octet), while ensuring that:

      - If VLAN 40 had DHCP enabled (server/relay), it still does after swap.
      - If VLAN 41 had DHCP enabled, it still does after swap.

    DHCP "enabledness" is tied to the VLAN ID (40 vs 41), not to the subnet.
    """
    if not isinstance(vlans, list):
        return vlans

    vlan_40 = None
    vlan_41 = None
    others: List[Dict[str, Any]] = []

    for v in vlans:
        try:
            vid_raw = v.get("id") or v.get("vlan") or v.get("vlanId")
            vid = int(str(vid_raw))
        except Exception:
            others.append(v)
            continue

        if vid == 40:
            vlan_40 = v
        elif vid == 41:
            vlan_41 = v
        else:
            others.append(v)

    if not (vlan_40 and vlan_41):
        # Nothing to do if one VLAN is missing
        return vlans

    # Base configs swapped (subnet, applianceIp, etc.)
    base_40 = dict(vlan_41)  # VLAN ID 40 will now look like old 41's config
    base_41 = dict(vlan_40)  # VLAN ID 41 will now look like old 40's config

    # Ensure IDs stay correct
    base_40["id"] = 40
    base_41["id"] = 41

    # Old reservations attached to each VLAN ID
    old_40_fixed = vlan_40.get("fixedIpAssignments")
    old_41_fixed = vlan_41.get("fixedIpAssignments")
    old_40_ranges = vlan_40.get("reservedIpRanges")
    old_41_ranges = vlan_41.get("reservedIpRanges")

    # Target subnets after the swap (the ones each VLAN will now use)
    target_subnet_for_40 = base_40.get("subnet")
    target_subnet_for_41 = base_41.get("subnet")

    # Re-IP reservations into the *new* subnet, preserving last octet
    new_fixed_for_40 = _readdress_fixed_assignments_preserve_last_octet(
        old_40_fixed, target_subnet_for_40
    )
    new_fixed_for_41 = _readdress_fixed_assignments_preserve_last_octet(
        old_41_fixed, target_subnet_for_41
    )

    new_ranges_for_40 = _readdress_reserved_ranges_preserve_last_octet(
        old_40_ranges, target_subnet_for_40
    )
    new_ranges_for_41 = _readdress_reserved_ranges_preserve_last_octet(
        old_41_ranges, target_subnet_for_41
    )

    # IMPORTANT: Preserve DHCP "enabledness" per VLAN ID.
    # VLAN 40 keeps its original DHCP mode, but on the new subnet & reservations.
    _preserve_dhcp_state_for_vlan(
        original_vlan=vlan_40,
        new_vlan=base_40,
        new_fixed=new_fixed_for_40,
        new_ranges=new_ranges_for_40,
    )

    # Same for VLAN 41.
    _preserve_dhcp_state_for_vlan(
        original_vlan=vlan_41,
        new_vlan=base_41,
        new_fixed=new_fixed_for_41,
        new_ranges=new_ranges_for_41,
    )

    logging.info(
        "Swapped VLAN 40/41 configs, re-IP'd DHCP reservations (preserving last octet), "
        "and preserved DHCP enabled state per VLAN."
    )
    print(
        "🔁 Swapping VLAN 40/41, re-addressing DHCP reservations (preserving last octet), "
        "and preserving DHCP enabled state."
    )

    new_list = others + [base_40, base_41]
    # Keep output in a stable order by VLAN ID if possible
    try:
        new_list.sort(key=lambda v: int(str(v.get("id") or 0)))
    except Exception:
        pass

    return new_list

def list_and_rebind_template(
    org_id: str,
    network_id: str,
    current_id: Optional[str],
    network_name: str,
    *,
    pre_change_devices: Optional[List[Dict[str, Any]]] = None,
    pre_change_vlans: Optional[List[Dict[str, Any]]] = None,
    pre_change_template: Optional[str] = None,
    claimed_serials: Optional[List[str]] = None,
    removed_serials: Optional[List[str]] = None,
    ms_list: Optional[List[Dict[str, Any]]] = None,
    mx_model_filter: Optional[str] = None,
) -> Tuple[Optional[str], Optional[str], bool, bool]:
    """
    Returns:
      (new_template_id, new_template_name, rollback_triggered, cloudstore_presce_flow)
    """
    skip_attempts = 0
    cloudstore_presce_flow = False

    # ------------------------------------------------------------------
    # Fetch templates and bound counts
    # ------------------------------------------------------------------
    all_templates_raw = meraki_get(f"/organizations/{org_id}/configTemplates")
    all_templates: List[Dict[str, Any]] = (
        all_templates_raw if isinstance(all_templates_raw, list) else []
    )

    all_counts = _template_counts_for_org(org_id)

    eligible: List[Dict[str, Any]] = []
    unknown: List[Dict[str, Any]] = []

    for t in all_templates:
        tid = t.get("id")
        if not tid:
            continue

        cnt = all_counts.get(tid)
        t2 = dict(t)
        t2["_boundCount"] = cnt

        if cnt is None:
            unknown.append(t2)
        elif cnt < 90:
            eligible.append(t2)

    if not eligible and not unknown:
        print("ℹ️ No templates available.")
        return current_id, None, False, False

    filtered = eligible + unknown if eligible else unknown[:]

    # Optional MX model suffix filter
    if mx_model_filter in {"MX67", "MX75"}:
        suffix = mx_model_filter.upper()
        subset = [
            t for t in filtered
            if (t.get("name") or "").strip().upper().endswith(suffix)
        ]
        if subset:
            filtered = subset
        else:
            print(f"(No templates ending with {suffix}; showing all eligible templates.)")

    # ------------------------------------------------------------------
    # VLAN count + current template name
    # ------------------------------------------------------------------
    vlan_count = _current_vlan_count(network_id)
    current_tpl_name = _get_template_name(org_id, current_id) if current_id else None
    curr_up = (current_tpl_name or "").strip().upper()

    # ------------------------------------------------------------------
    # SINGLE flatten decision point (authoritative)
    # ------------------------------------------------------------------
    flatten_requested = False

    if curr_up.startswith("GBR-CT-CLOUDSTORE-ENABLEMX-PORT3"):
        print("ℹ️ Flatten option disabled: current template is GBR-CT-CloudStore-EnableMX-Port3.")
    elif current_tpl_name and re.match(r"^GBR-CT-CloudStore-\d{3}$", current_tpl_name, re.IGNORECASE):
        flatten_requested = _prompt_yes_no(
            "Flatten VLANs / standardise the network? "
            "(Only swaps VLAN 40/41 if you answer YES)",
            default_no=True,
        )

    # Persist operator intent for later VLAN logic
    if _current_checkpoint is not None:
        _current_checkpoint.step_status = _current_checkpoint.step_status or {}
        _current_checkpoint.step_status["flatten_requested"] = bool(flatten_requested)
        _current_checkpoint.save()

    # ------------------------------------------------------------------
    # VLAN-count based filtering
    # ------------------------------------------------------------------
    if vlan_count == 4:
        rx_no_legacy = re.compile(r"NO\s*LEGACY.*MX", re.IGNORECASE)
        rx_3xdata = re.compile(r"3\s*X\s*DATA.*MX75", re.IGNORECASE)
        filtered = [
            t for t in filtered
            if not rx_no_legacy.search(t.get("name") or "")
            and not rx_3xdata.search(t.get("name") or "")
        ]
        if not filtered:
            print("⚠️ No templates remain after 4-VLAN filtering.")
            return current_id, None, False, False

    if vlan_count in (3, 5) and current_tpl_name and not flatten_requested:
        up = curr_up
        if "PRESCE" in up:
            filtered = [
                t for t in filtered
                if "PRESCE" in (t.get("name") or "").upper()
            ]
        else:
            filtered = [
                t for t in filtered
                if "CLOUDSTORE" in (t.get("name") or "").upper()
                and "PRESCE" not in (t.get("name") or "").upper()
            ]

    # ------------------------------------------------------------------
    # Flatten restriction: ONLY PreSCE-***-MX75
    # ------------------------------------------------------------------
    if flatten_requested:
        presce = [
            t for t in filtered
            if PRESCE_MX75_RE.match((t.get("name") or "").strip())
        ]
        if presce:
            filtered = presce
        else:
            print("⚠️ No PreSCE-***-MX75 templates found; flatten disabled.")
            flatten_requested = False
            if _current_checkpoint is not None:
                _current_checkpoint.step_status = _current_checkpoint.step_status or {}
                _current_checkpoint.step_status["flatten_requested"] = False
                _current_checkpoint.save()


    # ------------------------------------------------------------------
    # Suggestion logic
    # ------------------------------------------------------------------
    try:
        suggested_tpl = _pick_template_by_vlan_count(
            filtered,
            vlan_count,
            current_template_name=current_tpl_name,
        )
    except ValueError as e:
        print(f"❌ {e}")
        suggested_tpl = None

    suggested_id = suggested_tpl.get("id") if suggested_tpl else None
    if suggested_id:
        filtered = (
            [t for t in filtered if t.get("id") == suggested_id]
            + [t for t in filtered if t.get("id") != suggested_id]
        )

    # ------------------------------------------------------------------
    # Selection loop
    # ------------------------------------------------------------------
    while True:
        print(f"\nCurrent network: {network_name} (ID: {network_id})")

        if current_id:
            print(f"Bound template: {current_tpl_name} (ID: {current_id})\n")
        else:
            print("No template bound.\n")

        print("Available templates (< 90 bound or unknown):")
        for i, t in enumerate(filtered, 1):
            cnt = t.get("_boundCount")
            cnt_str = "?" if cnt is None else str(cnt)
            auto = " [AUTO]" if suggested_id == t.get("id") else ""
            print(f"{i}. {t.get('name','')}{auto} — {cnt_str} bound")

        sel = input("Select template # (or 'a' for suggestion): ").strip().lower()

        if sel in {"", "skip", "cancel"}:
            skip_attempts += 1
            if skip_attempts == 1:
                print("⚠️ Cancel again to rollback.")
                continue

            rollback_all_changes(
                network_id=network_id,
                pre_change_devices=pre_change_devices or [],
                pre_change_vlans=pre_change_vlans or [],
                pre_change_template=pre_change_template,
                org_id=org_id,
                claimed_serials=claimed_serials or [],
                removed_serials=removed_serials or [],
                ms_list=ms_list or [],
                network_name=network_name,
            )
            return current_id, None, True, False

        if sel == "a" and suggested_tpl:
            chosen = suggested_tpl
        elif sel.isdigit() and 1 <= int(sel) <= len(filtered):
            chosen = filtered[int(sel) - 1]
        else:
            print("Invalid selection.")
            continue

        if chosen["id"] == current_id:
            print("No change.")
            return current_id, chosen.get("name"), False, False

        chosen_name = (chosen.get("name") or "").strip()
        if flatten_requested and PRESCE_MX75_RE.match(chosen_name):
            cloudstore_presce_flow = True

        try:
            if current_id:
                do_action(meraki_post, f"/networks/{network_id}/unbind")
            do_action(
                meraki_post,
                f"/networks/{network_id}/bind",
                data={"configTemplateId": chosen["id"]},
            )

            _ORG_NETWORKS_CACHE.pop(org_id, None)
            _TEMPLATE_COUNT_CACHE.pop(org_id, None)

            print(f"✅ Bound to {chosen_name}")
            return chosen["id"], chosen_name, False, cloudstore_presce_flow

        except Exception as e:
            logging.exception("Template bind failed")
            rollback_all_changes(
                network_id=network_id,
                pre_change_devices=pre_change_devices or [],
                pre_change_vlans=pre_change_vlans or [],
                pre_change_template=pre_change_template,
                org_id=org_id,
                claimed_serials=claimed_serials or [],
                removed_serials=removed_serials or [],
                ms_list=ms_list or [],
                network_name=network_name,
            )
            return current_id, None, True, False

def _current_vlan_count(network_id: str) -> Optional[int]:
    vlans = fetch_vlan_details(network_id)
    return len(vlans) if isinstance(vlans, list) else None

# ---------- Template rebind helpers (with rollback) ----------
def bind_network_to_template(
    org_id: str,
    network_id: str,
    tpl_id: Optional[str],
    vlan_list: List[Dict[str, Any]],
    network_name: str,
    *,
    pre_change_devices,
    pre_change_vlans,
    pre_change_template,
    claimed_serials,
    removed_serials,
    ms_list,
    # NEW: explicit operator intent. Default False => no swap unless explicitly requested.
    flatten_requested: bool = False,
):
    if not tpl_id:
        return

    time.sleep(5)

    # -------- Resolve old/new template names (best-effort) --------
    try:
        new_tpl_name = _get_template_name(org_id, tpl_id)
    except Exception:
        logging.exception("Could not resolve new template name in bind_network_to_template")
        new_tpl_name = None

    try:
        old_tpl_name = _get_template_name(org_id, pre_change_template) if pre_change_template else None
    except Exception:
        logging.exception("Could not resolve old template name in bind_network_to_template")
        old_tpl_name = None

    # -------- Decide whether to swap VLAN 40/41 --------
    # Default: NEVER swap unless explicitly requested.
    swap_40_41: bool = False

    old_up = (old_tpl_name or "").strip().upper()
    new_up = (new_tpl_name or "").strip().upper()

    # Explicit exclusions / gating rules
    enablemx_port3_source = old_up.startswith("GBR-CT-CLOUDSTORE-ENABLEMX-PORT3")
    classic_cloudstore_source = bool(re.match(r"^GBR-CT-CLOUDSTORE-\d{3}$", old_up))
    presce_mx75_target = bool(re.match(r"^GBR-CT-CLOUDSTORE-PRESCE-\d{3}-MX75$", new_up))

    if enablemx_port3_source:
        # Flatten not allowed; therefore swap never allowed.
        if flatten_requested:
            logging.warning("flatten_requested=True but source template is EnableMX-Port3; forcing no-swap.")
        flatten_requested = False

    # Only swap if operator explicitly requested flatten AND target is PreSCE-###-MX75
    # AND (optionally) old template is the classic CloudStore-### family.
    if flatten_requested and presce_mx75_target and classic_cloudstore_source and not enablemx_port3_source:
        swap_40_41 = True

    if swap_40_41:
        print("🔁 Flatten requested: swapping VLAN 40 and 41 (and DHCP reservations) for PreSCE-###-MX75.")
        log_change(
            "vlan_swap_enabled",
            "Operator requested flatten; VLAN 40/41 swap enabled",
            network_id=network_id,
            network_name=network_name,
            misc=json.dumps({
                "old_template": old_tpl_name,
                "new_template": new_tpl_name,
            }),
        )
        # Flattening case: do the 40/41 swap + DHCP re-IP (your helper must handle DHCP safely)
        vlans_to_push = maybe_swap_vlan_40_41(vlan_list)
    else:
        # No flattening: push VLANs exactly as-is. DHCP stays with same VLAN IDs.
        vlans_to_push = vlan_list
        logging.info(
            "No VLAN swap (swap_40_41=False). old=%r new=%r flatten_requested=%s classic_source=%s enablemx_port3=%s",
            old_tpl_name, new_tpl_name, flatten_requested, classic_cloudstore_source, enablemx_port3_source
        )

    # -------- Safety: VLANs must still be enabled post-bind --------
    enabled = vlans_enabled(network_id)

    if enabled is False:
        print("❌ VLANs are disabled on this network after binding. Rolling back immediately...")
        rollback_all_changes(
            network_id=network_id,
            pre_change_devices=pre_change_devices or [],
            pre_change_vlans=pre_change_vlans or [],
            pre_change_template=pre_change_template,
            org_id=org_id,
            claimed_serials=claimed_serials or [],
            removed_serials=removed_serials or [],
            ms_list=ms_list or [],
            network_name=network_name,
        )
        log_change('workflow_end', 'Exited after rollback due to VLANs disabled (pre-check)')
        raise SystemExit(1)

    if enabled is None:
        print("❌ Could not verify VLAN settings after multiple attempts. Rolling back immediately...")
        rollback_all_changes(
            network_id=network_id,
            pre_change_devices=pre_change_devices or [],
            pre_change_vlans=pre_change_vlans or [],
            pre_change_template=pre_change_template,
            org_id=org_id,
            claimed_serials=claimed_serials or [],
            removed_serials=removed_serials or [],
            ms_list=ms_list or [],
            network_name=network_name,
        )
        log_change('workflow_end', 'Exited after rollback due to VLAN settings check failure')
        raise SystemExit(1)

    try:
        update_vlans(network_id, network_name, vlans_to_push)
    except requests.exceptions.HTTPError as e:
        if is_vlans_disabled_error(e):
            print("❌ VLANs disabled error during VLAN update. Rolling back immediately...")
            rollback_all_changes(
                network_id=network_id,
                pre_change_devices=pre_change_devices or [],
                pre_change_vlans=pre_change_vlans or [],
                pre_change_template=pre_change_template,
                org_id=org_id,
                claimed_serials=claimed_serials or [],
                removed_serials=removed_serials or [],
                ms_list=ms_list or [],
                network_name=network_name,
            )
            log_change('workflow_end', 'Exited after rollback due to VLANs disabled during VLAN update')
            raise SystemExit(1)
        raise

def _parse_switch_model_series_ports(model_str: str) -> Tuple[str, Optional[int]]:
    """
    Parse a Meraki switch model into (series, port_count), e.g.:
      'MS225-24P'  -> ('MS225', 24)
      'MS225-48LP' -> ('MS225', 48)
    If parsing fails, ports is None.
    """
    if not model_str:
        return "", None
    m = re.match(r"^(MS\d+)-(\d+)", model_str.upper())
    if not m:
        return model_str.upper(), None
    series = m.group(1)
    try:
        ports = int(m.group(2))
    except ValueError:
        ports = None
    return series, ports

def _profile_supports_switch_model(profile: Dict[str, Any], switch_model: str) -> bool:
    """
    True if this switch profile is suitable for the given switch model.
    - Same series (e.g. MS225)
    - Same port count when both are known (24 vs 48)
    """
    if not switch_model:
        return False

    sw_series, sw_ports = _parse_switch_model_series_ports(switch_model)
    if not sw_series:
        return False

    prof_models = profile.get("model")
    if not prof_models:
        return False

    # Meraki SDK may give a string or a list of models
    if isinstance(prof_models, str):
        prof_models_list = [prof_models]
    else:
        prof_models_list = list(prof_models)

    for pm in prof_models_list:
        p_series, p_ports = _parse_switch_model_series_ports(str(pm))
        if not p_series:
            continue
        if p_series != sw_series:
            continue
        # If we know both port counts and they differ, skip
        if sw_ports is not None and p_ports is not None and sw_ports != p_ports:
            continue
        # Otherwise treat as compatible
        return True

    return False

def select_switch_profile_interactive_by_model(
    tpl_profiles: List[Dict[str, Any]],
    tpl_profile_map: Dict[str, str],
    switch_model: str
) -> Optional[str]:
    """
    Show only profiles that are compatible with the given switch model
    (same series, same 24/48 count where known).
    """
    candidates = [p for p in tpl_profiles if _profile_supports_switch_model(p, switch_model)]
    if not candidates:
        print(f"No switch profiles in template support {switch_model}.")
        return None

    print(f"\nAvailable switch profiles for {switch_model}:")
    for idx, p in enumerate(candidates, 1):
        name = str(p.get("name") or "<unnamed>")
        print(f"{idx}. {name}")

    # Force this to be List[str] so type-checkers are happy
    profile_names: List[str] = []
    for p in candidates:
        profile_names.append(str(p.get("name") or ""))

    while True:
        choice = input("Select switch profile by number (or Enter to skip): ").strip()
        if not choice:
            return None
        if choice.isdigit():
            idx = int(choice) - 1
            if 0 <= idx < len(profile_names):
                selected_name: str = profile_names[idx]
                profile_id = tpl_profile_map.get(selected_name)
                if profile_id:
                    return profile_id
                print("Selected profile has no ID mapping; please choose another.")
                continue
        print("Invalid selection. Please try again.")

def select_switch_profile_for_network(
    tpl_profiles: List[Dict[str, Any]],
    example_switch_model: Optional[str] = None
) -> Optional[str]:
    """
    Show switch profiles on the new template for a representative switch model.
    If example_switch_model is provided, only show compatible profiles
    (same series, same 24/48 count).
    """
    if not tpl_profiles:
        print("No switch profiles found on the new template.")
        return None

    if example_switch_model:
        candidates = [
            p for p in tpl_profiles
            if _profile_supports_switch_model(p, example_switch_model)
        ]
        if not candidates:
            print(f"No switch profiles compatible with {example_switch_model} found.")
            return None
    else:
        candidates = tpl_profiles

    print("\nSwitch profiles available on the new template:")
    for idx, p in enumerate(candidates, 1):
        print(f"{idx}. {p.get('name', '<unnamed>')}  (ID: {p.get('switchProfileId')})")

    while True:
        raw = input("Select switch profile # to use for this network (or press Enter to skip): ").strip()
        if not raw:
            print("Skipping network-level switch profile selection.")
            return None
        if not raw.isdigit():
            print("Please enter a valid number from the list.")
            continue
        idx = int(raw)
        if not (1 <= idx <= len(candidates)):
            print("Number out of range.")
            continue
        chosen = candidates[idx - 1]
        return chosen.get("switchProfileId")


# =====================
# Rollback
# =====================

def rollback_all_changes(
    network_id: str,
    pre_change_devices: List[Dict[str, Any]],
    pre_change_vlans: List[Dict[str, Any]],
    pre_change_template: Optional[str],
    org_id: str,
    *,
    claimed_serials: Optional[List[str]] = None,
    removed_serials: Optional[List[str]] = None,
    ms_list: Optional[List[Dict[str, Any]]] = None,
    network_name: str,
) -> None:
    """
    Roll back safely:
      - Only REMOVE devices that were ADDED during this run.
      - Never remove MG unless it was added during this run.
      - Only RE-ADD devices that were present pre-change and are now missing.
      - Retry VLAN restore with backoff.
      - Produce rollback summary JSON.
      - NEVER crash midway through rollback.
    """
    print("=== Starting rollback to previous network state ===")

    rb_failures: List[Dict[str, Any]] = []
    per_device_attempts: Dict[str, int] = {}
    failed_devices: Dict[str, List[Dict[str, Any]]] = {}

    def _stage_fail(stage: str, exc: Exception) -> None:
        logging.exception("Rollback stage failed: %s", stage)
        rb_failures.append({
            "stage": stage,
            "error_type": type(exc).__name__,
            "error": str(exc),
        })

    def _record_device_failure(
        serial: str,
        *,
        stage: str,
        outcome: RetryOutcome,
        extra: Optional[Dict[str, Any]] = None,
    ) -> None:
        failed_devices.setdefault(serial, []).append({
            "serial": serial,
            "stage": stage,
            "attempts": outcome.attempts,
            "error_type": outcome.error_type,
            "error": outcome.error,
            "http_status": outcome.http_status,
            "last_response_text": outcome.last_response_text,
            "extra": extra or {},
        })

    # -------------------------------------------------
    # Build lookup maps
    # -------------------------------------------------
    pre_serials: Set[str] = {
        d.get("serial", "") for d in pre_change_devices if d.get("serial")
    }
    pre_by_serial: Dict[str, Dict[str, Any]] = {
        d["serial"]: d for d in pre_change_devices if d.get("serial")
    }

    # -------------------------------------------------
    # Stage 1: Remove devices added during this run
    # -------------------------------------------------
    try:
        safe_claimed = list(claimed_serials or [])
        current_serials: Set[str] = set()

        if safe_claimed:
            try:
                current_devices = meraki_get(f"/networks/{network_id}/devices") or []
                if isinstance(current_devices, list):
                    current_serials = {
                        serial
                        for d in current_devices
                        if isinstance(d, dict)
                        and (serial := d.get("serial")) is not None
                    }
            except Exception as e:
                _stage_fail("fetch_current_devices_before_remove", e)


            for serial in safe_claimed:
                if not serial or serial in pre_serials:
                    continue
                if current_serials and serial not in current_serials:
                    continue

                try:
                    inv = get_inventory_device(org_id, serial) or {}
                    model = (inv.get("model") or "").upper()
                    if model.startswith("MG") and serial in pre_serials:
                        continue
                except Exception:
                    logging.exception(
                        "Rollback inventory check failed for %s; skipping removal",
                        serial,
                    )
                    continue

                print(f"Removing claimed device: {serial}")

                def _op_remove() -> None:
                    do_action(
                        meraki_post,
                        f"/networks/{network_id}/devices/remove",
                        data={"serial": serial},
                    )

                outcome = retry_with_backoff(
                    f"rollback_remove_{serial}",
                    _op_remove,
                    max_attempts=4,
                    base_sleep=1.0,
                    max_sleep=10.0,
                )
                per_device_attempts[serial] = outcome.attempts

                if outcome.ok:
                    log_change(
                        "rollback_device_removed",
                        "Removed device added during this run",
                        device_serial=serial,
                        network_id=network_id,
                        network_name=network_name,
                    )
                else:
                    _record_device_failure(
                        serial,
                        stage="remove_added_device",
                        outcome=outcome,
                    )
    except Exception as e:
        _stage_fail("stage_1_remove_added_devices", e)

    # -------------------------------------------------
    # Stage 2: Restore original template
    # -------------------------------------------------
    try:
        print("Restoring config template binding...")

        def _op_unbind() -> None:
            do_action(meraki_post, f"/networks/{network_id}/unbind")

        out_unbind = retry_with_backoff(
            "rollback_unbind",
            _op_unbind,
            max_attempts=4,
            base_sleep=1.0,
            max_sleep=10.0,
        )
        if not out_unbind.ok:
            rb_failures.append({"stage": "unbind", **asdict(out_unbind)})

        if pre_change_template:
            def _op_bind() -> None:
                do_action(
                    meraki_post,
                    f"/networks/{network_id}/bind",
                    data={"configTemplateId": pre_change_template},
                )

            out_bind = retry_with_backoff(
                "rollback_bind_original_template",
                _op_bind,
                max_attempts=4,
                base_sleep=2.0,
                max_sleep=15.0,
            )
            if not out_bind.ok:
                rb_failures.append({"stage": "bind_original_template", **asdict(out_bind)})

        log_change(
            "rollback_template",
            f"Restored template binding {pre_change_template}",
            device_name=f"Network: {network_id}",
            network_id=network_id,
            network_name=network_name,
        )

        print("Waiting for template binding to take effect (sleeping 15 seconds)...")
        time.sleep(15)
    except Exception as e:
        _stage_fail("stage_2_restore_template", e)

    # -------------------------------------------------
    # Stage 3: Re-add devices present pre-change but missing
    # -------------------------------------------------
    try:
        current_serials: Set[str] = set()
        try:
            current_devices = meraki_get(f"/networks/{network_id}/devices") or []
            if isinstance(current_devices, list):
                current_serials = {
                    serial
                    for d in current_devices
                    if isinstance(d, dict)
                    and (serial := d.get("serial")) is not None
                }
        except Exception as e:
            _stage_fail("fetch_current_devices_after_template", e)

        for serial in sorted(pre_serials - current_serials):
            if not serial:
                continue

            try:
                inv = get_inventory_device(org_id, serial) or {}
                if inv.get("networkId"):
                    continue
            except Exception:
                logging.exception(
                    "Rollback inventory lookup failed for %s; skipping re-add",
                    serial,
                )
                continue

            print(f"Re-adding previously present device: {serial}")

            def _op_claim() -> None:
                do_action(
                    meraki_post,
                    f"/networks/{network_id}/devices/claim",
                    data={"serials": [serial]},
                )

            outcome = retry_with_backoff(
                f"rollback_readd_{serial}",
                _op_claim,
                max_attempts=4,
                base_sleep=2.0,
                max_sleep=15.0,
            )
            per_device_attempts[serial] = max(
                per_device_attempts.get(serial, 0),
                outcome.attempts,
            )

            if outcome.ok:
                log_change(
                    "rollback_device_readded",
                    "Device re-added during rollback",
                    device_serial=serial,
                    network_id=network_id,
                    network_name=network_name,
                )
            else:
                _record_device_failure(
                    serial,
                    stage="readd_missing_device",
                    outcome=outcome,
                )
    except Exception as e:
        _stage_fail("stage_3_readd_devices", e)

    # -------------------------------------------------
    # Stage 4: Restore VLANs (retry with backoff)
    # -------------------------------------------------
    vlan_restore_outcome: Optional[RetryOutcome] = None
    try:
        print("Restoring VLANs and DHCP assignments...")
        time.sleep(5)

        def _op_vlan_restore() -> None:
            update_vlans(network_id, network_name, pre_change_vlans)

        vlan_restore_outcome = retry_with_backoff(
            "rollback_vlan_restore",
            _op_vlan_restore,
            max_attempts=6,
            base_sleep=2.0,
            max_sleep=20.0,
            verbose=True,
        )

        if vlan_restore_outcome.ok:
            log_change(
                "rollback_vlans",
                "Restored VLANs and DHCP assignments",
                device_name=f"Network: {network_id}",
                network_id=network_id,
                network_name=network_name,
            )
        else:
            log_change(
                "rollback_vlans_failed",
                "Failed restoring VLANs during rollback",
                device_name=f"Network: {network_id}",
                network_id=network_id,
                network_name=network_name,
            )
            rb_failures.append({
                "stage": "vlan_restore",
                **asdict(vlan_restore_outcome),
            })
    except Exception as e:
        _stage_fail("stage_4_restore_vlans", e)

    # -------------------------------------------------
    # Write rollback summary JSON
    # -------------------------------------------------
    try:
        summary = {
            "timestamp": datetime.now().isoformat(),
            "org_id": org_id,
            "network_id": network_id,
            "network_name": network_name,
            "pre_change_template": pre_change_template,
            "failures": rb_failures,
            "failed_devices": failed_devices,
            "attempts_by_serial": per_device_attempts,
            "vlan_restore_outcome": asdict(vlan_restore_outcome)
            if vlan_restore_outcome else None,
        }
        rb_path = os.path.join(
            "logs",
            f"rollback_summary_{org_id}_{network_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
        )
        _write_json_log(rb_path, summary)
        print(f"📝 Rollback summary saved to {rb_path}")
    except Exception:
        logging.exception("Failed writing rollback summary JSON")

    # -------------------------------------------------
    # Console summary
    # -------------------------------------------------
    if failed_devices or rb_failures:
        print("\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        print("⚠️ Rollback Completed With Issues")
        print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        for serial, failures in failed_devices.items():
            for f in failures:
                print(
                    f"  • {serial}: stage={f['stage']} "
                    f"attempts={f['attempts']} "
                    f"http={f['http_status']} "
                    f"err={f['error']}"
                )
        print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")

    print("=== Rollback complete ===")

    # -------------------------------------------------
    # Clear checkpoint (best-effort)
    # -------------------------------------------------
    global _current_checkpoint
    try:
        cp = _current_checkpoint
        if cp and cp.org_id == org_id and cp.network_id == network_id:
            cp_path = cp.path()
            if os.path.isfile(cp_path):
                os.remove(cp_path)
                print("🧹 Cleared resume checkpoint after rollback.")
            _current_checkpoint = None
    except Exception:
        logging.exception("Failed to clear checkpoint after rollback")


# =====================
# Step Summary helpers (✅ / ❌ and skip N/A)
# =====================
StatusVal = Union[bool, str]  # True/False/"NA"

def _fmt(val: StatusVal) -> str:
    if val is True:
        return "✅ Success"
    if val is False:
        return "❌ Failed"
    return str(val)

def write_run_summary_json(
    *,
    org_id: str,
    network_id: str,
    network_name: str,
    step_status: Dict[str, StatusVal],
    claimed_serials: List[str],
    removed_serials: List[str],
    extra: Optional[Dict[str, Any]] = None,
) -> None:
    payload: Dict[str, Any] = {
        "timestamp": datetime.now().isoformat(),
        "org_id": org_id,
        "network_id": network_id,
        "network_name": network_name,
        "step_status": step_status,
        "vlan_updates": {
            "ok": sorted(set(VLAN_UPDATES_OK)),
            "skipped": sorted(set(VLAN_UPDATES_SKIPPED)),
            "failed": sorted(set(VLAN_UPDATES_FAILED)),
        },
        "claimed_serials_delta": claimed_serials,
        "removed_serials_delta": removed_serials,
        "extra": extra or {},
    }
    path = os.path.join("logs", f"run_summary_{org_id}_{network_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
    _write_json_log(path, payload)
    if DEBUG_MODE:
        print(f"📝 Run summary saved: {path}")


def print_summary(step_status: Dict[str, StatusVal]) -> None:
    order = [
        'template_bound',
        'vlans_updated',
        'devices_claimed',
        'mx_removed',
        'mr33_removed',
        'configured',
        'old_mx',
        'old_mr33',
    ]
    print("\nStep Summary:")
    for step in order:
        val = step_status.get(step, "NA")
        if isinstance(val, str) and val.upper() == "NA":
            continue
        print(f" - {step}: {_fmt(val)}")

def print_debug_json_log_summary() -> None:
    """
    If DEBUG_MODE is enabled, print all JSON diagnostic files created in this run.
    """
    if not DEBUG_MODE:
        return

    print("\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print("🐞 DEBUG MODE: JSON Diagnostic Files Created")
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")

    if not JSON_DEBUG_FILES_CREATED:
        print("No JSON debug logs were created during this run.")
    else:
        for path in JSON_DEBUG_FILES_CREATED:
            print(f"📄 {path}")

    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")

def print_api_error_log_summary() -> None:
    """
    At the end of the workflow, inform the user whether API error logs
    were generated during this run.
    """
    networks_log = os.path.join("logs", "last_networks_error.json")
    vlan_log = os.path.join("logs", "last_vlan_error.json")

    networks_exists = os.path.isfile(networks_log)
    vlan_exists = os.path.isfile(vlan_log)

    print("\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print("📄 API Error Log Summary")
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")

    if networks_exists:
        print(f"❗ Network listing errors detected — see: {networks_log}")
    else:
        print("✔ No network listing errors recorded.")

    if vlan_exists:
        print(f"❗ VLAN retrieval errors detected — see: {vlan_log}")
    else:
        print("✔ No VLAN retrieval errors recorded.")

    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")

def _slug_filename(s: str) -> str:
    s = re.sub(r'[^A-Za-z0-9._-]+', '-', s).strip('-_')
    return s[:80]

def _json(x: Any) -> str:
    try:
        return json.dumps(x, ensure_ascii=False)
    except Exception:
        return str(x)

def _normalize_tags_list(val) -> List[str]:
    if isinstance(val, list):
        return sorted(str(t) for t in val)
    if isinstance(val, str):
        return sorted([t for t in val.split() if t])
    return []

def _autosize(ws):
    max_col: int = ws.max_column
    max_row: int = ws.max_row
    for col_idx in range(1, max_col + 1):
        max_len = 0
        for row_idx in range(1, max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                s = str(val)
                if len(s) > max_len:
                    max_len = len(s)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

def _write_snapshot_sheet(
    ws,
    *,
    org_id: str,
    network_id: str,
    network_name: str,
    template_id: Optional[str],
    vlan_list: List[Dict[str, Any]],
    mx_list: List[Dict[str, Any]],
    ms_list: List[Dict[str, Any]],
    mr_list: List[Dict[str, Any]],
    profileid_to_name: Optional[Dict[str, str]] = None,
    tpl_name_lookup: Optional[Callable[[str], str]] = None,
):
    header: List[str] = [
        "section", "network_id", "network_name", "item_type",
        "col1", "col2", "col3", "col4", "col5",
        "switch_profile_id", "switch_profile_name", "extra_info"
    ]
    ws.append(header)

    tpl_name = ""
    if template_id:
        try:
            if tpl_name_lookup:
                tpl_name = tpl_name_lookup(template_id) or ""
            else:
                tpl_name = _get_template_name(org_id, template_id) or ""
        except Exception:
            logging.exception("Could not fetch template name for snapshot")


    ws.append([
        "template", network_id, network_name, "template",
        template_id or "", tpl_name, "", "", "",
        "", "", ""
    ])

    for v in vlan_list:
        ws.append([
            "vlans", network_id, network_name, "vlan",
            str(v.get("id", "")),
            str(v.get("name", "") or ""),
            str(v.get("subnet", "") or ""),
            str(v.get("applianceIp", "") or ""),
            str(v.get("dhcpHandling", "") or ""),
            "", "",
            _json({k: v.get(k) for k in v.keys() - {"id", "name", "subnet", "applianceIp", "dhcpHandling"}}),
        ])

    def device_row(d: Dict[str, Any]) -> List[str]:
        tags_list = _normalize_tags_list(d.get("tags", []))
        sp_id: str = str(d.get("switchProfileId", "") or "")
        sp_name: str = str(d.get("switchProfileName", "") or "")
        if (not sp_name) and sp_id and profileid_to_name:
            sp_name = profileid_to_name.get(sp_id, "") or ""
        return [
            "devices", network_id, network_name, "device",
            str(d.get("serial", "") or ""),
            str(d.get("model", "") or ""),
            str(d.get("name", "") or ""),
            str(d.get("address", "") or ""),
            " ".join(tags_list),
            sp_id,
            sp_name,
            ""
        ]

    for d in (mx_list + ms_list + mr_list):
        ws.append(device_row(d))

    for sw in ms_list:
        changes_by_port: Dict[str, Dict[str, Any]] = sw.get("port_overrides") or {}
        if not isinstance(changes_by_port, dict) or not changes_by_port:
            continue
        for port_id, changes in changes_by_port.items():
            if not isinstance(changes, dict):
                continue
            for fld, val in changes.items():
                ws.append([
                    "port_overrides", network_id, network_name, "port_override",
                    str(sw.get("serial", "") or ""), str(port_id), str(fld),
                    "" if isinstance(val, (dict, list)) else str(val),
                    "", "", "",
                    _json(val) if isinstance(val, (dict, list)) else "",
                ])
def export_combined_snapshot_xlsx(
    *,
    org_id: str,
    network_id: str,
    network_name: str,

    # PRE
    pre_template_id: Optional[str],
    pre_vlan_list: List[Dict[str, Any]],
    pre_mx_list: List[Dict[str, Any]],
    pre_ms_list: List[Dict[str, Any]],
    pre_mr_list: List[Dict[str, Any]],
    pre_profileid_to_name: Optional[Dict[str, str]] = None,

    # POST
    post_template_id: Optional[str],
    post_vlan_list: List[Dict[str, Any]],
    post_mx_list: List[Dict[str, Any]],
    post_ms_list: List[Dict[str, Any]],
    post_mr_list: List[Dict[str, Any]],
    post_profileid_to_name: Optional[Dict[str, str]] = None,

    outfile: Optional[str] = None
) -> None:
    """
    Creates ONE workbook with 3 sheets: PRE, POST, SUMMARY.

    PRE  -> snapshot before changes
    POST -> snapshot after changes
    SUMMARY -> VLAN update summary (OK/skipped/failed) for this run.
    """
    if outfile:
        out_path = outfile
    else:
        base = _slug_filename(_network_tag_from_name(network_name))
        out_path = f"{base}_combined_{timestamp}.xlsx"

    wb = Workbook()

    # PRE sheet
    ws_pre = wb.active
    assert ws_pre is not None
    ws_pre.title = "PRE"
    _write_snapshot_sheet(
        ws_pre,
        org_id=org_id,
        network_id=network_id,
        network_name=network_name,
        template_id=pre_template_id,
        vlan_list=pre_vlan_list,
        mx_list=pre_mx_list,
        ms_list=pre_ms_list,
        mr_list=pre_mr_list,
        profileid_to_name=pre_profileid_to_name
    )
    _autosize(ws_pre)

    # POST sheet
    ws_post = wb.create_sheet("POST")
    _write_snapshot_sheet(
        ws_post,
        org_id=org_id,
        network_id=network_id,
        network_name=network_name,
        template_id=post_template_id,
        vlan_list=post_vlan_list,
        mx_list=post_mx_list,
        ms_list=post_ms_list,
        mr_list=post_mr_list,
        profileid_to_name=post_profileid_to_name
    )
    _autosize(ws_post)

    # SUMMARY sheet (new)
    ws_summary = wb.create_sheet("SUMMARY")

    # Header/meta
    ws_summary.append(["Network ID", network_id])
    ws_summary.append(["Network Name", network_name])
    ws_summary.append(["Org ID", org_id])
    ws_summary.append(["Export Timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_summary.append([])

    ws_summary.append(["Section", "Count", "VLAN IDs (comma-separated)"])

    ok_sorted = sorted(set(VLAN_UPDATES_OK), key=lambda x: int(x) if x.isdigit() else x)
    skipped_sorted = sorted(set(VLAN_UPDATES_SKIPPED), key=lambda x: int(x) if x.isdigit() else x)
    failed_sorted = sorted(set(VLAN_UPDATES_FAILED), key=lambda x: int(x) if x.isdigit() else x)

    ws_summary.append([
        "Updated OK",
        len(ok_sorted),
        ", ".join(ok_sorted) if ok_sorted else "",
    ])
    ws_summary.append([
        "Skipped (not present on MX)",
        len(skipped_sorted),
        ", ".join(skipped_sorted) if skipped_sorted else "",
    ])
    ws_summary.append([
        "Failed to update",
        len(failed_sorted),
        ", ".join(failed_sorted) if failed_sorted else "",
    ])

    _autosize(ws_summary)

    wb.save(out_path)
    print(f"📗 Combined PRE/POST/SUMMARY snapshot exported to: {out_path}")
    log_change(
        "snapshot_export_combined",
        f"Exported combined PRE/POST/SUMMARY snapshot to {out_path}",
        network_id=network_id,
        network_name=network_name,
    )


def _network_tag_from_name(name: str) -> str:
    parts = name.split('-')
    if len(parts) >= 2 and parts[1].isdigit():
        return f"{parts[0]}-{parts[1]}"
    return name

# ======= New extracted helpers to eliminate duplication =======

def maybe_prompt_and_rollback(org_id, network_id, pre_change_devices, pre_change_vlans,
                              pre_change_template, ms_list, network_name,
                              claimed_serials=None, removed_serials=None) -> None:
    choice = prompt_rollback_big()
    if choice in {'yes', 'y'}:
        print("\nRolling back all changes...")
        log_change('rollback_start', 'User requested rollback')
        rollback_all_changes(
            network_id=network_id,
            pre_change_devices=pre_change_devices,
            pre_change_vlans=pre_change_vlans,
            pre_change_template=pre_change_template,
            org_id=org_id,
            claimed_serials=claimed_serials or [],
            removed_serials=removed_serials or [],
            ms_list=ms_list,
            network_name=network_name,
        )
        print("✅ Rollback complete.")
        log_change('rollback_end', 'Rollback completed')
    elif choice in {'no', 'n'}:
        print("\nProceeding without rollback. Rollback option will no longer be available.")
        log_change('workflow_end', 'Script finished (no rollback)')
    else:
        print("\n❌ No rollback selected (Enter pressed).")
        print("⚠️  Rollback is no longer available. Please ensure the network is functional and all required checks have been carried out.")
        log_change('workflow_end', 'Script finished (rollback skipped with Enter)')

# =====================
# Robust network selector
# =====================
def select_network_interactive(org_id: str) -> Tuple[str, str]:
    while True:
        partial = input("Enter partial network name to search (or press Enter to cancel): ").strip()
        if not partial:
            print("\n❌ No Network selected -----------\n   Please retry when Network is known *******")
            sys.exit(1)

        networks = fetch_matching_networks(org_id, partial)

        if not networks:
            # No matches for this search term – but is that because
            #   a) there are genuinely no networks? or
            #   b) the search doesn't match any of them?
            try:
                all_nets = meraki_list_networks_all(org_id)
            except SystemExit:
                # _get_org_networks_cached has already printed a fatal API error.
                raise
            except Exception as e:
                logging.exception("Error while verifying org networks after no matches")
                print(
                    f"\n🚨 ERROR: Could not verify networks for org {org_id} after a "
                    f"zero-match search. Details: {e}\n"
                )
                sys.exit(1)

            if not all_nets:
                # This is the "API says there are zero networks in this org" case
                print("\n🚨 Meraki API returned ZERO networks for this organization.")
                print("    This is unlikely to be just a search typo.")
                print("    Please verify:")
                print("      - Your API key has access to this org")
                print("      - The org ID is correct")
                print("      - Meraki dashboard/API is healthy\n")
                sys.exit(1)

            # At this point we *do* have networks, just none matching the filter
            print("\n❌ No matching networks found for that search term -----------")
            print(f"   (Org has {len(all_nets)} networks in total.)")
            retry = input("Search again with a different name? (y/N): ").strip().lower()
            if retry != 'y':
                print("\n❌ No Network selected -----------\n   Please retry when Network is known *******")
                sys.exit(1)
            continue

        if len(networks) == 1:
            only = networks[0]
            print(f"\n1 match: {only['name']} (ID: {only['id']})")
            confirm = input("Use this network? (Y/n): ").strip().lower()
            if confirm in {"", "y", "yes"}:
                print(f"Selected network: {only['name']} (ID: {only['id']})")
                return only['id'], only['name']
            else:
                continue

        print("\nMultiple networks found:")
        for idx, net in enumerate(networks, 1):
            print(f"{idx}. {net['name']} (ID: {net['id']})")

        while True:
            raw = input("Select the network by number (or press Enter to cancel): ").strip()
            if not raw:
                print("\n❌ No Network selected -----------\n   Please retry when Network is known *******")
                sys.exit(1)
            if raw.isdigit():
                choice = int(raw)
                if 1 <= choice <= len(networks):
                    chosen = networks[choice - 1]
                    print(f"Selected network #{choice}: {chosen['name']} (ID: {chosen['id']})")
                    return chosen['id'], chosen['name']
            print("❌ Invalid selection. Please enter a valid number from the list.")

# =====================
# Org selector
# =====================
def select_org() -> str:
    orgs = meraki_get("/organizations")
    if not orgs:
        print("\n❌ No Organisations returned from API -----------")
        print("   Please retry when Org is known *******")
        sys.exit(1)

    print("Organizations:")
    for idx, org in enumerate(orgs, 1):
        print(f"{idx}. {org['name']} (ID: {org['id']})")

    raw = input("Select organization by number (or press Enter to cancel): ").strip()
    if not raw:
        print("\n❌ No Organisation selected -----------")
        print("   Please retry when Org is known *******")
        sys.exit(1)

    try:
        org_idx = int(raw)
        if org_idx < 1 or org_idx > len(orgs):
            raise ValueError("out of range")
    except Exception:
        print("\n❌ Invalid Organisation selection -----------")
        print("   Please retry when Org is known *******")
        sys.exit(1)

    return orgs[org_idx - 1]['id']

# ------------- Change Rollback Font -------------
def prompt_rollback_big() -> str:
    try:
        from rich.console import Console
        from rich.panel import Panel
        from rich.text import Text
        import pyfiglet
    except Exception:
        return prompt_rollback_basic()

    console = Console()
    banner = pyfiglet.figlet_format("ROLLBACK", font="slant")
    console.print(f"[bold red]{banner}[/bold red]")
    console.print(Panel.fit(
        Text(
            "Type 'yes' to rollback changes, 'no' to continue without rollback, or just press Enter to skip.\n"
            "IMPORTANT: If you skip (press Enter), rollback will no longer be available.\n"
            "Have you ensured the network is fully functional and all required checks have been carried out?",
            style="bold white"
        ),
        title="⚠️  ROLLBACK OPTION",
        title_align="left",
        border_style="red"
    ))
    return input("> ").strip().lower()

def prompt_rollback_basic() -> str:
    print("\n" + "!"*78)
    print("⚠️  R O L L B A C K   O P T I O N  ⚠️".center(78))
    print("!"*78)
    print("Type 'yes' to rollback changes, 'no' to continue without rollback, or just press Enter to skip.")
    print("IMPORTANT: If you skip (Enter), rollback will no longer be available.")
    print("Have you ensured the network is fully functional and all required checks have been carried out?")
    return input("> ").strip().lower()

# =====================
# Main
# =====================
if __name__ == '__main__':
    print(f"\n🔧 Meraki Rebind Tool — Version {SCRIPT_VERSION}\n")
    # -------------------------------------------------------------
    # Helpful CLI hints for debugging API issues (new addition)
    # -------------------------------------------------------------
    print("💡 If you encounter errors fetching networks or VLANs, check:")
    print("      • logs/last_networks_error.json   (org-wide network listing failures)")
    print("      • logs/last_vlan_error.json       (VLAN fetch failures)")
    print("    These files contain raw Meraki API responses and diagnostics.\n")
    log_change('workflow_start', 'Script started')
    logging.info(f"Starting Meraki Rebind Tool (version {SCRIPT_VERSION})")
        
    step_status: Dict[str, StatusVal] = {}

    # -------- Select Org --------
    org_id = select_org()

    # -------- Prompt/validate serials (org-level), then summarize --------
    prevalidated_serials = prompt_and_validate_serials(org_id)
    detected_mx_models = summarize_devices_in_org(org_id, prevalidated_serials)

    mx_model_filter: Optional[str] = None
    if detected_mx_models == {'MX67'}:
        mx_model_filter = 'MX67'
    elif detected_mx_models == {'MX75'}:
        mx_model_filter = 'MX75'

    # -------- Select Network --------
    network_id, network_name = select_network_interactive(org_id)
    
    # ---- Resume checkpoint (create or load) ----
    cp_existing = Checkpoint.load(org_id, network_id)
    if cp_existing:
        print(f"\n🟨 Found previous session for {network_name} ({network_id}).")
        ans = input("Resume where you left off? (Y/n): ").strip().lower()
        if ans in {"", "y", "yes"}:
            _current_checkpoint = cp_existing
            # Carry over previous step_status; we still recalc live state below
            step_status.update(_current_checkpoint.step_status or {})
        else:
            _current_checkpoint = Checkpoint(
                org_id=org_id, network_id=network_id, network_name=network_name,
                step_status={}, pre_change_template=None,
                pre_change_devices=[], pre_change_vlans=[],
                claimed_serials=[], removed_serials=[]
            )
            _current_checkpoint.save()
    else:
        _current_checkpoint = Checkpoint(
            org_id=org_id, network_id=network_id, network_name=network_name,
            step_status={}, pre_change_template=None,
            pre_change_devices=[], pre_change_vlans=[],
            claimed_serials=[], removed_serials=[]
        )
        _current_checkpoint.save()

    net_info = meraki_get(f"/networks/{network_id}")
    old_template: Optional[str] = net_info.get('configTemplateId')

    # -------- Pre-change snapshot incl. MS port overrides --------
    mx, ms, mr = fetch_devices(org_id, network_id, template_id=old_template)
    pre_change_devices = mx + ms + mr
    pre_change_vlans = fetch_vlan_details(network_id)
    pre_change_template = old_template
    pre_change_serials: Set[str] = {d['serial'] for d in pre_change_devices}
    
    # Save PRE state into checkpoint
    _current_checkpoint.pre_change_template = pre_change_template
    _current_checkpoint.pre_change_devices = pre_change_devices
    _current_checkpoint.pre_change_vlans = pre_change_vlans
    _current_checkpoint.save()

    # For snapshot/xlsx mapping: template profileId -> name
    old_profileid_to_name: Dict[str, str] = {}
    if old_template:
        try:
            old_tpl_profiles = meraki_get(f"/organizations/{org_id}/configTemplates/{old_template}/switch/profiles") or []
            old_profileid_to_name = {p['switchProfileId']: p['name'] for p in old_tpl_profiles}
        except Exception:
            logging.exception("Failed fetching old template switch profiles")

    # --- Export PRE snapshot ---
    
    # -------- MX gate --------
    current_mx_models = sorted({d['model'] for d in mx})
    is_mx64_present = any(m.startswith('MX64') for m in current_mx_models)

    # ------------------------------------------------------------------
    # PATH A: Current network has MX (not MX64) -> light flow (no rebind)
    # ------------------------------------------------------------------
    if current_mx_models and not is_mx64_present:
        print(f"\nCurrent network: {network_name} (ID: {network_id})")
        # if old_template:
        #     try:
        #         curr_tpl = meraki_get(f"/organizations/{org_id}/configTemplates/{old_template}")
        #         print(f"Bound template: {curr_tpl.get('name','<unknown>')} (ID: {old_template})")
        #     except Exception:
        #         print(f"Bound template ID: {old_template}")
        # else:
        #     print("No template bound.")
        # print(f"Detected MX model(s): {', '.join(current_mx_models)}")
        if old_template:
            curr_name = _get_template_name(org_id, old_template) or "<unknown>"
            print(f"Bound template: {curr_name} (ID: {old_template})")
        else:
            print("No template bound.")
        print(f"Detected MX model(s): {', '.join(current_mx_models)}")


            # Initialize PATH A status fields only once (safe for resume)
        if not _current_checkpoint.done('init_path_a_status'):
            step_status.setdefault('template_bound', "NA")
            step_status.setdefault('vlans_updated', "NA")
            step_status.setdefault('mx_removed', "NA")
            _current_checkpoint.mark('init_path_a_status', True)


        # Optional: VLAN-count based template suggestion in light flow
        try:
            all_templates_raw: Any = meraki_get(f"/organizations/{org_id}/configTemplates")
            all_templates: List[Dict[str, Any]] = (
                all_templates_raw if isinstance(all_templates_raw, list) else []
            )

            vlan_count = _current_vlan_count(network_id)
            # Current bound template name (may be None/empty)
            current_tpl_name = _get_template_name(org_id, old_template) if old_template else None

            suggested_tpl = _pick_template_by_vlan_count(
                all_templates,
                vlan_count,
                current_template_name=current_tpl_name,
            )

            if suggested_tpl and (not old_template or suggested_tpl.get('id') != old_template):
                print(
                    f"\nSuggestion: Based on VLAN count ({vlan_count}), "
                    f"'{suggested_tpl.get('name','')}' looks appropriate "
                    f"(ID: {suggested_tpl.get('id','')})."
                )
                ans = input(
                    "Press 'a' to bind to the suggested template, or Enter to keep current template: "
                ).strip().lower()

                if ans == 'a':
                    try:
                        new_template = suggested_tpl.get('id')
                        if old_template:
                            out_unbind = retry_with_backoff(
                                "lightflow_unbind",
                                lambda: do_action(meraki_post, f"/networks/{network_id}/unbind"),
                                max_attempts=4,
                                base_sleep=1.0,
                                max_sleep=10.0,
                                verbose=True,
                            )
                            if not out_unbind.ok:
                                raise RuntimeError(f"Unbind failed: {out_unbind.error}")

                        out_bind = retry_with_backoff(
                            "lightflow_bind",
                            lambda: do_action(
                                meraki_post,
                                f"/networks/{network_id}/bind",
                                data={"configTemplateId": new_template},
                            ),
                            max_attempts=4,
                            base_sleep=2.0,
                            max_sleep=15.0,
                            verbose=True,
                        )
                        if not out_bind.ok:
                            raise RuntimeError(f"Bind failed: {out_bind.error}")

                        print(f"✅ Bound to {suggested_tpl.get('name','')}")

                        # Re-apply VLANs after (re)bind
                        bind_network_to_template(
                            org_id=org_id,
                            network_id=network_id,
                            tpl_id=new_template,
                            vlan_list=pre_change_vlans,
                            network_name=network_name,
                            pre_change_devices=pre_change_devices,
                            pre_change_vlans=pre_change_vlans,
                            pre_change_template=pre_change_template,
                            claimed_serials=[],
                            removed_serials=[],
                            ms_list=ms,
                        )
                        step_status['template_bound'] = True
                        step_status['vlans_updated'] = True
                        old_template = new_template

                    except requests.exceptions.HTTPError as e:
                        resp = e.response
                        status = resp.status_code if resp is not None else "?"
                        text = resp.text if resp is not None else ""
                        logging.exception(
                            "Light-flow suggested bind failed: HTTP %s: %s", status, text
                        )
                        print("❌ Failed to bind suggested template in light flow.")
                        step_status['template_bound'] = False
                    except Exception:
                        logging.exception("Light-flow suggested bind failed (unexpected)")
                        print(
                            "❌ Failed to bind suggested template in light flow "
                            "(unexpected error)."
                        )
                        step_status['template_bound'] = False
            else:
                logging.debug(
                    "No VLAN-based suggestion available in light flow (vlan_count=%s).",
                    vlan_count,
                )

        except Exception:
            logging.exception("Suggestion stage in light flow failed")

        # Wireless pre-check + claim
        safe_to_claim, mr_removed_serials, mr_claimed_serials = run_wireless_precheck_and_filter_claims(
            org_id, network_id, prevalidated_serials  # allow wireless
        )
        if not _current_checkpoint.done('devices_claimed'):
            claimed = claim_devices(org_id, network_id, prevalidated_serials=safe_to_claim)
            step_status['devices_claimed'] = bool(claimed)
            _current_checkpoint.claimed_serials = claimed or _current_checkpoint.claimed_serials or []
            _current_checkpoint.mark('devices_claimed', step_status['devices_claimed'])
        else:
            print("⏭️  Skipping device claim (already completed).")
            claimed = _current_checkpoint.claimed_serials or []


        # Enable WAN2
        safe_enable_wan2_on_claimed_mx(org_id, claimed)

        # Primary / order
        primary_mx_serial = select_primary_mx(org_id, claimed)
        ensure_primary_mx(network_id, primary_mx_serial)
        mr_order = select_device_order(org_id, claimed, 'MR')
        ms_order = select_device_order(org_id, claimed, 'MS')

        # Template profiles (if any)
        try:
            if old_template:
                tpl_profiles = meraki_get(f"/organizations/{org_id}/configTemplates/{old_template}/switch/profiles") or []
                tpl_profile_map = {p['name']: p['switchProfileId'] for p in tpl_profiles}
            else:
                tpl_profile_map = {}
                tpl_profiles = []
        except Exception:
            logging.exception("Failed fetch template switch profiles")
            tpl_profile_map = {}
            tpl_profiles = []

        # Naming & config
        try:
            name_and_configure_claimed_devices(
                org_id=org_id,
                network_id=network_id,
                network_name=network_name,
                serials=claimed,
                ms_list=ms,
                tpl_profile_map=tpl_profile_map,
                old_mx_devices=mx,
                old_mr_devices=mr,
                primary_mx_serial=primary_mx_serial,
                mr_order=mr_order,
                ms_order=ms_order,
            )
            step_status['configured'] = True
        except Exception:
            logging.exception("Configuration of claimed devices failed")
            step_status['configured'] = False

        # Remove legacy MR33 only if new wireless was claimed
        try:
            inv_models_claimed = _get_inventory_models_for_serials(org_id, claimed)
            claimed_has_wireless = any(_is_wireless_model(m) for m in inv_models_claimed.values())
            if claimed_has_wireless:
                removed_mr33_ok = remove_existing_mr33_devices(org_id, network_id)
                step_status['mr33_removed'] = removed_mr33_ok
                if removed_mr33_ok:
                    log_change('mr33_removed', "Removed old MR33 after new AP claim", misc=f"claimed_serials={claimed}")
            else:
                step_status['mr33_removed'] = "NA"
        except Exception:
            logging.exception("MR33 removal failed")
            step_status['mr33_removed'] = False

        step_status.setdefault('old_mx', "NA")
        step_status.setdefault('old_mr33', "NA")

        remove_recently_added_tag(network_id)
        print_summary(step_status)

        # --- Export POST snapshot (extracted) ---
        
        # -------- Enhanced rollback prompt (extracted) --------
        post_change_devices = meraki_get(f"/networks/{network_id}/devices")
        post_change_serials = {d['serial'] for d in post_change_devices}
        claimed_serials_rb = list(post_change_serials - pre_change_serials)
        removed_serials_rb = list(pre_change_serials - post_change_serials)
        _current_checkpoint.claimed_serials = claimed_serials_rb
        _current_checkpoint.removed_serials = removed_serials_rb
        _current_checkpoint.save()

        
        # --- Build POST state & export one combined workbook (PATH A) ---
        final_tpl_id = meraki_get(f"/networks/{network_id}").get('configTemplateId')
        final_mx, final_ms, final_mr = fetch_devices(org_id, network_id, template_id=final_tpl_id)
        final_vlans = fetch_vlan_details(network_id)
        profileid_to_name_post: Dict[str, str] = {}
        if final_tpl_id:
            try:
                final_profiles = meraki_get(f"/organizations/{org_id}/configTemplates/{final_tpl_id}/switch/profiles") or []
                profileid_to_name_post = {p['switchProfileId']: p['name'] for p in final_profiles}
            except Exception:
                logging.exception("Failed fetching final template switch profiles")

        export_combined_snapshot_xlsx(
            org_id=org_id, network_id=network_id, network_name=network_name,
            pre_template_id=pre_change_template,
            pre_vlan_list=pre_change_vlans,
            pre_mx_list=mx,
            pre_ms_list=ms,
            pre_mr_list=mr,
            pre_profileid_to_name=old_profileid_to_name,
            post_template_id=final_tpl_id,
            post_vlan_list=final_vlans,
            post_mx_list=final_mx,
            post_ms_list=final_ms,
            post_mr_list=final_mr,
            post_profileid_to_name=profileid_to_name_post,
            outfile=f"{_slug_filename(_network_tag_from_name(network_name))}_combined_{timestamp}.xlsx",
        )
        print_vlan_update_summary()
        print_debug_json_log_summary()
        
        write_run_summary_json(
            org_id=org_id,
            network_id=network_id,
            network_name=network_name,
            step_status=step_status,
            claimed_serials=claimed_serials_rb,
            removed_serials=removed_serials_rb,
            extra={"path": "A"},
        )
     
        maybe_prompt_and_rollback(
            org_id, network_id,
            pre_change_devices, pre_change_vlans, pre_change_template,
            ms, network_name,
            claimed_serials=claimed_serials_rb, removed_serials=removed_serials_rb
        )
        
        print_api_error_log_summary()
        
        try:
            if _current_checkpoint:
                os.remove(_current_checkpoint.path())
                print("🧹 Cleared resume checkpoint.")
        except Exception:
            pass

        raise SystemExit(0)

    # ------------------------------------------------------------------
    # PATH B: MX64 present -> full rebind/VLAN flow
    # ------------------------------------------------------------------
    vlan_list = fetch_vlan_details(network_id)
    old_mx, prebind_ms_devices, old_mr = fetch_devices(org_id, network_id, template_id=old_template)
    ms_serial_to_profileid: Dict[str, Optional[str]] = {sw['serial']: sw.get('switchProfileId') for sw in prebind_ms_devices}
    prebind_overrides_by_serial: Dict[str, Dict[str, Any]] = {
        sw['serial']: (sw.get('port_overrides') or {}) for sw in prebind_ms_devices
    }

       # Choose & (re)bind template (with rollback on failure)
    # --- Template selection / rebind (checkpointed & idempotent) ---
      # Default: assume we're NOT in the CloudStore→PreSCE flow
    cloudstore_presce: bool = False

    # If we have a checkpoint with a stored flag, restore it
    if _current_checkpoint is not None and isinstance(_current_checkpoint.step_status, dict):
        cloudstore_presce = bool(_current_checkpoint.step_status.get('cloudstore_presce_flow'))

    if not _current_checkpoint.done('template_bound'):
        try:
            new_template, new_tpl_name, rollback_needed, cloudstore_presce = list_and_rebind_template(
                org_id=org_id,
                network_id=network_id,
                current_id=old_template,
                network_name=network_name,
                pre_change_devices=pre_change_devices,
                pre_change_vlans=pre_change_vlans,
                pre_change_template=pre_change_template,
                claimed_serials=_current_checkpoint.claimed_serials or [],
                removed_serials=_current_checkpoint.removed_serials or [],
                ms_list=prebind_ms_devices,
                mx_model_filter=mx_model_filter,
            )

            # If the inner function already rolled back and decided to bail,
            # it may have cleared _current_checkpoint. Don't touch it, just exit.
            if rollback_needed:
                print("ℹ️ Rollback requested during template bind. Exiting workflow.")
                raise SystemExit(0)

            step_status['template_bound'] = bool(new_template and new_template != old_template)

            if _current_checkpoint is not None:
                _current_checkpoint.bound_template_id = new_template or old_template
                _current_checkpoint.mark('template_bound', step_status['template_bound'])

                # Persist the CloudStore→PreSCE flow flag for later steps
                _current_checkpoint.step_status = _current_checkpoint.step_status or {}
                _current_checkpoint.step_status['cloudstore_presce_flow'] = cloudstore_presce
                _current_checkpoint.save()
            else:
                logging.warning(
                    "template_bound: _current_checkpoint was None after list_and_rebind_template; "
                    "cannot persist template state."
                )

        except SystemExit:
            # Let explicit SystemExit bubble up cleanly
            raise

        except Exception:
            logging.exception("Template bind step failed")
            step_status['template_bound'] = False

            if _current_checkpoint is not None:
                _current_checkpoint.mark('template_bound', False)
                # Fall back to whatever we last knew, or the original template
                new_template = _current_checkpoint.bound_template_id or old_template
            else:
                # Checkpoint already cleared (e.g., rollback). Just fall back to old template.
                new_template = old_template
    else:
        print("⏭️  Skipping template selection/bind (already completed).")
        new_template = _current_checkpoint.bound_template_id or old_template
        # cloudstore_presce already restored from checkpoint above

    # VLAN update after rebind
    if not _current_checkpoint.done('vlans_updated'):
        try:
            flatten_requested = False
            if _current_checkpoint is not None and isinstance(_current_checkpoint.step_status, dict):
                flatten_requested = bool(_current_checkpoint.step_status.get("flatten_requested", False))

            bind_network_to_template(
                org_id=org_id,
                network_id=network_id,
                tpl_id=new_template,
                vlan_list=pre_change_vlans,
                network_name=network_name,
                pre_change_devices=pre_change_devices,
                pre_change_vlans=pre_change_vlans,
                pre_change_template=pre_change_template,
                claimed_serials=(_current_checkpoint.claimed_serials or []) if _current_checkpoint else [],
                removed_serials=(_current_checkpoint.removed_serials or []) if _current_checkpoint else [],
                ms_list=prebind_ms_devices,
                flatten_requested=flatten_requested,
            )

            step_status['vlans_updated'] = True
            _current_checkpoint.mark('vlans_updated', True)
        except Exception:
            logging.exception("VLAN update step failed after template bind")
            step_status['vlans_updated'] = False
            _current_checkpoint.mark('vlans_updated', False)
    else:
        print("⏭️  Skipping VLAN update (already completed).")

    # Fetch new template profiles for post-bind MS mapping
    try:
        tpl_profiles = meraki_get(
            f"/organizations/{org_id}/configTemplates/{new_template}/switch/profiles"
        ) if new_template else []
        tpl_profile_map = {p['name']: p['switchProfileId'] for p in (tpl_profiles or [])}
    except Exception:
        logging.exception("Failed fetch template switch profiles")
        tpl_profile_map = {}
        tpl_profiles = []

    # Re-assign switch profiles to match previous names / user choice
    _, postbind_ms_devices, _ = fetch_devices(org_id, network_id, template_id=new_template)

    # --- CloudStore→PreSCE flow: ask ONCE which profile to use, using
    #     the first MS as the representative model (24 vs 48).
    network_level_profile_id: Optional[str] = None
    if cloudstore_presce and tpl_profiles and postbind_ms_devices:
        example_model = postbind_ms_devices[0].get("model")
        network_level_profile_id = select_switch_profile_for_network(
            tpl_profiles,
            example_switch_model=example_model,
        )


        # Track per-switch attempts + failures for summary JSON
    sw_profile_attempts: Dict[str, int] = {}
    sw_profile_failures: Dict[str, Dict[str, Any]] = {}

    for sw in postbind_ms_devices:
        serial = str(sw.get("serial") or "")
        if not serial:
            continue

        old_profile_id = ms_serial_to_profileid.get(serial)
        old_profile_name = old_profileid_to_name.get(old_profile_id) if isinstance(old_profile_id, str) else None

        # Decide which profile to apply
        chosen_profile_id: Optional[str] = None

        if network_level_profile_id:
            chosen_profile_id = network_level_profile_id
            # Compatibility check (best-effort)
            if not any(
                _profile_supports_switch_model(p, str(sw.get("model") or ""))
                for p in tpl_profiles
                if str(p.get("switchProfileId") or "") == str(chosen_profile_id)
            ):
                chosen_profile_id = None

        if not chosen_profile_id:
            chosen_profile_id = tpl_profile_map.get(old_profile_name) if old_profile_name else None

        if not chosen_profile_id and tpl_profiles:
            chosen_profile_id = select_switch_profile_interactive_by_model(
                tpl_profiles, tpl_profile_map, str(sw.get("model") or "")
            )

        if not chosen_profile_id:
            logging.warning("No switch profile chosen for %s (%s); skipping", serial, sw.get("model"))
            continue

        def _op_assign_profile() -> None:
            do_action(meraki_put, f"/devices/{serial}", data={"switchProfileId": chosen_profile_id})

        outcome = retry_with_backoff(
            f"assign_switch_profile_{serial}",
            _op_assign_profile,
            max_attempts=5,
            base_sleep=1.0,
            max_sleep=10.0,
            verbose=True,
        )
        sw_profile_attempts[serial] = outcome.attempts

        if outcome.ok:
            log_change(
                "switch_profile_assign",
                f"Assigned switchProfileId {chosen_profile_id} to {serial}",
                device_serial=serial,
                device_name=str(sw.get("name") or ""),
                network_id=network_id,
                network_name=network_name,
                misc=json.dumps({
                    "attempts": outcome.attempts,
                    "profile_name_previous": old_profile_name or "",
                    "profile_id_assigned": chosen_profile_id,
                    "switch_model": sw.get("model"),
                }),
            )

            # Port overrides: best-effort, but still retry individual ports at apply_port_overrides level if you want
            try:
                preserved = prebind_overrides_by_serial.get(serial) or (sw.get("port_overrides") or {})
                if preserved:
                    apply_port_overrides(serial, preserved)
            except Exception as e:
                logging.exception("Failed applying port overrides for %s", serial)
                sw_profile_failures[serial] = {
                    "stage": "apply_port_overrides",
                    "error_type": type(e).__name__,
                    "error": str(e),
                }
        else:
            logging.error("Failed assigning switch profile for %s after %d attempts: %s",
                          serial, outcome.attempts, outcome.error)
            sw_profile_failures[serial] = {
                "stage": "assign_profile",
                **asdict(outcome),
                "switch_model": sw.get("model"),
                "profile_id": chosen_profile_id,
                "previous_profile_name": old_profile_name,
            }

    # Optionally write a JSON summary for this sub-step (useful when PreSCE mapping goes wrong)
    try:
        path = os.path.join("logs", f"switch_profile_assignment_{org_id}_{network_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        _write_json_log(path, {
            "timestamp": datetime.now().isoformat(),
            "org_id": org_id,
            "network_id": network_id,
            "network_name": network_name,
            "attempts_by_switch": sw_profile_attempts,
            "failures_by_switch": sw_profile_failures,
        })
        if DEBUG_MODE:
            print(f"📝 Switch profile assignment summary saved: {path}")
    except Exception:
        logging.exception("Failed writing switch profile assignment summary JSON")

    if sw_profile_failures:
        print("\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        print("⚠️ Switch Profile Assignment Issues")
        print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        for ser, info in sw_profile_failures.items():
            print(f"- {ser}: stage={info.get('stage')} attempts={info.get('attempts')} http={info.get('http_status')} err={info.get('error')}")
        print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
  
    # Wireless pre-check + claim
    safe_to_claim, mr_removed_serials, mr_claimed_serials = run_wireless_precheck_and_filter_claims(
        org_id, network_id, prevalidated_serials
    )
    if not _current_checkpoint.done('devices_claimed'):
        claimed = claim_devices(org_id, network_id, prevalidated_serials=safe_to_claim)
        step_status['devices_claimed'] = bool(claimed)
        _current_checkpoint.claimed_serials = claimed or _current_checkpoint.claimed_serials or []
        _current_checkpoint.mark('devices_claimed', step_status['devices_claimed'])
    else:
        print("⏭️  Skipping device claim (already completed).")
        claimed = _current_checkpoint.claimed_serials or []

    # Enable WAN2
    if claimed and not _current_checkpoint.done('wan2_enabled'):
        safe_enable_wan2_on_claimed_mx(org_id, claimed)
        _current_checkpoint.mark('wan2_enabled', True)

    
    #  Primary / order (idempotent with checkpoint) ---

    # Primary selection + swap
    if not _current_checkpoint.done('primary_selected'):
        primary_mx_serial = select_primary_mx(org_id, claimed)
        ensure_primary_mx(network_id, primary_mx_serial)
        _current_checkpoint.primary_mx_serial = primary_mx_serial or _current_checkpoint.primary_mx_serial
        _current_checkpoint.mark('primary_selected', True)
    else:
        print("⏭️  Skipping primary MX selection (already completed).")
        primary_mx_serial = _current_checkpoint.primary_mx_serial

    # MR ordering
    if not _current_checkpoint.done('mr_order'):
        mr_order = select_device_order(org_id, claimed, 'MR')
        _current_checkpoint.mr_order = mr_order or _current_checkpoint.mr_order or []
        _current_checkpoint.mark('mr_order', True)
    else:
        print("⏭️  Skipping MR order (already completed).")
        mr_order = _current_checkpoint.mr_order or []

    # MS ordering
    if not _current_checkpoint.done('ms_order'):
        ms_order = select_device_order(org_id, claimed, 'MS')
        _current_checkpoint.ms_order = ms_order or _current_checkpoint.ms_order or []
        _current_checkpoint.mark('ms_order', True)
    else:
        print("⏭️  Skipping MS order (already completed).")
        ms_order = _current_checkpoint.ms_order or []


    # Compute deltas for rollback (after all device changes)
    post_change_devices_raw = meraki_get(f"/networks/{network_id}/devices")
    post_change_devices = post_change_devices_raw if isinstance(post_change_devices_raw, list) else []
    post_change_serials = {d.get('serial') for d in post_change_devices if d.get('serial')}
    claimed_serials_rb = list(post_change_serials - pre_change_serials)
    removed_serials_rb = list(pre_change_serials - post_change_serials)
    _current_checkpoint.claimed_serials = claimed_serials_rb
    _current_checkpoint.removed_serials = removed_serials_rb
    _current_checkpoint.save()



    if claimed:
        new_mx, ms_list, mr_list = fetch_devices(org_id, network_id)
        step_status['old_mx'] = bool([d['serial'] for d in old_mx])
        step_status['old_mr33'] = bool([d['serial'] for d in old_mr if d['model'] == 'MR33'])

    # Legacy removals (checkpointed & idempotent) ---

        # Cache claimed models once for both MX64/MR33 decisions
        if _current_checkpoint.claimed_models is None:
            try:
                _current_checkpoint.claimed_models = _get_inventory_models_for_serials(org_id, claimed)
            except Exception:
                logging.exception("Failed to read claimed models for legacy removal checks")
                _current_checkpoint.claimed_models = {}

        # Always coerce to a dict for type safety
        claimed_models: Dict[str, str] = _current_checkpoint.claimed_models or {}

        # MX64 removal (only if a newer MX was actually claimed)
        if not _current_checkpoint.done('mx64_removed'):
            try:
                newer_claimed = any(
                    (m or "").startswith(("MX67", "MX75"))
                    for m in claimed_models.values()
                )
                if newer_claimed:
                    ok = remove_existing_mx64_devices(org_id, network_id)
                    step_status['mx_removed'] = ok
                    if ok:
                        log_change(
                            'mx_removed',
                            "Removed old MX64 after new MX claim",
                            misc=f"claimed_serials={claimed}"
                        )
                else:
                    step_status['mx_removed'] = "NA"
                _current_checkpoint.mark('mx64_removed', True)
            except Exception:
                logging.exception("MX64 removal stage failed")
                step_status['mx_removed'] = False
        else:
            print("⏭️  Skipping MX64 removal (already completed).")

        # MR33 removal (only if any wireless device was claimed this run)
        if not _current_checkpoint.done('mr33_removed'):
            try:
                claimed_has_wireless = any(
                    _is_wireless_model(m) for m in claimed_models.values()
                )
                if claimed_has_wireless:
                    ok = remove_existing_mr33_devices(org_id, network_id)
                    step_status['mr33_removed'] = ok
                    if ok:
                        log_change(
                            'mr33_removed',
                            "Removed old MR33 after new AP claim",
                            misc=f"claimed_serials={claimed}"
                        )
                else:
                    step_status['mr33_removed'] = "NA"
                _current_checkpoint.mark('mr33_removed', True)
            except Exception:
                logging.exception("MR33 removal stage failed")
                step_status['mr33_removed'] = False
        else:
            print("⏭️  Skipping MR33 removal (already completed).")
   
        # Naming & configuration for claimed devices
        try:
            name_and_configure_claimed_devices(
                org_id=org_id,
                network_id=network_id,
                network_name=network_name,
                serials=claimed,
                ms_list=ms_list,
                tpl_profile_map=tpl_profile_map,
                old_mx_devices=old_mx,
                old_mr_devices=old_mr,
                primary_mx_serial=primary_mx_serial,
                mr_order=mr_order,
                ms_order=ms_order,
            )
            remove_recently_added_tag(network_id)
            step_status['configured'] = True
        except Exception:
            logging.exception("Configuration of claimed devices failed")
            step_status['configured'] = False
    else:
        step_status.setdefault('mx_removed', "NA")
        step_status.setdefault('mr33_removed', "NA")
        step_status.setdefault('configured', "NA")
        step_status.setdefault('old_mx', "NA")
        step_status.setdefault('old_mr33', "NA")

    print_summary(step_status)

    # --- Build POST state & export one combined workbook (PATH A) ---
    final_tpl_id = meraki_get(f"/networks/{network_id}").get('configTemplateId')
    final_mx, final_ms, final_mr = fetch_devices(org_id, network_id, template_id=final_tpl_id)
    final_vlans = fetch_vlan_details(network_id)
    profileid_to_name_post: Dict[str, str] = {}
    if final_tpl_id:
        try:
            final_profiles = meraki_get(f"/organizations/{org_id}/configTemplates/{final_tpl_id}/switch/profiles") or []
            profileid_to_name_post = {p['switchProfileId']: p['name'] for p in final_profiles}
        except Exception:
            logging.exception("Failed fetching final template switch profiles")

    if not _current_checkpoint.done('snapshot_exported'):    
        export_combined_snapshot_xlsx(
            org_id=org_id, network_id=network_id, network_name=network_name,
            pre_template_id=pre_change_template,
            pre_vlan_list=pre_change_vlans,
            pre_mx_list=mx,
            pre_ms_list=ms,
            pre_mr_list=mr,
            pre_profileid_to_name=old_profileid_to_name,
            post_template_id=final_tpl_id,
            post_vlan_list=final_vlans,
            post_mx_list=final_mx,
            post_ms_list=final_ms,
            post_mr_list=final_mr,
            post_profileid_to_name=profileid_to_name_post,
            outfile=f"{_slug_filename(_network_tag_from_name(network_name))}_combined_{timestamp}.xlsx",
        )
        _current_checkpoint.mark('snapshot_exported', True)
    else:
        print("⏭️  Skipping snapshot export (already completed).")

    print_vlan_update_summary()
    print_api_error_log_summary()
    print_debug_json_log_summary()
    
    write_run_summary_json(
        org_id=org_id,
        network_id=network_id,
        network_name=network_name,
        step_status=step_status,
        claimed_serials=claimed_serials_rb,
        removed_serials=removed_serials_rb,
        extra={"path": "B", "cloudstore_presce_flow": bool(cloudstore_presce)},
    )
 
    # -------- Enhanced rollback prompt (extracted) --------
    maybe_prompt_and_rollback(
        org_id, network_id,
        pre_change_devices, pre_change_vlans, pre_change_template,
        ms, network_name,
        claimed_serials=claimed_serials_rb, removed_serials=removed_serials_rb
    )
    

