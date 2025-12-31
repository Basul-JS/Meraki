#!/usr/bin/env python3
# dryrun_network_workflow.py

import requests
import logging
import re
import json
from getpass import getpass
from datetime import datetime
import csv
import os
import time
import signal
import sys
from typing import Any, Dict, List, Optional, Tuple, Set, Union, cast
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
import unicodedata
from difflib import SequenceMatcher

# =====================
# Config & Constants
# =====================
EXCLUDED_VLANS = {100, 110, 210, 220, 230, 235, 240}
REQUEST_TIMEOUT = 30  # seconds
BASE_URL = "https://api.meraki.com/api/v1"
MAX_RETRIES = 5
DRY_RUN = True  # <-- Hard-wired: this file is DRY-RUN ONLY

# Logging setup
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
logging.basicConfig(
    filename=f"meraki_script_dryrun_{timestamp}.log",
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
CSV_LOGFILE = f"meraki_dryrun_audit_{timestamp}.csv"

# =====================
# Utility: CSV audit log (messages only; no mutations happen)
# =====================
def log_change(
    event: str,
    details: str,
    *,
    username: Optional[str] = None,
    device_serial: Optional[str] = None,
    device_name: Optional[str] = None,
    misc: Optional[str] = None,
    org_id: Optional[str] = None,
    org_name: Optional[str] = None,
    network_id: Optional[str] = None,
    network_name: Optional[str] = None,
) -> None:
    file_exists = os.path.isfile(CSV_LOGFILE)
    with open(CSV_LOGFILE, mode='a', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        if not file_exists:
            writer.writerow([
                'timestamp', 'event', 'details', 'user',
                'device_serial', 'device_name', 'misc',
                'org_id', 'org_name', 'network_id', 'network_name'
            ])
        writer.writerow([
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            event,
            details,
            username or "<dryrun>",
            device_serial or '',
            device_name or '',
            misc or '',
            org_id or '',
            org_name or '',
            network_id or '',
            network_name or ''
        ])

# =====================
# Auth (read-only)
# =====================
def validate_api_key(key: str) -> bool:
    return bool(re.fullmatch(r'[A-Fa-f0-9]{40}', key or ''))

API_KEY: Optional[str] = None
for _ in range(4):
    API_KEY = getpass("Enter your Meraki API key (hidden): ")
    if validate_api_key(API_KEY):
        break
else:
    print("❌ Invalid API key after 4 attempts")
    sys.exit(1)


HEADERS = {
    "X-Cisco-Meraki-API-Key": API_KEY,
    "Content-Type": "application/json",
    "Accept": "application/json",
}

# Graceful abort
_aborted = False
def _handle_sigint(signum, frame):
    global _aborted
    _aborted = True
    print("\nReceived Ctrl+C — attempting graceful shutdown...")
    log_change('workflow_abort', 'User interrupted with SIGINT')
signal.signal(signal.SIGINT, _handle_sigint)

# =====================
# HTTP (GET only in practice; mutations are simulated)
# =====================
class MerakiAPIError(Exception):
    def __init__(self, status_code: int, text: str, json_body: Optional[Any], url: str):
        super().__init__(f"Meraki API error: {status_code} {text}")
        self.status_code = status_code
        self.text = text
        self.json_body = json_body
        self.url = url

def _request(method: str, path: str, *, params=None, json_data=None) -> Any:
    url = f"{BASE_URL}{path}"
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            if method == 'GET':
                resp = requests.get(url, headers=HEADERS, params=params, timeout=REQUEST_TIMEOUT)
            elif method in ('POST', 'PUT', 'DELETE'):
                # Dry run: never actually call mutating endpoints
                logging.info("DRY-RUN: would %s %s, json=%s", method, url, json.dumps(json_data, ensure_ascii=False))
                return {"dryRun": True, "method": method, "url": url, "payload": json_data}
            else:
                raise ValueError("Unknown HTTP method")

            if resp.status_code == 429:
                ra = resp.headers.get("Retry-After")
                wait = min(float(ra) if ra else 2 ** (attempt - 1), 30.0)
                logging.warning(f"429 rate limit for {url}. Sleeping {wait}s and retrying...")
                time.sleep(wait)
                continue

            if not resp.ok:
                try:
                    body = resp.json()
                except Exception:
                    body = None
                logging.error(f"{method} {url} -> {resp.status_code} {resp.text}")
                raise MerakiAPIError(resp.status_code, resp.text, body, url)

            if resp.text:
                try:
                    return resp.json()
                except Exception:
                    return resp.text
            return None
        except MerakiAPIError:
            raise
        except Exception as e:
            if attempt == MAX_RETRIES:
                logging.exception(f"HTTP error for {url}: {e}")
                raise
            wait = min(2 ** attempt, 30)
            logging.warning(f"HTTP exception {e} for {url}. Retrying in {wait}s...")
            time.sleep(wait)

def meraki_get(path, params=None):
    return _request('GET', path, params=params)

def meraki_post(path, data=None):
    return _request('POST', path, json_data=data)

def meraki_put(path, data=None):
    return _request('PUT', path, json_data=data)

def meraki_delete(path):
    return _request('DELETE', path)

def do_action(func, *args, **kwargs):
    # In this dry-run script, do_action never mutates; it just logs the intent.
    if func in (meraki_post, meraki_put, meraki_delete):
        logging.info("DRY-RUN: would call %s args=%s kwargs=%s", getattr(func, '__name__', str(func)), args, kwargs)
        return {"dryRun": True, "fn": getattr(func, '__name__', ''), "args": args, "kwargs": kwargs}
    return func(*args, **kwargs)

# =====================
# Helpers (normalization, search)
# =====================
def _norm(s: Optional[str]) -> str:
    base: str = s or ""
    base = unicodedata.normalize("NFKC", base)
    base = base.replace("–", "-").replace("—", "-")
    base = re.sub(r"\s+", " ", base).strip()
    return base.casefold()

def _similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a, b).ratio()

def meraki_list_networks_all(org_id: str) -> List[Dict[str, Any]]:
    all_nets: List[Dict[str, Any]] = []
    per_page: int = 1000
    starting_after: Optional[str] = None
    while True:
        params: Dict[str, Any] = {"perPage": per_page}
        if starting_after:
            params["startingAfter"] = starting_after
        page_raw: Any = meraki_get(f"/organizations/{org_id}/networks", params=params)
        page: List[Dict[str, Any]] = page_raw if isinstance(page_raw, list) else []
        if not page:
            break
        all_nets.extend(page)
        if len(page) < per_page:
            break
        last = page[-1]
        starting_after = str(last.get("id") or "")
        if not starting_after:
            break
    return all_nets

def fetch_matching_networks(org_id: str, partial: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    nets: List[Dict[str, Any]] = meraki_list_networks_all(org_id)
    partial_n: str = _norm(partial)
    matches: List[Dict[str, Any]] = [n for n in nets if partial_n in _norm(n.get("name"))]
    if matches:
        return matches, []
    scored: List[Tuple[Dict[str, Any], float]] = [(n, _similarity(partial_n, _norm(n.get("name")))) for n in nets]
    scored.sort(key=lambda t: t[1], reverse=True)
    suggestions: List[Dict[str, Any]] = [n for (n, score) in scored[:5] if score >= 0.6]
    return [], suggestions

# =====================
# Inventory / VLAN / Devices
# =====================
def fetch_vlan_details(network_id: str) -> List[Dict[str, Any]]:
    try:
        vlans = meraki_get(f"/networks/{network_id}/appliance/vlans")
        filtered = [v for v in vlans if int(v.get('id')) not in EXCLUDED_VLANS]
        logging.debug(f"Fetched VLANs: {len(filtered)} (excluded {len(vlans) - len(filtered)})")
        return filtered
    except MerakiAPIError as e:
        # In dry-run, just report
        logging.exception("Failed to fetch VLANs: %s %s", e.status_code, e.text)
        return []
    except Exception:
        logging.exception("Failed to fetch VLANs")
        return []

def fetch_devices(org_id: str, network_id: str, template_id: Optional[str] = None) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    devs = meraki_get(f"/networks/{network_id}/devices") or []

    def _mk(d):
        tags = d.get('tags', [])
        if not isinstance(tags, list):
            tags = (tags or '').split()
        return {
            'serial': d['serial'],
            'model': d['model'],
            'tags': tags,
            'address': d.get('address', ''),
            'name': d.get('name', ''),
            'switchProfileId': d.get('switchProfileId'),
            'switchProfileName': d.get('switchProfileName'),
        }

    mx = [_mk(d) for d in devs if str(d.get('model','')).startswith('MX')]
    ms = [_mk(d) for d in devs if str(d.get('model','')).startswith('MS')]
    # Treat MR/CW wireless as MR list
    mr = [_mk(d) for d in devs if str(d.get('model','')).upper().startswith(("MR","CW"))]

    # If a template is known, compute port diffs (for reporting only)
    if template_id:
        for sw in ms:
            profile_id = sw.get('switchProfileId')
            if not profile_id:
                sw['port_overrides'] = {}
                continue
            try:
                live_ports = meraki_get(f"/devices/{sw['serial']}/switch/ports") or []
                tmpl_ports = meraki_get(f"/organizations/{org_id}/configTemplates/{template_id}/switch/profiles/{profile_id}/ports") or []
                sw['port_overrides'] = compute_port_overrides(live_ports, tmpl_ports)
            except Exception:
                logging.exception("Failed computing port overrides for %s", sw.get('serial'))
                sw['port_overrides'] = {}
    else:
        for sw in ms:
            sw['port_overrides'] = {}

    return mx, ms, mr

def vlans_enabled(network_id: str) -> Optional[bool]:
    try:
        settings = meraki_get(f"/networks/{network_id}/appliance/vlans/settings")
        return bool(settings.get("vlansEnabled"))
    except Exception:
        logging.exception("Could not read VLANs settings")
        return None

# =====================
# Port diff helpers (unchanged)
# =====================
_PORT_FIELDS = [
    "enabled", "name", "tags", "type", "vlan", "voiceVlan", "allowedVlans",
    "poeEnabled", "isolationEnabled", "rstpEnabled", "stpGuard",
    "linkNegotiation", "udld", "accessPolicyType", "accessPolicyNumber",
    "portScheduleId"
]

def _normalize_tags(value):
    if isinstance(value, list):
        return sorted(value)
    if isinstance(value, str):
        return sorted([t for t in value.split() if t])
    return []

def _port_dict_by_number(ports: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    for p in ports:
        pid = p.get("portId") or p.get("number") or p.get("name")
        if pid is None:
            continue
        out[str(pid)] = p
    return out

def compute_port_overrides(live_ports: List[Dict[str, Any]], tmpl_ports: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    overrides: Dict[str, Dict[str, Any]] = {}
    live = _port_dict_by_number(live_ports)
    tmpl = _port_dict_by_number(tmpl_ports)
    for pid, lp in live.items():
        tp = tmpl.get(pid)
        if not tp:
            continue
        for fld in _PORT_FIELDS:
            lv = lp.get(fld)
            tv = tp.get(fld)
            if fld == "tags":
                lv = _normalize_tags(lv)
                tv = _normalize_tags(tv)
            if lv is not None and lv != tv:
                overrides.setdefault(pid, {})[fld] = lv
    return overrides

# =====================
# Export snapshot (supports planned VLAN payloads)
# =====================
def _slug_filename(s: str) -> str:
    s = re.sub(r'[^A-Za-z0-9._-]+', '-', s).strip('-_')
    return s[:80]

def _network_tag_from_name(name: str) -> str:
    parts = (name or "").split('-')
    if len(parts) >= 2 and parts[1].isdigit():
        return f"{parts[0]}-{parts[1]}"
    return name or "network"

def _network_number_from_name(name: str) -> Optional[str]:
    m = re.search(r'\b(\d{2,8})\b', name or "")
    return m.group(1) if m else None

def export_network_snapshot_xlsx(
    org_id: str,
    network_id: str,
    network_name: str,
    template_id: Optional[str],
    vlan_list: List[Dict[str, Any]],
    mx_list: List[Dict[str, Any]],
    ms_list: List[Dict[str, Any]],
    mr_list: List[Dict[str, Any]],
    profileid_to_name: Optional[Dict[str, str]] = None,
    outfile: Optional[str] = None,
    filename_mode: str = "name",
    *,
    planned_vlan_payloads: Optional[Dict[str, Dict[str, Any]]] = None,  # <-- NEW
) -> None:
    def _json(x: Any) -> str:
        try:
            return json.dumps(x, ensure_ascii=False)
        except Exception:
            return str(x)

    if outfile:
        out_path: str = outfile
    else:
        if filename_mode == "number":
            base = _network_number_from_name(network_name) or _network_tag_from_name(network_name)
        else:
            base = _network_tag_from_name(network_name)
        out_path = f"{_slug_filename(base)}_{timestamp}.xlsx"

    wb: Workbook = Workbook()
    ws: Worksheet = cast(Worksheet, wb.active)
    ws.title = "Snapshot"

    header: List[str] = [
        "section", "network_id", "network_name", "item_type",
        "col1", "col2", "col3", "col4", "col5",
        "switch_profile_id", "switch_profile_name", "extra_info"
    ]
    ws.append(header)

    tpl_name: str = ""
    if template_id:
        try:
            tpl = meraki_get(f"/organizations/{org_id}/configTemplates/{template_id}")
            tpl_name = str(tpl.get("name", "") or "")
        except Exception:
            logging.exception("Could not fetch template name for snapshot")

    ws.append(["template", network_id, network_name, "template",
               template_id or "", tpl_name, "", "", "", "", "", ""])

    # VLANs (live)
    for v in vlan_list:
        ws.append([
            "vlans", network_id, network_name, "vlan",
            str(v.get("id", "")),
            str(v.get("name", "") or ""),
            str(v.get("subnet", "") or ""),
            str(v.get("applianceIp", "") or ""),
            str(v.get("dhcpHandling", "") or ""),
            "", "",
            _json({k: v.get(k) for k in v.keys() - {"id", "name", "subnet", "applianceIp", "dhcpHandling"}}),
        ])

    # Devices
    def _device_row(d: Dict[str, Any]) -> List[str]:
        tags_val = d.get("tags", [])
        if isinstance(tags_val, list):
            tags_list = [str(t) for t in tags_val]
        else:
            tags_list = [t for t in str(tags_val or "").split() if t]

        sp_id: str = str(d.get("switchProfileId", "") or "")
        sp_name: str = str(d.get("switchProfileName", "") or "")
        if (not sp_name) and sp_id and profileid_to_name:
            sp_name = profileid_to_name.get(sp_id, "") or ""

        return [
            "devices", network_id, network_name, "device",
            str(d.get("serial", "") or ""),
            str(d.get("model", "") or ""),
            str(d.get("name", "") or ""),
            str(d.get("address", "") or ""),
            " ".join(tags_list),
            sp_id,
            sp_name,
            ""
        ]

    for d in (mx_list + ms_list + mr_list):
        ws.append(_device_row(d))

    # MS port overrides (if any)
    for sw in ms_list:
        changes_by_port: Dict[str, Dict[str, Any]] = sw.get("port_overrides") or {}
        if not isinstance(changes_by_port, dict) or not changes_by_port:
            continue
        for port_id, changes in changes_by_port.items():
            if not isinstance(changes, dict):
                continue
            for fld, val in changes.items():
                ws.append([
                    "port_overrides", network_id, network_name, "port_override",
                    str(sw.get("serial", "") or ""), str(port_id), str(fld),
                    "" if isinstance(val, (dict, list)) else str(val),
                    "", "", "",
                    json.dumps(val) if isinstance(val, (dict, list)) else "",
                ])

    # NEW: planned VLAN payloads (what we'd PUT)
    if planned_vlan_payloads:
        for vid, payload in planned_vlan_payloads.items():
            ws.append([
                "vlans_planned_payload", network_id, network_name, "vlan_payload",
                str(vid), "", "", "", "", "", "",
                json.dumps(payload, ensure_ascii=False),
            ])

    # autosize
    max_col: int = ws.max_column
    max_row: int = ws.max_row
    for col_idx in range(1, max_col + 1):
        max_len = 0
        for row_idx in range(1, max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                s = str(val)
                if len(s) > max_len:
                    max_len = len(s)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    wb.save(out_path)
    print(f"📄 Snapshot exported to Excel: {out_path}")
    log_change("snapshot_export", f"Exported snapshot to {out_path}",
               network_id=network_id, network_name=network_name)

# =====================
# VLAN handling rules (mimic prod, but return 'would-send' payload)
# =====================
def _dhcp_mode(val: Optional[str]) -> str:
    v = (val or "").strip().lower()
    if v in {"run a dhcp server", "run dhcp server", "server", "enabled", "on"}:
        return "server"
    if "relay" in v:
        return "relay"
    if v in {"do not respond", "do not respond to dhcp requests", "off", "disabled", "none"}:
        return "off"
    return "off"

def _nonempty(x: Any) -> bool:
    if x is None:
        return False
    if isinstance(x, (list, dict)) and len(x) == 0:
        return False
    if isinstance(x, str) and x.strip() == "":
        return False
    return True

def build_vlan_update_payloads(vlan_list: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """
    Return { vlan_id: payload_we_would_send } honoring DHCP mode rules.
    """
    out: Dict[str, Dict[str, Any]] = {}
    for v in vlan_list:
        vlan_id = str(v.get("id", ""))
        payload: Dict[str, Any] = {}
        if _nonempty(v.get("applianceIp")):
            payload["applianceIp"] = v.get("applianceIp")
        if _nonempty(v.get("subnet")):
            payload["subnet"] = v.get("subnet")
        if _nonempty(v.get("dhcpHandling")):
            payload["dhcpHandling"] = v.get("dhcpHandling")

        mode = _dhcp_mode(v.get("dhcpHandling"))
        if mode == "server":
            if _nonempty(v.get("fixedIpAssignments")):
                payload["fixedIpAssignments"] = v.get("fixedIpAssignments")
            if _nonempty(v.get("reservedIpRanges")):
                payload["reservedIpRanges"] = v.get("reservedIpRanges")
            if _nonempty(v.get("dnsNameservers")):
                payload["dnsNameservers"] = v.get("dnsNameservers")
        elif mode == "relay":
            relay_ips = v.get("dhcpRelayServerIps") or v.get("dhcpRelayServerIp")
            if _nonempty(relay_ips):
                payload["dhcpRelayServerIps"] = relay_ips
        # mode "off": intentionally omit fixed/reserved/relay

        out[vlan_id] = payload
    return out

# =====================
# Template picking heuristic (what-if)
# =====================
def _current_vlan_count(network_id: str) -> Optional[int]:
    vlans = fetch_vlan_details(network_id)
    return len(vlans) if isinstance(vlans, list) else None

def _pick_template_by_vlan_count(
    templates: List[Dict[str, Any]],
    vlan_count: Optional[int],
) -> Optional[Dict[str, Any]]:
    if vlan_count not in (3, 5):
        return None
    patterns: List[str]
    if vlan_count == 3:
        patterns = [r'NO\s*LEGACY.*MX\b']
    else:  # vlan_count == 5
        patterns = [r'3\s*X\s*DATA[_\s-]*VLAN.*MX75\b']
    for pat in patterns:
        rx = re.compile(pat, re.IGNORECASE)
        for t in templates:
            name = (t.get('name') or '')
            if rx.search(name):
                return t
    return None

# =====================
# Simulation builders
# =====================
def build_simulated_post_devices_same_as_pre(
    pre_mx: List[Dict[str, Any]],
    pre_ms: List[Dict[str, Any]],
    pre_mr: List[Dict[str, Any]],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    For dry-run we keep devices the same; we only log what would be renamed or reassigned.
    """
    return [dict(d) for d in pre_mx], [dict(d) for d in pre_ms], [dict(d) for d in pre_mr]

def simulate_switch_profile_reapplication(
    org_id: str,
    new_template_id: Optional[str],
    ms_list: List[Dict[str, Any]],
    pre_profileid_to_name: Dict[str, str],
) -> Tuple[List[Dict[str, Any]], Dict[str, Dict[str, Dict[str, Any]]]]:
    """
    Returns:
      - ms_list_with_simulated_profileIds (unchanged in dry-run, but we set a 'simulatedNewProfileId' key for visibility)
      - simulated_port_override_patches_by_serial: { serial: { portId: patchDict } }
    """
    if not new_template_id or not ms_list:
        return ms_list, {}

    try:
        profiles = meraki_get(f"/organizations/{org_id}/configTemplates/{new_template_id}/switch/profiles") or []
        name_to_id = {p['name']: p['switchProfileId'] for p in profiles if 'name' in p and 'switchProfileId' in p}
    except Exception:
        logging.exception("Could not fetch new template profiles for simulation.")
        name_to_id = {}

    simulated_patches: Dict[str, Dict[str, Dict[str, Any]]] = {}
    out_ms: List[Dict[str, Any]] = []

    for sw in ms_list:
        sw_copy = dict(sw)
        old_profile_id = sw.get('switchProfileId')
        old_profile_name = pre_profileid_to_name.get(old_profile_id) if isinstance(old_profile_id, str) else None
        new_prof_id = name_to_id.get(old_profile_name) if old_profile_name else None

        if new_prof_id:
            sw_copy['simulatedNewProfileId'] = new_prof_id  # marker for the export/logs

        # Port overrides: reuse any precomputed diffs in sw['port_overrides']
        preserved = sw.get('port_overrides') or {}
        if preserved:
            simulated_patches[sw['serial']] = preserved  # what we'd PUT per port

        out_ms.append(sw_copy)

    return out_ms, simulated_patches

# =====================
# Org/Network pickers
# =====================
def select_org() -> str:
    orgs = meraki_get("/organizations") or []
    if not orgs:
        print("No organizations available.")
        sys.exit(1)

    print("Organizations:")
    for idx, org in enumerate(orgs, 1):
        print(f"{idx}. {org['name']} (ID: {org['id']})")

    raw = input("Select organization by number: ").strip()
    if not raw.isdigit() or not (1 <= int(raw) <= len(orgs)):
        print("Invalid selection.")
        sys.exit(1)
    return orgs[int(raw) - 1]['id']

def select_network_interactive(org_id: str) -> Tuple[str, str]:
    while True:
        partial = input("Enter partial network name to search: ").strip()
        matches, suggestions = fetch_matching_networks(org_id, partial)
        if matches:
            if len(matches) == 1:
                m = matches[0]
                print(f"Selected: {m['name']} ({m['id']})")
                return m['id'], m['name']
            print("\nMultiple matches:")
            for i, n in enumerate(matches, 1):
                print(f"{i}. {n['name']} (ID: {n['id']})")
            raw = input("Pick # : ").strip()
            if raw.isdigit() and 1 <= int(raw) <= len(matches):
                chosen = matches[int(raw)-1]
                return chosen['id'], chosen['name']
            print("Invalid choice.")
        else:
            print("No exact/substring matches found.")
            if suggestions:
                print("Did you mean:")
                for n in suggestions:
                    print(f" - {n['name']} (ID: {n['id']})")
        retry = input("Search again? (y/N): ").strip().lower()
        if retry != 'y':
            print("No network selected.")
            sys.exit(1)

# =====================
# Main (dry-run orchestration)
# =====================
if __name__ == '__main__':
    log_change('workflow_start', 'Dry-run script started')

    org_id = select_org()
    network_id, network_name = select_network_interactive(org_id)

    net_info = meraki_get(f"/networks/{network_id}") or {}
    old_template: Optional[str] = net_info.get('configTemplateId')

    # Pre-change snapshot (incl. MS port overrides vs template)
    mx, ms, mr = fetch_devices(org_id, network_id, template_id=old_template)
    pre_change_devices = mx + ms + mr
    pre_change_vlans = fetch_vlan_details(network_id)
    pre_change_template = old_template

    # Pre snapshot export (unchanged)
    old_profileid_to_name: Dict[str, str] = {}
    if old_template:
        try:
            old_tpl_profiles = meraki_get(f"/organizations/{org_id}/configTemplates/{old_template}/switch/profiles") or []
            old_profileid_to_name = {p['switchProfileId']: p['name'] for p in old_tpl_profiles if 'switchProfileId' in p}
        except Exception:
            logging.exception("Failed fetching old template switch profiles")

    export_network_snapshot_xlsx(
        org_id=org_id,
        network_id=network_id,
        network_name=network_name,
        template_id=old_template,
        vlan_list=pre_change_vlans,
        mx_list=mx,
        ms_list=ms,
        mr_list=mr,
        profileid_to_name=old_profileid_to_name,
        outfile=f"{_slug_filename(_network_tag_from_name(network_name))}_pre_{timestamp}.xlsx",
    )

    # -------------------------------
    # Template rebind WHAT-IF (dry-run)
    # -------------------------------
    vlan_count = _current_vlan_count(network_id)
    all_templates_raw: Any = meraki_get(f"/organizations/{org_id}/configTemplates") or []
    all_templates: List[Dict[str, Any]] = all_templates_raw if isinstance(all_templates_raw, list) else []
    suggested_tpl = _pick_template_by_vlan_count(all_templates, vlan_count)

    chosen_template_id = old_template  # default: keep same
    if suggested_tpl and suggested_tpl.get('id') != old_template:
        print(
            f"\nSuggestion (WHAT-IF): Based on VLAN count ({vlan_count}), "
            f"'{suggested_tpl.get('name','')}' looks appropriate (ID: {suggested_tpl.get('id','')})."
        )
        print("Dry-run: we will not bind, but will simulate as if we did.")
        chosen_template_id = suggested_tpl.get('id')

    # -------------------------------
    # VLAN updates: build would-send payloads (prod rules), but don't send
    # -------------------------------
    vlan_payloads = build_vlan_update_payloads(pre_change_vlans)
    print("\nVLAN update payloads (WHAT-IF):")
    for vid, payload in vlan_payloads.items():
        print(f" - VLAN {vid}: {json.dumps(payload, ensure_ascii=False)}")

    # -------------------------------
    # MS profile re-application + port overrides (simulated)
    # -------------------------------
    sim_ms, sim_port_patches = simulate_switch_profile_reapplication(
        org_id=org_id,
        new_template_id=chosen_template_id,
        ms_list=ms,
        pre_profileid_to_name=old_profileid_to_name
    )
    if sim_port_patches:
        print("\nPort override re-application (WHAT-IF):")
        for serial, patches in sim_port_patches.items():
            for port_id, patch in patches.items():
                print(f" - Would PUT /devices/{serial}/switch/ports/{port_id} with {json.dumps(patch, ensure_ascii=False)}")

    # -------------------------------
    # Build simulated POST snapshot
    # -------------------------------
    post_mx, post_ms, post_mr = build_simulated_post_devices_same_as_pre(mx, sim_ms, mr)
    simulated_template_id = chosen_template_id  # if suggestion active, reflect in snapshot

    # Export simulated POST snapshot (clearly labeled) + planned VLAN payloads
    export_network_snapshot_xlsx(
        org_id=org_id,
        network_id=network_id,
        network_name=network_name + " (SIMULATED POST)",
        template_id=simulated_template_id,
        vlan_list=pre_change_vlans,  # content unchanged; payloads above show what we'd modify
        mx_list=post_mx,
        ms_list=post_ms,
        mr_list=post_mr,
        profileid_to_name=old_profileid_to_name,  # names may still resolve
        outfile=f"{_slug_filename(_network_tag_from_name(network_name))}_post_SIMULATED_{timestamp}.xlsx",
        planned_vlan_payloads=vlan_payloads,  # <-- NEW
    )

    # -------------------------------
    # Rollback prompt (unchanged UX, but explains it’s a what-if)
    # -------------------------------
    print("\nRollback (WHAT-IF): If this were a live run and we had rebound/updated:")
    print(" - We would remove newly-claimed devices (none in dry-run).")
    print(" - We would re-claim previously removed devices (none in dry-run).")
    print(" - We would restore the original template binding "
          f"({pre_change_template or 'None'}) and re-apply preserved MS profiles/port overrides.")
    print(" - We would re-apply pre-change VLAN definitions.")

    # End
    log_change('workflow_end', 'Dry-run finished')
    print("\n✅ Dry-run complete. See console output and generated XLSX files for the simulated plan.")
