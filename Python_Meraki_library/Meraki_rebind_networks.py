# Created by J A Said
# Rebinds network to a New Template
# 20250903 - updated with new logic for MR33 presence
# 20250905 - updated to enable WAN2 on the new MX's
# 20250916 - updated to remove the block on MR33's been added
    # updated name of excel eports to correctly reference  pre and post 
    # removed redundant code
# 20250916 - Port override logic updated
# Python Dependencies
    # pip install meraki 
    # pip install openpyxl


import meraki
import logging
import re
import json
from datetime import datetime
from getpass import getpass
import csv
import os
import time
import signal
import sys
from typing import Any, Dict, List, Optional, Tuple, Set, Union, Iterable, cast
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ------------- Config & Constants -------------
EXCLUDED_VLANS: Set[int] = {100, 110, 210, 220, 230, 235, 240}
BASE_URL: str = "https://api.meraki.com/api/v1"

timestamp: str = datetime.now().strftime("%Y%m%d_%H%M%S")
logging.basicConfig(
    filename=f"meraki_script_{timestamp}.log",
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
CSV_LOGFILE: str = f"meraki_techboost25_rebind_{timestamp}.csv"

# ------------- CSV audit log -------------
def log_change(
    event: str,
    details: str,
    *,
    username: Optional[str] = None,
    device_serial: Optional[str] = None,
    device_name: Optional[str] = None,
    misc: Optional[str] = None,
    org_id: Optional[str] = None,
    org_name: Optional[str] = None,
    network_id: Optional[str] = None,
    network_name: Optional[str] = None,
) -> None:
    file_exists: bool = os.path.isfile(CSV_LOGFILE)
    with open(CSV_LOGFILE, mode='a', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        if not file_exists:
            writer.writerow([
                'timestamp', 'event', 'details', 'user',
                'device_serial', 'device_name', 'misc',
                'org_id', 'org_name', 'network_id', 'network_name'
            ])
        writer.writerow([
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            event,
            details,
            username or OPERATOR,
            device_serial or '',
            device_name or '',
            misc or '',
            org_id or '',
            org_name or '',
            network_id or '',
            network_name or ''
        ])

# ------------- Prompts / Safety -------------
OPERATOR: str = input("Enter your name or initials for audit logs: ")
DRY_RUN: bool = input("Run in dry-run mode? (yes/no): ").strip().lower() in {'yes', 'y'}
print(f"{'DRY RUN: ' if DRY_RUN else ''}Actions will {'not ' if DRY_RUN else ''}be executed.")

now = datetime.now()
cutoff_hour = 17
cutoff_minute = 40
if not DRY_RUN and ((now.hour < cutoff_hour) or (now.hour == cutoff_hour and now.minute < cutoff_minute)):
    print("\n" + "="*80)
    print("⚠️  WARNING: YOU ARE ABOUT TO MAKE LIVE CHANGES TO THE NETWORK ⚠️")
    print("This will bring down the network if applied during business hours.")
    print("Please Ensure the Store is closed before continuing.")
    print(f"Current time: {now.strftime('%H:%M')}")
    print("Recommended run time: AFTER 17:40.")
    print("="*80 + "\n")
    confirm: str = input("❗ Type 'YES' to proceed, or anything else to abort: ").strip()
    if confirm.upper() != "YES":
        print("❌ Aborting script.")
        raise SystemExit(1)

# ------------- API auth / Dashboard client -------------
def validate_api_key(key: str) -> bool:
    return bool(re.fullmatch(r'[A-Fa-f0-9]{40}', key or ''))

MAX_API_KEY_ATTEMPTS: int = 4
attempts: int = 0
API_KEY: Optional[str] = None
while attempts < MAX_API_KEY_ATTEMPTS:
    API_KEY = getpass("Enter your Meraki API key (hidden): ")
    if validate_api_key(API_KEY):
        break
    attempts += 1
    print(f"❌ Invalid API key. ({MAX_API_KEY_ATTEMPTS - attempts} attempt(s) left)")
else:
    print("❌ Maximum attempts reached. Exiting.")
    raise SystemExit(1)

dashboard = meraki.DashboardAPI(
    api_key=API_KEY,
    base_url=BASE_URL,
    output_log=False,
    print_console=False,
    suppress_logging=True,
    maximum_retries=5,
    wait_on_rate_limit=True,
)

# ------------- Graceful abort -------------
_aborted: bool = False
def _handle_sigint(signum, frame) -> None:
    global _aborted
    _aborted = True
    print("\nReceived Ctrl+C — attempting graceful shutdown...")
    log_change('workflow_abort', 'User interrupted with SIGINT')
signal.signal(signal.SIGINT, _handle_sigint)

# ------------- Error helpers -------------
def is_vlans_disabled_error(exc: Exception) -> bool:
    needle = "VLANs are not enabled for this network"
    try:
        if isinstance(exc, meraki.APIError):
            msg = getattr(exc, "message", "") or ""
            body = getattr(getattr(exc, "response", None), "text", "") or ""
            return (needle.lower() in msg.lower()) or (needle.lower() in body.lower())
        return needle.lower() in str(exc).lower()
    except Exception:
        return False

def do_action(func, *args, **kwargs):
    if DRY_RUN:
        logging.debug(f"DRY RUN: {func.__name__} args={args} kwargs={kwargs}")
        return None
    return func(*args, **kwargs)

# ======================================================
# ------------- Wireless pre-check helpers -------------
# ======================================================
WIRELESS_PREFIXES: Tuple[str, ...] = ("MR", "CW")

def _is_wireless_model(model: Optional[str]) -> bool:
    return bool(model) and model.upper().startswith(WIRELESS_PREFIXES)

def _is_mr33(model: Optional[str]) -> bool:
    return bool(model) and model.upper().startswith("MR33")

def _get_network_wireless_devices(network_id: str) -> List[Dict[str, Any]]:
    try:
        devices: List[Dict[str, Any]] = dashboard.networks.getNetworkDevices(network_id)  # type: ignore[attr-defined]
    except Exception:
        logging.exception("Failed to list devices for wireless check")
        return []
    return [d for d in devices if _is_wireless_model(cast(Optional[str], d.get("model")))]

def _get_inventory_models_for_serials(org_id: str, serials: List[str]) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for s in serials:
        try:
            inv: Dict[str, Any] = dashboard.organizations.getOrganizationInventoryDevice(org_id, s)  # type: ignore[attr-defined]
            mdl_opt: Optional[str] = cast(Optional[str], inv.get("model"))
            if mdl_opt:
                out[s] = mdl_opt
        except Exception:
            logging.exception("Inventory lookup failed for %s", s)
    return out

def _prompt_yes_no(question: str, default_no: bool = True) -> bool:
    prompt: str = " [y/N] " if default_no else " [Y/n] "
    ans_raw: str = input(question + prompt).strip().lower()
    if not ans_raw:
        return (not default_no)
    return ans_raw in ("y", "yes")

def _prompt_replacement_mapping(old_serials: List[str], new_serials: List[str]) -> List[Tuple[str, str]]:
    if not old_serials or not new_serials:
        return []
    print("\nEnter replacement pairs as 'OLD:NEW'. Leave blank to finish.")
    print(f"Old (present in network): {', '.join(old_serials)}")
    print(f"New (available to add):   {', '.join(new_serials)}")
    available_new: Set[str] = {s.upper() for s in new_serials}
    old_upper: Set[str] = {s.upper() for s in old_serials}
    mapping: List[Tuple[str, str]] = []
    def _restore_case(target: str, pool: List[str]) -> str:
        for p in pool:
            if p.upper() == target.upper():
                return p
        return target
    while True:
        line: str = input("Pair (OLD:NEW): ").strip()
        if not line:
            break
        if ":" not in line:
            print("  Format must be OLD:NEW")
            continue
        old_s, new_s = [p.strip() for p in line.split(":", 1)]
        if old_s.upper() not in old_upper:
            print(f"  {old_s} is not in the old-serials list.")
            continue
        if new_s.upper() not in available_new:
            print(f"  {new_s} is not in the new-serials list or already used.")
            continue
        mapping.append((_restore_case(old_s, old_serials), _restore_case(new_s, new_serials)))
        available_new.remove(new_s.upper())
    return mapping

def ensure_mr33_and_handle_wireless_replacements(
    org_id: str,
    network_id: str,
    serials_to_add: List[str],
) -> Tuple[List[str], List[str], List[str]]:
    add_models: Dict[str, str] = _get_inventory_models_for_serials(org_id, serials_to_add)
    incoming_wireless: List[str] = [s for s, m in add_models.items() if _is_wireless_model(m)]
    if not incoming_wireless:
        return serials_to_add, [], []
    wireless_now: List[Dict[str, Any]] = _get_network_wireless_devices(network_id)
    has_mr33_now: bool = any(_is_mr33(cast(Optional[str], d.get("model"))) for d in wireless_now)
    non_mr33_in_net: List[Dict[str, Any]] = [
        d for d in wireless_now
        if _is_wireless_model(cast(Optional[str], d.get("model"))) and not _is_mr33(cast(Optional[str], d.get("model")))
    ]
    adding_has_mr33: bool = any(_is_mr33(add_models.get(s)) for s in incoming_wireless)
    if not has_mr33_now and not adding_has_mr33:
        # Ask before touching wireless when no MR33 exists now nor being added
        proceed = _prompt_yes_no("No MR33 detected in network or incoming. Proceed with wireless changes?", default_no=True)
        if not proceed:
            # User said "No": continue the overall workflow, but do NOT add or remove wireless.
            #  - Leave existing wireless untouched (removed_old stays empty)
            #  - Prevent new wireless from being claimed by removing them from serials_to_add
            serials_to_add = [s for s in serials_to_add if s not in incoming_wireless]
            return serials_to_add, [], []

    removed_old: List[str] = []
    claimed_new: List[str] = []
    if non_mr33_in_net and _prompt_yes_no("Replace non-MR33 wireless with incoming?", default_no=False):
        mapping = _prompt_replacement_mapping(
            [cast(str, d.get("serial")) for d in non_mr33_in_net if d.get("serial")],
            incoming_wireless
        )
        for old_serial, new_serial in mapping:
            try:
                do_action(dashboard.devices.updateDevice, old_serial, name="", address="")  # type: ignore[attr-defined]
                do_action(dashboard.networks.removeNetworkDevices, network_id, old_serial)  # type: ignore[attr-defined]
                log_change('wireless_replace_remove', f"Removed old wireless {old_serial}", device_serial=old_serial)
                removed_old.append(old_serial)
            except Exception:
                logging.exception("Failed to remove %s", old_serial)
            try:
                do_action(dashboard.networks.claimNetworkDevices, network_id, serials=[new_serial])  # type: ignore[attr-defined]
                log_change('wireless_replace_claim', f"Claimed new wireless {new_serial}", device_serial=new_serial)
                claimed_new.append(new_serial)
            except Exception:
                logging.exception("Failed to claim %s", new_serial)
    claimed_new_set: Set[str] = set(claimed_new)
    serials_to_add = [s for s in serials_to_add if s not in claimed_new_set]
    return serials_to_add, removed_old, claimed_new


# ------------- Switch port helpers (diff + apply) -------------
_PORT_FIELDS = [
    "enabled", "name", "tags", "type", "vlan", "voiceVlan", "allowedVlans",
    "poeEnabled", "isolationEnabled", "rstpEnabled", "stpGuard",
    "linkNegotiation", "udld", "accessPolicyType", "accessPolicyNumber",
    "portScheduleId"
]

def _normalize_tags(value):
    if isinstance(value, list):
        return sorted(value)
    if isinstance(value, str):
        return sorted([t for t in value.split() if t])
    return []

def _port_dict_by_number(ports: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    for p in ports:
        pid = p.get("portId") or p.get("number") or p.get("name")
        if pid is None:
            continue
        out[str(pid)] = p
    return out

def compute_port_overrides(live_ports: List[Dict[str, Any]], tmpl_ports: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    overrides: Dict[str, Dict[str, Any]] = {}
    live = _port_dict_by_number(live_ports)
    tmpl = _port_dict_by_number(tmpl_ports)
    for pid, lp in live.items():
        tp = tmpl.get(pid)
        if not tp:
            continue
        for fld in _PORT_FIELDS:
            lv = lp.get(fld)
            tv = tp.get(fld)
            if fld == "tags":
                lv = _normalize_tags(lv)
                tv = _normalize_tags(tv)
            if lv is not None and lv != tv:
                overrides.setdefault(pid, {})[fld] = lv
    return overrides

def apply_port_overrides(serial: str, overrides: Dict[str, Dict[str, Any]]) -> None:
    for pid, patch in overrides.items():
        try:
            do_action(dashboard.switch.updateDeviceSwitchPort, serial, pid, **patch)
            logging.debug(f"Applied port overrides on {serial} port {pid}: {patch}")
            log_change('switch_port_override', f"Applied port overrides on port {pid}",
                       device_serial=serial, misc=json.dumps(patch))
        except Exception:
            logging.exception(f"Failed applying port overrides on {serial} port {pid}")

# ------------- Domain helpers (library) -------------
# ---- NEW / UPDATED HELPERS FOR NETWORK MATCHING & PAGINATION ----

_DASH_VARIANTS: Tuple[str, ...] = ("\u2010", "\u2011", "\u2012", "\u2013", "\u2014", "\u2212")

def _normalize_name(s: Optional[str]) -> str:
    """Normalize network names for reliable matching."""
    if not s:
        return ""
    out = s
    for ch in _DASH_VARIANTS:
        out = out.replace(ch, "-")
    # collapse whitespace, lower-case
    out = re.sub(r"\s+", " ", out).strip().lower()
    return out

def get_all_networks(org_id: str) -> List[Dict[str, Any]]:
    """
    Return *all* networks for the organization, handling pagination in a way that
    works across meraki SDK versions.
    """
    try:
        # Preferred path (SDK paginates internally)
        nets = dashboard.organizations.getOrganizationNetworks(
            org_id, total_pages="all", perPage=1000  # type: ignore[arg-type]
        )
        if isinstance(nets, list):
            return nets
    except TypeError:
        # Older SDKs may not accept total_pages argument
        pass
    except Exception:
        logging.exception("get_all_networks: total_pages=all path failed, falling back")

    # Fallback: manual page-walk using startingAfter
    results: List[Dict[str, Any]] = []
    starting_after: Optional[str] = None
    seen_ids: Set[str] = set()

    while True:
        try:
            batch: List[Dict[str, Any]]
            if starting_after:
                batch = dashboard.organizations.getOrganizationNetworks(
                    org_id, perPage=1000, startingAfter=starting_after
                )
            else:
                batch = dashboard.organizations.getOrganizationNetworks(
                    org_id, perPage=1000
                )
        except Exception:
            logging.exception("get_all_networks: page fetch failed")
            break

        if not batch:
            break

        for n in batch:
            nid = str(n.get("id", ""))
            if nid and nid not in seen_ids:
                results.append(n)
                seen_ids.add(nid)

        # The SDK auto-follows Link headers when possible; if not, approximate:
        if len(batch) < 1000:
            break
        starting_after = str(batch[-1].get("id", "")) or None
        if not starting_after:
            break

    return results

def fetch_matching_networks(org_id: str, partial: str) -> List[Dict[str, Any]]:
    """
    More reliable matching:
      - Accept direct Network ID (exact match).
      - Case/space/Unicode-dash insensitive substring matching.
      - If input looks numeric (e.g., store code), match as substring token too.
    """
    partial_norm: str = _normalize_name(partial)
    all_nets: List[Dict[str, Any]] = get_all_networks(org_id)

    # Direct ID match first
    direct = [n for n in all_nets if str(n.get("id", "")).strip() == partial.strip()]
    if direct:
        logging.debug("fetch_matching_networks: direct ID match")
        return direct

    # Normalized name contains
    matches: List[Dict[str, Any]] = []
    for n in all_nets:
        nname_raw: str = str(n.get("name", "") or "")
        nname_norm: str = _normalize_name(nname_raw)
        if partial_norm and partial_norm in nname_norm:
            matches.append(n)

    # If no matches and partial is a numeric token, try as a bare substring (without normalization)
    if not matches and re.fullmatch(r"\d{2,}", partial.strip()):
        num = partial.strip()
        matches = [n for n in all_nets if num in str(n.get("name", "") or "")]

    logging.debug("Found %d networks matching '%s'", len(matches), partial)
    return matches

def fetch_devices(
    org_id: str,
    network_id: str,
    template_id: Optional[str] = None
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Fetch MX/MS/MR device summaries for a network.
    If template_id is provided, compute per-MS port overrides by diffing
    live ports vs. the assigned switch profile ports from the config template.
    """
    devs = dashboard.networks.getNetworkDevices(network_id)

    def _mk(d):
        tags = d.get('tags', [])
        if not isinstance(tags, list):
            tags = (tags or '').split()
        return {
            'serial': d['serial'],
            'model': d['model'],
            'tags': tags,
            'address': d.get('address', ''),
            'name': d.get('name', ''),
            'switchProfileId': d.get('switchProfileId'),
            'switchProfileName': d.get('switchProfileName'),
        }

    mx = [_mk(d) for d in devs if d['model'].startswith('MX')]
    ms = [_mk(d) for d in devs if d['model'].startswith('MS')]
    mr = [_mk(d) for d in devs if _is_wireless_model(d.get('model'))]


    # Per-MS port overrides vs current template profile (if known)
    if template_id:
        for sw in ms:
            profile_id = sw.get('switchProfileId')
            if not profile_id:
                sw['port_overrides'] = {}
                continue

            try:
                # Live device ports
                live_ports = dashboard.switch.getDeviceSwitchPorts(sw['serial'])

                # Template profile ports (✅ correct endpoint & namespace)
                tmpl_ports = dashboard.switch.getOrganizationConfigTemplateSwitchProfilePorts(
                    org_id, template_id, profile_id
                )

                sw['port_overrides'] = compute_port_overrides(live_ports, tmpl_ports)
                logging.debug(
                    "Computed %d port overrides for %s (profile %s)",
                    len(sw['port_overrides']), sw['serial'], profile_id
                )
            except meraki.APIError as e:
                logging.exception("Failed computing port overrides for %s: %s", sw['serial'], e)
                sw['port_overrides'] = {}
            except Exception:
                logging.exception("Failed computing port overrides for %s", sw['serial'])
                sw['port_overrides'] = {}
    else:
        for sw in ms:
            sw['port_overrides'] = {}

    logging.debug("Fetched devices: MX=%d, MS=%d, MR=%d", len(mx), len(ms), len(mr))
    log_change(
        event='fetch_devices',
        details=f"Fetched devices for network {network_id}",
        network_id=network_id,
        misc=f"mx={json.dumps(mx)}, ms={json.dumps(ms)}, mr={json.dumps(mr)}"
    )
    return mx, ms, mr

def fetch_vlan_details(network_id: str) -> List[Dict[str, Any]]:
    try:
        vlans = dashboard.appliance.getNetworkApplianceVlans(network_id)
        filtered = [v for v in vlans if int(v.get('id')) not in EXCLUDED_VLANS]
        logging.debug(f"Fetched VLANs: {len(filtered)} (excluded {len(vlans) - len(filtered)})")
        return filtered
    except meraki.APIError as e:
        if is_vlans_disabled_error(e):
            logging.warning("VLAN endpoints unavailable because VLANs are disabled on this network (returning empty list).")
            return []
        logging.exception("Failed to fetch VLANs")
        return []
    except Exception:
        logging.exception("Failed to fetch VLANs")
        return []

def vlans_enabled(network_id: str) -> Optional[bool]:
    try:
        settings = dashboard.appliance.getNetworkApplianceVlansSettings(network_id)
        return bool(settings.get("vlansEnabled"))
    except Exception:
        logging.exception("Could not read VLANs settings")
        return None
# ---- UPDATED VLAN UPDATE (CONDITIONAL PAYLOAD FOR DHCP) ----
def _dhcp_is_server(dhcp_handling: Optional[str]) -> bool:
    """
    Return True only when VLAN DHCP is 'Run a DHCP server'.
    Be defensive to handle minor string variations.
    """
    val = (dhcp_handling or "").strip().lower()
    # Canonical UI/API value is 'Run a DHCP server'
    # Accept tolerant variants just in case.
    return val in {
        "run a dhcp server", "dhcp server", "server", "enabled", "on"
    }

def update_vlans(network_id: str, network_name: str, vlan_list: List[Dict[str, Any]]) -> None:
    for v in vlan_list:
        # Base payload (always safe)
        payload: Dict[str, Any] = {
            "applianceIp": v.get("applianceIp"),
            "subnet": v.get("subnet"),
            "dhcpHandling": v.get("dhcpHandling"),
        }

        # Only send static/reserved assignments when DHCP == SERVER
        if _dhcp_is_server(v.get("dhcpHandling")):
            if "fixedIpAssignments" in v and isinstance(v["fixedIpAssignments"], dict):
                payload["fixedIpAssignments"] = v["fixedIpAssignments"]
            if "reservedIpRanges" in v and isinstance(v["reservedIpRanges"], list):
                payload["reservedIpRanges"] = v["reservedIpRanges"]
        else:
            # Ensure these fields are NOT present for OFF/RELAY
            payload.pop("fixedIpAssignments", None)
            payload.pop("reservedIpRanges", None)

        try:
            do_action(
                dashboard.appliance.updateNetworkApplianceVlan,  # type: ignore[attr-defined]
                network_id, v["id"], **payload
            )
            logging.debug("Updated VLAN %s with payload keys: %s", v.get("id"), list(payload.keys()))
            log_change(
                "vlan_update",
                f"Updated VLAN {v.get('id')}",
                device_name=f"Network: {network_id}",
                network_id=network_id,
                network_name=network_name,
                misc=json.dumps(payload),
            )
        except meraki.APIError as e:
            if is_vlans_disabled_error(e):
                raise
            logging.exception("Failed to update VLAN %s", v.get("id"))
        except Exception:
            logging.exception("Failed to update VLAN %s", v.get("id"))

def classify_serials_for_binding(org_id: str, net_id: str, serials: List[str]):
    already, elsewhere, avail = [], [], []
    for s in serials:
        try:
            inv = dashboard.organizations.getOrganizationInventoryDevice(org_id, s)
            nid = inv.get('networkId')
            if nid == net_id:
                already.append(s)
            elif nid:
                elsewhere.append((s, inv.get('networkName') or nid))
            else:
                avail.append(s)
        except meraki.APIError as e:
            if getattr(e, "status", None) == 404:
                avail.append(s)
            else:
                logging.error(f"Error checking inventory for {s}: {e}")
        except Exception as e:
            logging.error(f"Error checking inventory for {s}: {e}")
    return already, elsewhere, avail

# ------------- Clear & remove by model -------------
def _clear_and_remove_models(org_id: str, network_id: str, models: Tuple[str, ...]) -> bool:
    """
    Clear (name/address) and remove all devices in `network_id` whose exact model
    matches any entry in `models` (e.g., ("MX64",) or ("MR33",)).

    Uses:
      - dashboard.devices.updateDevice(serial, name="", address="")
      - dashboard.networks.removeNetworkDevices(network_id, serial)  # per-serial
    """
    try:
        mx, ms, mr = fetch_devices(org_id, network_id)
        candidates = mx + ms + mr
        to_remove = [d for d in candidates if d.get('model') in models]
    except Exception:
        logging.exception("Failed to enumerate devices before removal")
        return False

    if not to_remove:
        logging.debug("No devices found matching models %s in network %s", models, network_id)
        return True

    # 1) Clear device metadata first (best-effort, per device)
    for d in to_remove:
        serial = d.get('serial')
        if not serial:
            continue
        try:
            do_action(dashboard.devices.updateDevice, serial, name="", address="")
            log_change('device_clear', f"Cleared config for {serial}", device_serial=serial)
        except Exception:
            logging.exception("Error clearing %s", serial)

    # 2) Remove devices one-by-one (correct SDK usage)
    all_ok = True
    for d in to_remove:
        serial = d.get('serial')
        if not serial:
            continue
        try:
            do_action(dashboard.networks.removeNetworkDevices, network_id, serial)
            log_change('device_removed', "Removed device from network", device_serial=serial)
        except Exception:
            all_ok = False
            logging.exception("Error removing device %s", serial)

    return all_ok


def remove_existing_mx64_devices(org_id: str, network_id: str) -> bool:
    return _clear_and_remove_models(org_id, network_id, ("MX64",))

def remove_existing_MR33_devices(org_id: str, network_id: str) -> bool:
    return _clear_and_remove_models(org_id, network_id, ("MR33",))

# ------------- Serial collection / claim flows -------------
def prompt_and_validate_serials(org_id: str) -> List[str]:
    MAX_SERIAL_ATTEMPTS = 4
    MAX_BLANK_ATTEMPTS = 4

    while True:
        count_raw = input("How many devices/serials will you add to this org? (Enter to skip): ").strip()
        if not count_raw:
            return []
        try:
            intended_count = int(count_raw)
            if intended_count <= 0:
                print("ℹ️  Count must be a positive integer.")
                continue
            break
        except ValueError:
            print("ℹ️  Please enter a whole number (e.g., 3).")

    blank_attempts = 0
    while True:
        print("\nEnter serial numbers:")
        print(" - Paste as comma-separated, or one per line (blank to finish).")
        first_line = input("Enter serial(s): ").strip().upper()

        raw_serials: List[str] = []
        if "," in first_line:
            raw_serials = [s.strip().upper() for s in first_line.split(",") if s.strip()]
        else:
            if first_line:
                raw_serials.append(first_line)
            while True:
                nxt = input("Enter next serial (or blank to finish): ").strip().upper()
                if not nxt:
                    break
                raw_serials.append(nxt)

        if not raw_serials:
            blank_attempts += 1
            remaining = MAX_BLANK_ATTEMPTS - blank_attempts
            if remaining <= 0:
                print("\n❌ No serial number(s) entered after 4 attempts -----------")
                print("   Please retry when serial(s) are known *******")
                sys.exit(1)
            print(f"ℹ️  No serials provided. Try again. (attempt {blank_attempts}/{MAX_BLANK_ATTEMPTS})")
            continue

        # dedupe preserving order
        seen: Set[str] = set()
        serial_list: List[str] = []
        for s in raw_serials:
            if s in seen:
                print(f"ℹ️  Duplicate serial '{s}' removed from input.")
                continue
            seen.add(s)
            serial_list.append(s)

        entered_count = len(serial_list)
        if entered_count != intended_count:
            print(f"⚠️  You said {intended_count} device(s) but entered {entered_count}.")
            choice = input("Proceed anyway? (yes to proceed / no to re-enter): ").strip().lower()
            if choice not in {"y", "yes"}:
                continue

        collected: List[str] = []
        for idx, original_serial in enumerate(serial_list, start=1):
            attempts = 0
            serial = original_serial
            while attempts < MAX_SERIAL_ATTEMPTS:
                if not re.fullmatch(r"[A-Z0-9]{4}-[A-Z0-9]{4}-[A-Z0-9]{4}", serial or ""):
                    attempts += 1
                    if attempts >= MAX_SERIAL_ATTEMPTS:
                        print(f"❌ Maximum attempts reached for serial #{idx} ({original_serial}). Skipping.")
                        break
                    serial = input(
                        f"Serial #{idx} '{serial}' is invalid. Re-enter (attempt {attempts+1}/{MAX_SERIAL_ATTEMPTS}): "
                    ).strip().upper()
                    continue

                try:
                    dashboard.organizations.getOrganizationInventoryDevice(org_id, serial)
                    print(f"✅ {serial} found in org inventory.")
                    collected.append(serial)
                    break
                except meraki.APIError as e:
                    if getattr(e, "status", None) == 404:
                        try:
                            do_action(dashboard.organizations.claimIntoOrganizationInventory, org_id, serials=[serial])
                            print(f"✅ Serial '{serial}' successfully claimed into org inventory.")
                            log_change('device_claimed_inventory', "Claimed serial into org inventory", device_serial=serial)
                            collected.append(serial)
                            break
                        except Exception as claim_ex:
                            attempts += 1
                            print(f"❌ Error claiming '{serial}' into org inventory: {claim_ex}")
                            if attempts >= MAX_SERIAL_ATTEMPTS:
                                print(f"❌ Maximum attempts reached for serial #{idx}. Skipping.")
                                break
                            serial = input(
                                f"Re-enter serial #{idx} (attempt {attempts+1}/{MAX_SERIAL_ATTEMPTS}): "
                            ).strip().upper()
                            continue
                    else:
                        print(f"API Error for serial '{serial}': {e}")
                        break
                except Exception as e:
                    print(f"API Error for serial '{serial}': {e}")
                    break

        if len(collected) != intended_count:
            print(f"⚠️  Intended: {intended_count}, Entered: {entered_count}, Validated: {len(collected)}.")
            choice = input("Proceed with validated devices anyway? (yes to proceed / no to re-enter all): ").strip().lower()
            if choice in {"y", "yes"}:
                return collected
            else:
                print("Okay, let's re-enter the serial list.")
                continue

        return collected

def summarize_devices_in_org(org_id: str, serials: List[str]) -> Set[str]:
    detected_mx_models: Set[str] = set()
    if not serials:
        print("No serials to summarize.")
        return detected_mx_models

    print("\nValidated / added to organization:")
    for s in serials:
        try:
            inv = dashboard.organizations.getOrganizationInventoryDevice(org_id, s)
            model = inv.get('model') or 'Unknown'
            ptypes = inv.get('productTypes') or []
            ptype = ptypes[0] if isinstance(ptypes, list) and ptypes else inv.get('productType') or 'Unknown'
            name = inv.get('name') or ''
            print(f" - {s}: {model} ({ptype}){f' — {name}' if name else ''}")

            if model.startswith('MX67'):
                detected_mx_models.add('MX67')
            elif model.startswith('MX75'):
                detected_mx_models.add('MX75')
        except Exception as e:
            print(f" - {s}: (lookup failed: {e})")

    return detected_mx_models

def claim_devices(org_id: str, network_id: str, prevalidated_serials: Optional[List[str]] = None) -> List[str]:
    if prevalidated_serials is not None:
        valids = prevalidated_serials
    else:
        valids = prompt_and_validate_serials(org_id)

    if not valids:
        print("❌ No valid serials.")
        return []

    already, elsewhere, avail = classify_serials_for_binding(org_id, network_id, valids)
    if elsewhere:
        print("⚠️ In use elsewhere:")
        for s, name in elsewhere:
            print(f" - {s} in {name}")

    mx_models: List[str] = []
    for s in avail:
        try:
            inv = dashboard.organizations.getOrganizationInventoryDevice(org_id, s)
            if (inv.get('model') or '').startswith('MX'):
                mx_models.append(inv['model'])
        except Exception:
            pass
    if len(set(mx_models)) > 1:
        print("❌ MX warm spare models mismatch. Aborting.")
        return []
    if not avail:
        print("ℹ️ No newly available devices to claim to the network (perhaps already present).")
        return already

    try:
        remove_existing_mx64_devices(org_id, network_id)
        do_action(dashboard.networks.claimNetworkDevices, network_id, serials=avail)
        for s in avail:
            log_change('device_claimed', f"Claimed device to network", device_serial=s)
        return avail
    except Exception:
        logging.exception("Failed to claim/bind")
        return []

# ------------- Ordering / warm spare -------------
def select_primary_mx(org_id: str, serials: List[str]) -> Optional[str]:
    mx_candidates: List[Tuple[str, str]] = []
    for s in serials:
        try:
            inv = dashboard.organizations.getOrganizationInventoryDevice(org_id, s)
            model = (inv.get('model') or '').upper()
            if model.startswith('MX'):
                mx_candidates.append((s, model))
        except Exception:
            logging.exception(f"Unable to read inventory for {s}")

    if len(mx_candidates) == 0:
        return None
    if len(mx_candidates) == 1:
        return mx_candidates[0][0]

    auto_choice = sorted([s for s, _ in mx_candidates])[0]
    print("\nMultiple MX devices detected in the claimed list:")
    for idx, (s, m) in enumerate(mx_candidates, 1):
        print(f" {idx}. {s}  ({m})")
    sel = input("Select which MX should be PRIMARY (mx-01). "
                "Enter number, or press Enter / type 'skip'/'cancel' to auto-select: ").strip().lower()

    if not sel or sel in {'skip', 'cancel'}:
        print(f"ℹ️  No explicit choice made. Auto-selecting PRIMARY MX: {auto_choice}")
        return auto_choice

    if sel.isdigit():
        i = int(sel)
        if 1 <= i <= len(mx_candidates):
            return mx_candidates[i-1][0]

    print(f"ℹ️  Invalid selection. Auto-selecting PRIMARY MX: {auto_choice}")
    return auto_choice

def select_device_order(org_id: str, serials: List[str], kind: str) -> List[str]:
    filtered: List[Tuple[str, str]] = []  # (serial, model)
    for s in serials:
        try:
            inv = dashboard.organizations.getOrganizationInventoryDevice(org_id, s)
            model = (inv.get('model') or '').upper()
            if kind == 'MR' and _is_wireless_model(model):
                filtered.append((s, model))
            elif kind == 'MS' and model.startswith('MS'):
                filtered.append((s, model))
        except Exception:
            logging.exception(f"Unable to read inventory for {s}")

    if len(filtered) <= 1:
        return [s for s, _ in filtered]

    auto_order = sorted([s for s, _ in filtered])

    print(f"\nSelect ordering for {kind} devices (enter a comma-separated list of indices).")
    for idx, (s, m) in enumerate(filtered, 1):
        print(f" {idx}. {s}  ({m})")
    raw = input(f"Desired order for {kind} (e.g. 2,1,3). "
                "Press Enter / type 'skip'/'cancel' to auto-order: ").strip().lower()

    if not raw or raw in {'skip', 'cancel'}:
        print(f"ℹ️  Auto-ordering {kind} devices by serial: {', '.join(auto_order)}")
        return auto_order

    parts = [p.strip() for p in raw.split(',') if p.strip()]
    if all(p.isdigit() and 1 <= int(p) <= len(filtered) for p in parts) and len(parts) == len(filtered):
        return [filtered[int(p)-1][0] for p in parts]

    print(f"ℹ️  Invalid list. Auto-ordering {kind} devices by serial: {', '.join(auto_order)}")
    return auto_order

def ensure_primary_mx(network_id: str, desired_primary_serial: Optional[str]) -> None:
    if not desired_primary_serial:
        return
    try:
        status = dashboard.appliance.getNetworkApplianceWarmSpare(network_id) or {}
        enabled = bool(status.get("enabled"))
        current_primary = status.get("primarySerial")
        if not enabled:
            print("ℹ️  Warm spare is not enabled on this network; cannot swap primary automatically.")
            log_change('mx_warmspare_not_enabled',
                       "Warm spare not enabled; no primary swap performed",
                       network_id=network_id)
            return
        if current_primary and current_primary.upper() == desired_primary_serial.upper():
            print(f"✅ Warm spare already has the correct primary ({desired_primary_serial}).")
            return
        print(f"🔁 Swapping warm spare primary to {desired_primary_serial} ...")
        do_action(dashboard.appliance.swapNetworkApplianceWarmSpare, network_id)
        log_change('mx_warmspare_swap',
                   f"Swapped warm spare primary to {desired_primary_serial}",
                   device_serial=desired_primary_serial, network_id=network_id)
        print("✅ Warm spare primary swap requested.")
    except Exception as e:
        logging.exception("Failed to ensure warm spare primary")
        print(f"❌ Failed to verify/swap warm spare primary: {e}")

# ------------- Naming & configuration -------------
def name_and_configure_claimed_devices(
    org_id: str,
    network_id: str,
    network_name: str,
    serials: List[str],
    ms_list: List[Dict[str, Any]],
    mr_list: List[Dict[str, Any]],
    tpl_profile_map: Dict[str, str],
    old_mx_devices: Optional[List[Dict[str, Any]]] = None,
    old_mr_devices: Optional[List[Dict[str, Any]]] = None,
    primary_mx_serial: Optional[str] = None,
    mr_order: Optional[List[str]] = None,
    ms_order: Optional[List[str]] = None,
):
    prefix = '-'.join(network_name.split('-')[:2]).lower()
    counts = {'MX': 1, 'MR': 1, 'MS': 1}
    old_mr33s = sorted([d for d in (old_mr_devices or []) if d['model'] == 'MR33'], key=lambda x: x.get('name', ''))
    old_mxs_sorted = sorted((old_mx_devices or []) if old_mx_devices else [], key=lambda x: x.get('name', ''))

    inv_by_serial: Dict[str, Dict[str, Any]] = {}
    for s in serials:
        try:
            inv_by_serial[s] = dashboard.organizations.getOrganizationInventoryDevice(org_id, s)
        except Exception:
            logging.exception(f"Failed inventory lookup for {s}")
            inv_by_serial[s] = {}

    mx_serials = [s for s in serials if (inv_by_serial.get(s, {}).get('model') or '').upper().startswith('MX')]
    mr_serials = [
        s for s in serials
        if _is_wireless_model((inv_by_serial.get(s, {}).get('model') or '').upper())
    ]

    ms_serials = [s for s in serials if (inv_by_serial.get(s, {}).get('model') or '').upper().startswith('MS')]

    if primary_mx_serial and primary_mx_serial in mx_serials:
        mx_serials = [primary_mx_serial] + [s for s in mx_serials if s != primary_mx_serial]

    if mr_order:
        mr_serials = [s for s in mr_order if s in mr_serials]
    if ms_order:
        ms_serials = [s for s in ms_order if s in ms_serials]

# MX
    mx_idx = 0
    for s in mx_serials:
        mdl = (inv_by_serial.get(s, {}).get('model') or '')
        mdl_upper = mdl.upper()

        data: Dict[str, Any] = {}
        data['name'] = f"{prefix}-mx-{counts['MX']:02}"
        if mx_idx < len(old_mxs_sorted):
            data['address'] = old_mxs_sorted[mx_idx].get('address', '')
            data['tags'] = old_mxs_sorted[mx_idx].get('tags', [])
        else:
            data['address'] = ''
            data['tags'] = []
        mx_idx += 1
        counts['MX'] += 1

    # Rename / re-tag first (as you already did)
        try:
            do_action(dashboard.devices.updateDevice, s, **data)
            log_change('device_update', f"Renamed and reconfigured device {s} ({mdl})",
                   device_serial=s, device_name=data.get('name', ''),
                   misc=f"tags={data.get('tags', [])}, address={data.get('address', '')}")
        except Exception:
            logging.exception(f"Failed configuring {s} (MX)")

    # Enable WAN2 ONLY on MX67 models
        if mdl_upper.startswith('MX67'):
            try:
                do_action(
                dashboard.appliance.updateDeviceApplianceUplinksSettings,
                s,
                interfaces={"wan2": {"enabled": True}}
                )
                log_change(
                'mx_wan2_enable',
                'Enabled WAN2 on MX67',
                device_serial=s,
                device_name=data.get('name', ''),
                misc='interfaces={"wan2":{"enabled":true}}'
            )
            except Exception:
                logging.exception(f"Failed to enable WAN2 uplink on {s}")
        else:
            logging.debug("Skipped WAN2 enable for non-MX67 %s (%s)", s, mdl)
# MX
    
    # MR
    ap_idx = 0
    for s in mr_serials:
        mdl = (inv_by_serial.get(s, {}).get('model') or '')
        data: Dict[str, Any] = {'name': f"{prefix}-ap-{counts['MR']:02}"}
        if ap_idx < len(old_mr33s):
            data['tags'] = old_mr33s[ap_idx].get('tags', [])
            data['address'] = old_mr33s[ap_idx].get('address', '')
        else:
            data['tags'] = []
            data['address'] = ''
        ap_idx += 1
        counts['MR'] += 1
        try:
            do_action(dashboard.devices.updateDevice, s, **data)
            log_change('device_update', f"Renamed and reconfigured device {s} ({mdl})",
                       device_serial=s, device_name=data.get('name', ''),
                       misc=f"tags={data.get('tags', [])}, address={data.get('address', '')}")
        except Exception:
            logging.exception(f"Failed configuring {s} (MR)")

    # MS
    for s in ms_serials:
        mdl = (inv_by_serial.get(s, {}).get('model') or '')
        data: Dict[str, Any] = {'name': f"{prefix}-ms-{counts['MS']:02}"}
        counts['MS'] += 1
        prof_name = ms_list[0].get('switchProfileName') if ms_list else None
        prof_id = tpl_profile_map.get(prof_name) if prof_name else None
        if prof_id:
            data['switchProfileId'] = prof_id
        try:
            do_action(dashboard.devices.updateDevice, s, **data)
            log_change('device_update', f"Renamed and reconfigured device {s} ({mdl})",
                       device_serial=s, device_name=data.get('name', ''),
                       misc=f"tags={data.get('tags', [])}, address={data.get('address', '')}")
        except Exception:
            logging.exception(f"Failed configuring {s} (MS)")

def remove_recently_added_tag(network_id: str):
    devs = dashboard.networks.getNetworkDevices(network_id)
    for d in devs:
        tags = d.get('tags', [])
        if not isinstance(tags, list):
            tags = (tags or '').split()
        if 'recently-added' in tags:
            updated_tags = [t for t in tags if t != 'recently-added']
            print(f"Removing 'recently-added' tag from {d['model']} {d['serial']}")
            try:
                do_action(dashboard.devices.updateDevice, d['serial'], tags=updated_tags)
                log_change(
                    'tag_removed', "Removed 'recently-added' tag",
                    device_serial=d['serial'], device_name=d.get('name', ''),
                    misc=f"old_tags={tags}, new_tags={updated_tags}"
                )
            except Exception:
                logging.exception(f"Failed to remove 'recently-added' from {d['serial']}")

# ------------- Template rebind helpers -------------
def _pick_template_by_vlan_count(templates: List[Dict[str, Any]], vlan_count: Optional[int]) -> Optional[Dict[str, Any]]:
    if vlan_count not in (3, 5):
        return None

    patterns = []
    if vlan_count == 3:
        patterns = [r'NO\s*LEGACY.*MX*\b']
    elif vlan_count == 5:
        patterns = [r'3\s*X\s*DATA[_\s-]*VLAN.*MX75\b']

    for pat in patterns:
        rx = re.compile(pat, re.IGNORECASE)
        for t in templates:
            name = (t.get('name') or '')
            if rx.search(name):
                return t
    return None

def list_and_rebind_template(
    org_id: str,
    network_id: str,
    current_id: Optional[str],
    network_name: str,
    *,
    pre_change_devices: Optional[List[Dict[str, Any]]] = None,
    pre_change_vlans: Optional[List[Dict[str, Any]]] = None,
    pre_change_template: Optional[str] = None,
    claimed_serials: Optional[List[str]] = None,
    removed_serials: Optional[List[str]] = None,
    ms_list: Optional[List[Dict[str, Any]]] = None,
    mx_model_filter: Optional[str] = None,
    vlan_count: Optional[int] = None,
) -> Tuple[Optional[str], Optional[str], bool]:
    skip_attempts = 0

    while True:
        print(f"\nCurrent network: {network_name} (ID: {network_id})")
        log_change('current_network_info', f"Current network: {network_name}",
                   org_id=org_id, network_id=network_id, network_name=network_name)

        if current_id:
            try:
                curr = dashboard.organizations.getOrganizationConfigTemplate(org_id, current_id)
                print(f"Bound template: {curr.get('name','<unknown>')} (ID: {current_id})\n")
                log_change('bound_template_info',
                           f"Bound template {curr.get('name','<unknown>')} ({current_id})",
                           network_id=network_id, network_name=network_name)
            except Exception:
                print(f"Bound template ID: {current_id}\n")
                log_change('bound_template_info', f"Bound template ID: {current_id}",
                           network_id=network_id, network_name=network_name)
        else:
            print("No template bound.\n")

        temps = dashboard.organizations.getOrganizationConfigTemplates(org_id)
        filtered = temps
        if mx_model_filter in {'MX67', 'MX75'}:
            suffix = mx_model_filter.upper()
            filtered = [t for t in temps if (t.get('name') or '').strip().upper().endswith(suffix)]
            if not filtered:
                print(f"(No templates ending with {suffix}; showing all templates instead.)")
                filtered = temps

        # VLAN-based suggestion
        offered_auto = False
        if not offered_auto:
            offered_auto = True
            auto_choice = _pick_template_by_vlan_count(filtered, vlan_count) or _pick_template_by_vlan_count(temps, vlan_count)
            if auto_choice:
                print(f"Suggested template based on VLAN count ({vlan_count}): {auto_choice['name']} (ID: {auto_choice['id']})")
                resp = input("Use this template? (Y/n): ").strip().lower()
                if resp in {"", "y", "yes"}:
                    try:
                        if current_id:
                            do_action(dashboard.networks.unbindNetwork, network_id)
                        do_action(dashboard.networks.bindNetwork, network_id, configTemplateId=auto_choice['id'])
                        log_change('template_bind',
                                   f"Auto-bound (confirmed) to template {auto_choice['name']} (ID: {auto_choice['id']})",
                                   device_name=network_name, network_id=network_id, network_name=network_name)
                        print(f"✅ Bound to {auto_choice['name']}")
                        return auto_choice['id'], auto_choice['name'], False
                    except meraki.APIError as e:
                        logging.error(f"Error binding suggested template: {e}")
                        must_rollback = True if current_id else False
                        if is_vlans_disabled_error(e):
                            print("❌ VLANs are disabled for this network. Binding failed.")
                            must_rollback = True
                        if must_rollback:
                            print("🚨 Initiating rollback due to failed bind...")
                            rollback_all_changes(
                                network_id=network_id,
                                pre_change_devices=pre_change_devices or [],
                                pre_change_vlans=pre_change_vlans or [],
                                pre_change_template=pre_change_template,
                                org_id=org_id,
                                claimed_serials=claimed_serials or [],
                                removed_serials=removed_serials or [],
                                ms_list=ms_list or [],
                                network_name=network_name,
                            )
                            return current_id, None, True
                        print("Auto-bind failed; falling back to manual selection.\n")
                    except Exception as e:
                        logging.error(f"Unexpected error during auto-bind: {e}")
                        if current_id:
                            print("🚨 Unexpected error after unbind — initiating rollback...")
                            rollback_all_changes(
                                network_id=network_id,
                                pre_change_devices=pre_change_devices or [],
                                pre_change_vlans=pre_change_vlans or [],
                                pre_change_template=pre_change_template,
                                org_id=org_id,
                                claimed_serials=claimed_serials or [],
                                removed_serials=removed_serials or [],
                                ms_list=ms_list or [],
                                network_name=network_name,
                            )
                            return current_id, None, True
                        print("Auto-bind failed; falling back to manual selection.\n")
                else:
                    print("Okay — we’ll choose a different template from the list below.\n")

        for i, t in enumerate(filtered, 1):
            print(f"{i}. {t['name']} (ID: {t['id']})")

        sel = input(
            "Select template # (or press Enter / type 'skip'/'cancel' to cancel — a second cancel will ROLLBACK): "
        ).strip().lower()

        if sel in {"", "skip", "cancel"}:
            skip_attempts += 1
            if skip_attempts == 1:
                print("⚠️  You chose to cancel template selection.")
                print("If you cancel again, the process will be ROLLED BACK immediately.")
                continue
            print("🚨 Cancelled twice — initiating rollback...")
            log_change('rollback_trigger', 'User cancelled twice during template selection')
            rollback_all_changes(
                network_id=network_id,
                pre_change_devices=pre_change_devices or [],
                pre_change_vlans=pre_change_vlans or [],
                pre_change_template=pre_change_template,
                org_id=org_id,
                claimed_serials=claimed_serials or [],
                removed_serials=removed_serials or [],
                ms_list=ms_list or [],
                network_name=network_name,
            )
            return current_id, None, True

        if not sel.isdigit():
            print("Invalid selection. Please enter a valid number or press Enter to cancel.")
            continue

        idx = int(sel) - 1
        if idx < 0 or idx >= len(filtered):
            print("Invalid template number.")
            continue

        chosen = filtered[idx]
        if chosen['id'] == current_id:
            print("No change (already bound to that template).")
            return current_id, chosen['name'], False

        try:
            if current_id:
                do_action(dashboard.networks.unbindNetwork, network_id)
            do_action(dashboard.networks.bindNetwork, network_id, configTemplateId=chosen['id'])
            log_change('template_bind',
                       f"Bound to template {chosen['name']} (ID: {chosen['id']})",
                       device_name=network_name, network_id=network_id, network_name=network_name)
            print(f"✅ Bound to {chosen['name']}")
            return chosen['id'], chosen['name'], False

        except meraki.APIError as e:
            logging.error(f"Error binding template: {e}")
            must_rollback = True if current_id else False
            if is_vlans_disabled_error(e):
                print("❌ VLANs are not enabled for this network. Binding failed and state may be partial.")
                must_rollback = True
            if must_rollback:
                print("🚨 Initiating rollback due to failed bind...")
                rollback_all_changes(
                    network_id=network_id,
                    pre_change_devices=pre_change_devices or [],
                    pre_change_vlans=pre_change_vlans or [],
                    pre_change_template=pre_change_template,
                    org_id=org_id,
                    claimed_serials=claimed_serials or [],
                    removed_serials=removed_serials or [],
                    ms_list=ms_list or [],
                    network_name=network_name,
                )
                return current_id, None, True
            print(f"❌ Failed to bind template: {e}. You can try again or cancel.")
            continue

        except Exception as e:
            logging.error(f"Unexpected error during bind: {e}")
            if current_id:
                print("🚨 Unexpected error after unbind — initiating rollback...")
                rollback_all_changes(
                    network_id=network_id,
                    pre_change_devices=pre_change_devices or [],
                    pre_change_vlans=pre_change_vlans or [],
                    pre_change_template=pre_change_template,
                    org_id=org_id,
                    claimed_serials=claimed_serials or [],
                    removed_serials=removed_serials or [],
                    ms_list=ms_list or [],
                    network_name=network_name,
                )
                return current_id, None, True
            print(f"❌ Unexpected error: {e}. You can try again or cancel.")
            continue

def bind_network_to_template(
    org_id: str,
    network_id: str,
    tpl_id: Optional[str],
    vlan_list: List[Dict[str, Any]],
    network_name: str,
    *,
    pre_change_devices,
    pre_change_vlans,
    pre_change_template,
    claimed_serials,
    removed_serials,
    ms_list
):
    if not tpl_id:
        return
    time.sleep(5)

    enabled = vlans_enabled(network_id)
    if enabled is False:
        print("❌ VLANs are disabled on this network after binding. Rolling back immediately...")
        rollback_all_changes(
            network_id=network_id,
            pre_change_devices=pre_change_devices or [],
            pre_change_vlans=pre_change_vlans or [],
            pre_change_template=pre_change_template,
            org_id=org_id,
            claimed_serials=claimed_serials or [],
            removed_serials=removed_serials or [],
            ms_list=ms_list or [],
            network_name=network_name,
        )
        log_change('workflow_end', 'Exited after rollback due to VLANs disabled (pre-check)')
        raise SystemExit(1)

    try:
        update_vlans(network_id, network_name, vlan_list)
    except meraki.APIError as e:
        if is_vlans_disabled_error(e):
            print("❌ VLANs disabled error during VLAN update. Rolling back immediately...")
            rollback_all_changes(
                network_id=network_id,
                pre_change_devices=pre_change_devices or [],
                pre_change_vlans=pre_change_vlans or [],
                pre_change_template=pre_change_template,
                org_id=org_id,
                claimed_serials=claimed_serials or [],
                removed_serials=removed_serials or [],
                ms_list=ms_list or [],
                network_name=network_name,
            )
            log_change('workflow_end', 'Exited after rollback due to VLANs disabled during VLAN update')
            raise SystemExit(1)
        raise

def select_switch_profile_interactive_by_model(tpl_profiles: List[Dict[str, Any]], tpl_profile_map: Dict[str, str], switch_model: str) -> Optional[str]:
    candidates = [p for p in tpl_profiles if switch_model in p.get('model', [])]
    if not candidates:
        print(f"No switch profiles in template support {switch_model}.")
        return None
    print(f"\nAvailable switch profiles for {switch_model}:")
    for idx, p in enumerate(candidates, 1):
        print(f"{idx}. {p['name']}")
    profile_names = [p['name'] for p in candidates]
    while True:
        choice = input("Select switch profile by number (or Enter to skip): ").strip()
        if not choice:
            return None
        if choice.isdigit():
            idx = int(choice) - 1
            if 0 <= idx < len(profile_names):
                return tpl_profile_map[profile_names[idx]]
        print("Invalid selection. Please try again.")

def device_in_inventory(org_id: str, serial: str) -> bool:
    try:
        inv = dashboard.organizations.getOrganizationInventoryDevice(org_id, serial)
        return inv.get('networkId') is None
    except Exception:
        return False

# ------------- Rollback -------------
def rollback_all_changes(
    network_id: str,
    pre_change_devices: List[Dict[str, Any]],
    pre_change_vlans: List[Dict[str, Any]],
    pre_change_template: Optional[str],
    org_id: str,
    *,
    claimed_serials: Optional[List[str]] = None,
    removed_serials: Optional[List[str]] = None,
    ms_list: Optional[List[Dict[str, Any]]] = None,
    network_name: str,
):
    print("=== Starting rollback to previous network state ===")

    if claimed_serials:
        for serial in claimed_serials:
            try:
                do_action(dashboard.networks.removeNetworkDevices, network_id, serial)
                log_change('rollback_device_removed', "Removed claimed device in rollback", device_serial=serial)
            except Exception:
                logging.exception("Failed to remove claimed device during rollback: %s", serial)

    if removed_serials:
        try:
            do_action(dashboard.networks.claimNetworkDevices, network_id, serials=removed_serials)
            for serial in removed_serials:
                log_change('rollback_device_reclaimed', f"Re-claimed previously removed device", device_serial=serial)
        except Exception:
            logging.exception("Failed to re-claim devices during rollback")

    print("Restoring config template binding...")
    try:
        do_action(dashboard.networks.unbindNetwork, network_id)
        if pre_change_template:
            do_action(dashboard.networks.bindNetwork, network_id, configTemplateId=pre_change_template)
        log_change('rollback_template', f"Restored template binding {pre_change_template}", device_name=f"Network: {network_id}")
    except Exception:
        logging.exception("Failed to restore original template binding")

    print("Waiting for template binding to take effect (sleeping 15 seconds)...")
    time.sleep(15)

    current_devices = dashboard.networks.getNetworkDevices(network_id)
    current_serials = {d['serial'] for d in current_devices}

    # Re-add missing devices from snapshot if available in org inventory and not assigned
    for dev in pre_change_devices:
        if dev["serial"] not in current_serials:
            try:
                inv = dashboard.organizations.getOrganizationInventoryDevice(org_id, dev['serial'])
                if not inv.get('networkId'):
                    print(f"Re-adding device {dev['serial']} ({dev['model']}) to network...")
                    do_action(dashboard.networks.claimNetworkDevices, network_id, serials=[dev["serial"]])
                    log_change('rollback_device_readded', f"Device re-added during rollback", device_serial=dev['serial'])
                else:
                    print(f"Device {dev['serial']} is assigned elsewhere. Skipping.")
            except Exception as e:
                print(f"Could not check/claim device {dev['serial']}: {e}")

    current_devices = dashboard.networks.getNetworkDevices(network_id)
    current_serials = {d['serial'] for d in current_devices}

    try:
        restored_tpl_profiles = dashboard.switch.getOrganizationConfigTemplateSwitchProfiles(org_id, pre_change_template) if pre_change_template else []
        profile_id_set = {p['switchProfileId'] for p in restored_tpl_profiles}
        profile_name_to_id = {p['name']: p['switchProfileId'] for p in restored_tpl_profiles}
    except Exception:
        logging.exception("Could not fetch switch profiles for restored template")
    # default if error:
        restored_tpl_profiles = []
        profile_id_set = set()
        profile_name_to_id = {}

    for dev in pre_change_devices:
        if dev["serial"] not in current_serials:
            continue

        update_args: Dict[str, Any] = {"name": dev.get("name", ""), "address": dev.get("address", ""), "tags": dev.get("tags", [])}
        if dev["model"].startswith("MS"):
            serial = dev["serial"]
            orig_profile_id = dev.get('switchProfileId')
            if orig_profile_id and orig_profile_id in profile_id_set:
                print(f"Auto-restoring MS {serial} to profile ID {orig_profile_id}")
                update_args["switchProfileId"] = orig_profile_id
            else:
                orig_profile_name = dev.get('switchProfileName')
                new_profile_id = profile_name_to_id.get(orig_profile_name)
                if new_profile_id:
                    print(f"Auto-restoring MS {serial} to profile '{orig_profile_name}' (ID: {new_profile_id})")
                    update_args["switchProfileId"] = new_profile_id

        try:
            do_action(dashboard.devices.updateDevice, dev['serial'], **update_args)
            log_change(
                'rollback_device_update',
                f"Restored device config during rollback",
                device_serial=dev['serial'],
                device_name=dev.get('name', ''),
                misc=f"tags={dev.get('tags', [])}, address={dev.get('address', '')}"
            )
        except Exception:
            logging.exception(f"Failed to update device {dev['serial']} during rollback")
            continue

        if dev["model"].startswith("MS"):
            try:
                preserved = (dev.get('port_overrides') or {})
                if preserved:
                    apply_port_overrides(dev['serial'], preserved)
            except Exception:
                logging.exception(f"Failed applying preserved port overrides during rollback for {dev['serial']}")

    print("Restoring VLANs and DHCP assignments...")
    time.sleep(5)
    update_vlans(network_id, network_name, pre_change_vlans)
    log_change('rollback_vlans', "Restored VLANs and DHCP assignments", device_name=f"Network: {network_id}")

    print("=== Rollback complete ===")

# ------------- Step Summary helpers -------------
StatusVal = Union[bool, str]  # True/False/"NA"

def _fmt(val: StatusVal) -> str:
    if val is True:
        return "✅ Success"
    if val is False:
        return "❌ Failed"
    return str(val)

def print_summary(step_status: Dict[str, StatusVal]) -> None:
    order = [
        'template_bound',
        'vlans_updated',
        'devices_claimed',
        'mx_removed',
        'mr33_removed',
        'configured',
        'old_mx',
        'old_mr33',
    ]
    print("\nStep Summary:")
    for step in order:
        val = step_status.get(step, "NA")
        if isinstance(val, str) and val.upper() == "NA":
            continue
        print(f" - {step}: {_fmt(val)}")

def get_profileid_to_name(org_id: str, template_id: Optional[str]) -> Dict[str, str]:
    if not template_id:
        return {}
    try:
        profs = dashboard.switch.getOrganizationConfigTemplateSwitchProfiles(org_id, template_id) or []
        return {p.get("switchProfileId"): p.get("name", "") for p in profs if p.get("switchProfileId")}
    except Exception:
        logging.exception("Failed fetching template switch profiles for ID->name mapping")
        return {}

# ------------- Wireless pre-check and filtering -------------

def run_wireless_precheck_and_filter_claims(
    org_id: str,
    network_id: str,
    prevalidated_serials: List[str],
) -> Tuple[List[str], List[str], List[str]]:
    """
    Run the MR33 pre-check/replacement flow.

    Behavior:
      - If there are non-MR33 APs present, the helper will PROMPT the user about replacements.
      - It may remove old APs and claim mapped new ones (with confirmation).
      - It returns serials to claim next, with ONLY those already claimed in this function removed.
        (No blanket wireless filtering — wireless proceeds automatically.)

    Returns:
        safe_to_claim, mr_removed_serials, mr_claimed_serials
    """
    mr_removed_serials: List[str] = []
    mr_claimed_serials: List[str] = []

    try:
        updated_serials, mr_removed_serials, mr_claimed_serials = ensure_mr33_and_handle_wireless_replacements(
            org_id, network_id, prevalidated_serials
        )
    except SystemExit:
        raise
    except Exception:
        logging.exception("Wireless pre-check/replacement step failed")
        updated_serials = prevalidated_serials  # fall back

    # IMPORTANT: Do NOT filter out wireless here.
    # Only avoid double-claiming anything already claimed in the helper.
    safe_to_claim = [s for s in updated_serials if s not in set(mr_claimed_serials)]
    return safe_to_claim, mr_removed_serials, mr_claimed_serials


def cleanup_after_claims(
    org_id: str,
    network_id: str,
    network_name: str,
    *,
    claimed: List[str],
    old_mx_devices: List[Dict[str, Any]],
    old_mr_devices: List[Dict[str, Any]],
    tpl_profile_map: Dict[str, str],
    ms_list: Optional[List[Dict[str, Any]]] = None,
    mr_list: Optional[List[Dict[str, Any]]] = None,
    primary_mx_serial: Optional[str] = None,
    mr_order: Optional[List[str]] = None,
    ms_order: Optional[List[str]] = None,
    step_status: Optional[Dict[str, StatusVal]] = None,
) -> None:
    """
    After new devices are claimed, perform:
      - MX64 removal if newer MX was claimed
      - MR33 removal
      - naming & config of claimed devices (MX/MR/MS)
      - 'recently-added' tag removal
      - update step_status keys

    All failures are logged; step_status is updated defensively.
    """

    if step_status is None:
        step_status = {}

    # Refresh device lists if not provided
    if ms_list is None or mr_list is None:
        try:
            _, ms_list_fresh, mr_list_fresh = fetch_devices(org_id, network_id)
            if ms_list is None:
                ms_list = ms_list_fresh
            if mr_list is None:
                mr_list = mr_list_fresh
        except Exception:
            logging.exception("Failed to refresh device lists post-claim")
            ms_list = ms_list or []
            mr_list = mr_list or []

    # Determine if any newly-claimed device is a newer MX
    try:
        mx_models = []
        for s in claimed:
            try:
                inv = dashboard.organizations.getOrganizationInventoryDevice(org_id, s)
                mx_models.append(inv.get('model', '') or '')
            except Exception:
                pass
        if any(m.startswith('MX67') or m.startswith('MX75') for m in mx_models):
            remove_existing_mx64_devices(org_id, network_id)
            log_change('mx_removed', "Removed old MX64 after new MX claim", misc=f"claimed_serials={claimed}")
        step_status['mx_removed'] = True
    except Exception:
        logging.exception("MX64 removal stage failed")
        step_status['mx_removed'] = False

    # Remove legacy MR33 after new AP claim
    try:
        mr33_ok = remove_existing_MR33_devices(org_id, network_id)
        step_status['mr33_removed'] = mr33_ok
        if mr33_ok:
            log_change('mr33_removed', "Removed old MR33 after new AP claim", misc=f"claimed_serials={claimed}")
    except Exception:
        logging.exception("MR33 removal stage failed")
        step_status['mr33_removed'] = False

    # Naming & configuration for claimed devices
    try:
        name_and_configure_claimed_devices(
            org_id,
            network_id,
            network_name,
            claimed,
            ms_list or [],
            mr_list or [],
            tpl_profile_map,
            old_mx_devices=old_mx_devices,
            old_mr_devices=old_mr_devices,
            primary_mx_serial=primary_mx_serial,
            mr_order=mr_order,
            ms_order=ms_order,
        )
        step_status['configured'] = True
    except Exception:
        logging.exception("Configuration stage failed")
        step_status['configured'] = False

    # Remove the 'recently-added' tag from any devices
    try:
        remove_recently_added_tag(network_id)
    except Exception:
        logging.exception("Failed removing 'recently-added' tag")

# ------------- Excel snapshot export -------------

def _slug_filename(s: str) -> str:
    # keep letters, numbers, dot, underscore, hyphen; replace others with '-'
    s = re.sub(r'[^A-Za-z0-9._-]+', '-', s).strip('-_')
    return s[:80]  # keep it tidy

def _network_tag_from_name(name: str) -> str:
    # If your convention is "UK-0593-Whatever", return "UK-0593"
    parts = name.split('-')
    if len(parts) >= 2 and parts[1].isdigit():
        return f"{parts[0]}-{parts[1]}"
    return name

def _network_number_from_name(name: str) -> str | None:
    # Grab the first standalone number block (e.g., 0593)
    m = re.search(r'\b(\d{2,8})\b', name)
    return m.group(1) if m else None

def export_network_snapshot_xlsx(
    org_id: str,
    network_id: str,
    network_name: str,
    template_id: Optional[str],
    vlan_list: List[Dict[str, Any]],
    mx_list: List[Dict[str, Any]],
    ms_list: List[Dict[str, Any]],
    mr_list: List[Dict[str, Any]],
    profileid_to_name: Optional[Dict[str, str]] = None,
    outfile: Optional[str] = None,
    filename_mode: str = "name",  # NEW: "name" | "number"
) -> None:
    def _json(x: Any) -> str:
        try:
            return json.dumps(x, ensure_ascii=False)
        except Exception:
            return str(x)

    # --- choose default filename if caller didn't pass one ---
    if outfile:
        out_path: str = outfile
    else:
        if filename_mode == "number":
            base = _network_number_from_name(network_name) or _network_tag_from_name(network_name)
        else:  # "name"
            base = _network_tag_from_name(network_name)
        out_path = f"{_slug_filename(_network_tag_from_name(network_name))}_pre_{timestamp}.xlsx"
        
    wb: Workbook = Workbook()
    ws: Worksheet = cast(Worksheet, wb.active)
    ws.title = "Snapshot"

    header: List[str] = [
        "section", "network_id", "network_name", "item_type",
        "col1", "col2", "col3", "col4", "col5",
        "switch_profile_id", "switch_profile_name", "extra_info"
    ]
    ws.append(header)

    tpl_name: str = ""
    if template_id:
        try:
            tpl = dashboard.organizations.getOrganizationConfigTemplate(org_id, template_id)
            tpl_name = str(tpl.get("name", "") or "")
        except Exception:
            logging.exception("Could not fetch template name for snapshot")

    ws.append([
        "template", network_id, network_name, "template",
        template_id or "", tpl_name, "", "", "",
        "", "", ""
    ])

    for v in vlan_list:
        ws.append([
            "vlans", network_id, network_name, "vlan",
            str(v.get("id", "")),
            str(v.get("name", "") or ""),
            str(v.get("subnet", "") or ""),
            str(v.get("applianceIp", "") or ""),
            str(v.get("dhcpHandling", "") or ""),
            "", "",
            _json({k: v.get(k) for k in v.keys() - {"id", "name", "subnet", "applianceIp", "dhcpHandling"}}),
        ])

    def _device_row(d: Dict[str, Any]) -> List[str]:
        tags_val = d.get("tags", [])
        tags_list: List[str]
        if isinstance(tags_val, list):
            tags_list = [str(t) for t in tags_val]
        else:
            tags_list = [t for t in str(tags_val or "").split() if t]

        sp_id: str = str(d.get("switchProfileId", "") or "")
        sp_name: str = str(d.get("switchProfileName", "") or "")
        if (not sp_name) and sp_id and profileid_to_name:
            sp_name = profileid_to_name.get(sp_id, "") or ""

        return [
            "devices", network_id, network_name, "device",
            str(d.get("serial", "") or ""),
            str(d.get("model", "") or ""),
            str(d.get("name", "") or ""),
            str(d.get("address", "") or ""),
            " ".join(tags_list),
            sp_id,
            sp_name,
            ""
        ]

    for d in (mx_list + ms_list + mr_list):
        ws.append(_device_row(d))

    for sw in ms_list:
        po_val = sw.get("port_overrides") or {}
        changes_by_port: Dict[str, Dict[str, Any]] = po_val if isinstance(po_val, dict) else {}
        if not changes_by_port:
            continue
        for port_id, changes in changes_by_port.items():
            if not isinstance(changes, dict):
                continue
            for fld, val in changes.items():
                ws.append([
                    "port_overrides", network_id, network_name, "port_override",
                    str(sw.get("serial", "") or ""), str(port_id), str(fld),
                    "" if isinstance(val, (dict, list)) else str(val),
                    "", "", "",
                    _json(val) if isinstance(val, (dict, list)) else "",
                ])

    max_col: int = ws.max_column
    max_row: int = ws.max_row
    for col_idx in range(1, max_col + 1):
        max_len = 0
        for row_idx in range(1, max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                s = str(val)
                if len(s) > max_len:
                    max_len = len(s)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    wb.save(out_path)
    print(f"📄 Snapshot exported to Excel: {out_path}")
    log_change("snapshot_export", f"Exported network snapshot to {out_path}", network_id=network_id, network_name=network_name)

def _sanitize_port_patch_for_new_template(patch: Dict[str, Any]) -> Dict[str, Any]:
    """On error, remove/adjust fields that are commonly non-portable across templates."""
    cleaned = dict(patch)
    # These often break after rebind; remove them if they cause 4xx
    for k in ('portScheduleId',):
        cleaned.pop(k, None)
    # If access policy mismatches, fall back to open
    if cleaned.get('accessPolicyType') or cleaned.get('accessPolicyNumber') is not None:
        cleaned.pop('accessPolicyNumber', None)
        cleaned['accessPolicyType'] = 'Open'
    return cleaned

def apply_port_overrides_with_retry(serial: str, overrides: Dict[str, Dict[str, Any]], retries: int = 5, delay: int = 3) -> None:
    for pid, patch in (overrides or {}).items():
        attempt = 0
        sanitized_once = False
        while True:
            try:
                do_action(dashboard.switch.updateDeviceSwitchPort, serial, pid, **patch)
                log_change('switch_port_override', f"Applied port overrides on port {pid}",
                           device_serial=serial, misc=json.dumps(patch))
                break
            except Exception as e:
                attempt += 1
                # Try one sanitation pass if we hit a 4xx-style error
                if not sanitized_once:
                    patch = _sanitize_port_patch_for_new_template(patch)
                    sanitized_once = True
                if attempt >= retries:
                    logging.exception(f"Failed applying port overrides on {serial} port {pid} after {retries} tries: {e}")
                    break
                logging.debug(f"Retrying {serial} port {pid} in {delay}s due to: {e}")
                time.sleep(delay)

def export_post_change_snapshot(org_id: str, network_id: str, network_name: str) -> None:
    """
    Export a post-change snapshot using export_network_snapshot_xlsx().
    File name will follow: <tag>_post_<timestamp>.xlsx
    """
    try:
        net = dashboard.networks.getNetwork(network_id)
        current_template = net.get('configTemplateId')
    except Exception:
        logging.exception("Failed to read current template for post-change snapshot")
        current_template = None

    # Re-fetch live state after changes
    mx_list, ms_list, mr_list = fetch_devices(org_id, network_id, template_id=current_template)
    vlan_list = fetch_vlan_details(network_id)
    profileid_to_name = get_profileid_to_name(org_id, current_template)

    outfile = f"{_slug_filename(_network_tag_from_name(network_name))}_post_{timestamp}.xlsx"

    export_network_snapshot_xlsx(
        org_id=org_id,
        network_id=network_id,
        network_name=network_name,
        template_id=current_template,
        vlan_list=vlan_list,
        mx_list=mx_list,
        ms_list=ms_list,
        mr_list=mr_list,
        profileid_to_name=profileid_to_name,
        outfile=outfile,
    )
    log_change("snapshot_export_post", f"Exported post-change snapshot to {outfile}",
               network_id=network_id, network_name=network_name)

# ------------- Selectors (org/network) -------------
# ---- UPDATED SELECTOR TO LEVERAGE ROBUST MATCHING (accepts ID too) ----
def select_network_interactive(org_id: str) -> Tuple[str, str]:
    while True:
        partial: str = input("Enter partial network name or network ID (or press Enter to cancel): ").strip()
        if not partial:
            print("\n❌ No Network selected -----------\n   Please retry when Network is known *******")
            sys.exit(1)

        networks: List[Dict[str, Any]] = fetch_matching_networks(org_id, partial)

        if not networks:
            print("\n❌ No matching networks found -----------")
            retry = input("Search again? (y/N): ").strip().lower()
            if retry == "y":
                continue
            print("\n❌ No Network selected -----------\n   Please retry when Network is known *******")
            sys.exit(1)

        if len(networks) == 1:
            only = networks[0]
            print(f"\n1 match: {only['name']} (ID: {only['id']})")
            confirm = input("Use this network? (Y/n): ").strip().lower()
            if confirm in {"", "y", "yes"}:
                print(f"Selected network: {only['name']} (ID: {only['id']})")
                return cast(str, only['id']), cast(str, only['name'])
            continue

        # Multiple matches — show a tidy, deterministic order
        networks_sorted = sorted(networks, key=lambda n: _normalize_name(str(n.get("name", ""))))
        print("\nMultiple networks found:")
        for idx, net in enumerate(networks_sorted, 1):
            print(f"{idx}. {net['name']} (ID: {net['id']})")

        while True:
            raw = input("Select the network by number (or press Enter to cancel): ").strip()
            if not raw:
                print("\n❌ No Network selected -----------\n   Please retry when Network is known *******")
                sys.exit(1)
            if raw.isdigit():
                choice = int(raw)
                if 1 <= choice <= len(networks_sorted):
                    chosen = networks_sorted[choice - 1]
                    print(f"Selected network #{choice}: {chosen['name']} (ID: {chosen['id']})")
                    return cast(str, chosen['id']), cast(str, chosen['name'])
            print("❌ Invalid selection. Please enter a valid number from the list.")

def select_org() -> str:
    orgs = dashboard.organizations.getOrganizations()
    if not orgs:
        print("\n❌ No Organisations returned from API -----------")
        print("   Please retry when Org is known *******")
        sys.exit(1)

    print("Organizations:")
    for idx, org in enumerate(orgs, 1):
        print(f"{idx}. {org['name']} (ID: {org['id']})")

    raw = input("Select organization by number (or press Enter to cancel): ").strip()
    if not raw:
        print("\n❌ No Organisation selected -----------")
        print("   Please retry when Org is known *******")
        sys.exit(1)

    try:
        org_idx = int(raw)
        if org_idx < 1 or org_idx > len(orgs):
            raise ValueError("out of range")
    except Exception:
        print("\n❌ Invalid Organisation selection -----------")
        print("   Please retry when Org is known *******")
        sys.exit(1)

    return orgs[org_idx - 1]['id']

# ------------- Change Rollback Font -------------
def prompt_rollback_big() -> str:
    # Import locally so names are always bound for Pylance,
    # and gracefully fall back if Rich/pyfiglet isn't installed.
    try:
        from rich.console import Console
        from rich.panel import Panel
        from rich.text import Text
        import pyfiglet
    except Exception:
        return prompt_rollback_basic()

    console = Console()
    banner = pyfiglet.figlet_format("ROLLBACK", font="slant")
    console.print(f"[bold red]{banner}[/bold red]")
    console.print(Panel.fit(
        Text(
            "Type 'yes' to rollback changes, 'no' to continue without rollback, or just press Enter to skip.\n"
            "IMPORTANT: If you skip (press Enter), rollback will no longer be available.\n"
            "Have you ensured the network is fully functional and all required checks have been carried out?",
            style="bold white"
        ),
        title="⚠️  ROLLBACK OPTION",
        title_align="left",
        border_style="red"
    ))
    return input("> ").strip().lower()

def prompt_rollback_basic() -> str:
    print("\n" + "!"*78)
    print("⚠️  R O L L B A C K   O P T I O N  ⚠️".center(78))
    print("!"*78)
    print("Type 'yes' to rollback changes, 'no' to continue without rollback, or just press Enter to skip.")
    print("IMPORTANT: If you skip (Enter), rollback will no longer be available.")
    print("Have you ensured the network is fully functional and all required checks have been carried out?")
    return input("> ").strip().lower()

# ------------- Main -------------
if __name__ == '__main__':
    log_change('workflow_start', 'Script started')

    step_status: Dict[str, StatusVal] = {}

    org_id = select_org()

    prevalidated_serials = prompt_and_validate_serials(org_id)
    detected_mx_models = summarize_devices_in_org(org_id, prevalidated_serials)

    mx_model_filter = None
    if detected_mx_models == {'MX67'}:
        mx_model_filter = 'MX67'
    elif detected_mx_models == {'MX75'}:
        mx_model_filter = 'MX75'

    network_id, network_name = select_network_interactive(org_id)
    net_info = dashboard.networks.getNetwork(network_id)
    old_template = net_info.get('configTemplateId')

    # Pre-change snapshot incl. MS port overrides
    mx, ms, mr = fetch_devices(org_id, network_id, template_id=old_template)
    pre_change_devices = mx + ms + mr
    pre_change_vlans = fetch_vlan_details(network_id)
    pre_change_template = old_template
    pre_change_serials = {d['serial'] for d in pre_change_devices}
    profileid_to_name = get_profileid_to_name(org_id, old_template)

    # Export PRE snapshot with new filename format
    export_network_snapshot_xlsx(
        org_id=org_id,
        network_id=network_id,
        network_name=network_name,
        template_id=old_template,
        vlan_list=pre_change_vlans,
        mx_list=mx,
        ms_list=ms,
        mr_list=mr,
        profileid_to_name=profileid_to_name,
        outfile=f"{_slug_filename(_network_tag_from_name(network_name))}_pre_{timestamp}.xlsx",
    )

    current_mx_models = sorted({d['model'] for d in mx})
    is_mx64_present = any(m.startswith('MX64') for m in current_mx_models)

    # ------------------------------------------------------------------
    # PATH A: Current network has MX (not MX64) -> light flow (no rebind)
    # ------------------------------------------------------------------
    if current_mx_models and not is_mx64_present:
        print(f"\nCurrent network: {network_name} (ID: {network_id})")
        if old_template:
            try:
                curr_tpl = dashboard.organizations.getOrganizationConfigTemplate(org_id, old_template)
                print(f"Bound template: {curr_tpl.get('name','<unknown>')} (ID: {old_template})")
            except Exception:
                print(f"Bound template ID: {old_template}")
        else:
            print("No template bound.")
        print(f"Detected MX model(s): {', '.join(current_mx_models)}")

        step_status['template_bound'] = "NA"
        step_status['vlans_updated'] = "NA"
        step_status['mx_removed'] = "NA"

        # --- wireless MR33 pre-check (prompts for non-MR33 removals) ---
        safe_to_claim, mr_removed_serials, mr_claimed_serials = run_wireless_precheck_and_filter_claims(
            org_id, network_id, prevalidated_serials
        )

        # Claim remaining devices (wireless included), excluding those already claimed in helper
        claimed = claim_devices(org_id, network_id, prevalidated_serials=safe_to_claim)
        step_status['devices_claimed'] = bool(claimed)

        # Order / primary selection
        primary_mx_serial = select_primary_mx(org_id, claimed)
        ensure_primary_mx(network_id, primary_mx_serial)
        mr_order = select_device_order(org_id, claimed, 'MR')
        ms_order = select_device_order(org_id, claimed, 'MS')

        # Template profiles (if any) for post-claim config
        try:
            if old_template:
                tpl_profiles = dashboard.switch.getOrganizationConfigTemplateSwitchProfiles(org_id, old_template)
                tpl_profile_map = {p['name']: p['switchProfileId'] for p in tpl_profiles}
            else:
                tpl_profile_map = {}
                tpl_profiles = []
        except Exception:
            logging.exception("Failed fetch template switch profiles")
            tpl_profile_map = {}
            tpl_profiles = []

        # Unified cleanup/config (also handles MX64/MR33 removals and naming)
        cleanup_after_claims(
            org_id,
            network_id,
            network_name,
            claimed=claimed,
            old_mx_devices=mx,
            old_mr_devices=mr,
            tpl_profile_map=tpl_profile_map,
            ms_list=ms,
            mr_list=mr,
            primary_mx_serial=primary_mx_serial,
            mr_order=mr_order,
            ms_order=ms_order,
            step_status=step_status,
        )

        # Export POST snapshot
        export_post_change_snapshot(org_id, network_id, network_name)

        # Deltas for rollback option
        post_change_devices = dashboard.networks.getNetworkDevices(network_id)
        post_change_serials = {d['serial'] for d in post_change_devices}
        claimed_serials_rb = list(post_change_serials - pre_change_serials)
        removed_serials_rb = list(pre_change_serials - post_change_serials)

        print_summary(step_status)

        rollback_choice = prompt_rollback_big()

        if rollback_choice in {'yes', 'y'}:
            print("\nRolling back all changes...")
            log_change('rollback_start', 'User requested rollback')
            rollback_all_changes(
                network_id,
                pre_change_devices,
                pre_change_vlans,
                pre_change_template,
                org_id,
                claimed_serials=claimed_serials_rb,
                removed_serials=removed_serials_rb,
                ms_list=ms,
                network_name=network_name,
            )
            print("✅ Rollback complete.")
            log_change('rollback_end', 'Rollback completed')

        elif rollback_choice in {'no', 'n'}:
            print("\nProceeding without rollback. Rollback option will no longer be available.")
            log_change('workflow_end', 'Script finished (no rollback)')

        else:
            print("\n❌ No rollback selected (Enter pressed).")
            print("⚠️  Rollback is no longer available. Please ensure the network is functional and all required checks have been carried out.")
            log_change('workflow_end', 'Script finished (rollback skipped with Enter)')

        raise SystemExit(0)

    # ------------------------------------------------------------------
    # PATH B: MX64 present -> full rebind/VLAN flow
    # ------------------------------------------------------------------
    vlan_list = fetch_vlan_details(network_id)
    old_mx, prebind_ms_devices, old_mr = fetch_devices(org_id, network_id, template_id=old_template)
    ms_serial_to_profileid = {sw['serial']: sw.get('switchProfileId') for sw in prebind_ms_devices}

    # After: old_mx, prebind_ms_devices, old_mr = fetch_devices(org_id, network_id, template_id=old_template)
    prebind_overrides_by_serial = {
        sw['serial']: (sw.get('port_overrides') or {})
        for sw in prebind_ms_devices
    }

    if old_template:
        try:
            old_tpl_profiles = dashboard.switch.getOrganizationConfigTemplateSwitchProfiles(org_id, old_template)
            old_profileid_to_name = {p['switchProfileId']: p['name'] for p in old_tpl_profiles}
        except Exception:
            logging.exception("Failed fetching old template switch profiles")
            old_profileid_to_name = {}
    else:
        old_profileid_to_name = {}

    # Choose & (re)bind template (with rollback on failure)
    try:
        new_template, _, rolled_back = list_and_rebind_template(
            org_id,
            network_id,
            old_template,
            network_name,
            pre_change_devices=pre_change_devices,
            pre_change_vlans=pre_change_vlans,
            pre_change_template=pre_change_template,
            claimed_serials=[],
            removed_serials=[],
            ms_list=ms,
            mx_model_filter=mx_model_filter,
            vlan_count=len(vlan_list),
        )
        if rolled_back:
            log_change('workflow_end', 'Exited after rollback during template stage')
            print("Rollback complete. Exiting.")
            raise SystemExit(1)
        step_status['template_bound'] = (new_template is not None) and (new_template != old_template)
    except SystemExit:
        raise
    except Exception:
        logging.exception("Template bind failed")
        new_template = old_template
        step_status['template_bound'] = False

    # Validate VLANs after bind + update VLANs
    try:
        bind_network_to_template(
            org_id, network_id, new_template, vlan_list, network_name,
            pre_change_devices=pre_change_devices,
            pre_change_vlans=pre_change_vlans,
            pre_change_template=pre_change_template,
            claimed_serials=[],
            removed_serials=[],
            ms_list=ms
        )
        step_status['vlans_updated'] = True
    except SystemExit:
        raise
    except Exception:
        logging.exception("VLAN update failed")
        step_status['vlans_updated'] = False

    # Fetch new template profiles for post-bind MS mapping
    try:
        tpl_profiles = dashboard.switch.getOrganizationConfigTemplateSwitchProfiles(org_id, new_template) if new_template else []
        tpl_profile_map = {p['name']: p['switchProfileId'] for p in tpl_profiles}
    except Exception:
        logging.exception("Failed fetch template switch profiles")
        tpl_profile_map = {}
        tpl_profiles = []

    # Re-assign switch profiles to match previous names / user choice
    _, postbind_ms_devices, _ = fetch_devices(org_id, network_id)
    for sw in postbind_ms_devices:
        old_profile_id = ms_serial_to_profileid.get(sw['serial'])
        old_profile_name = old_profileid_to_name.get(old_profile_id)
        new_profile_id = tpl_profile_map.get(old_profile_name) if old_profile_name else None
        if not new_profile_id:
            new_profile_id = select_switch_profile_interactive_by_model(tpl_profiles, tpl_profile_map, sw['model']) if tpl_profiles else None
            if not new_profile_id:
                continue
     # After computing prebind_overrides_by_serial earlier

        try:
            do_action(dashboard.devices.updateDevice, sw['serial'], switchProfileId=new_profile_id)
            log_change('switch_profile_assign', f"Assigned switchProfileId {new_profile_id} to {sw['serial']}",
                    device_serial=sw['serial'], device_name=sw.get('name', ''),
                    misc=f"profile_name={old_profile_name or ''}")

            time.sleep(3)  # let profile settle

            preserved = prebind_overrides_by_serial.get(sw['serial'], {})
            if preserved:
                apply_port_overrides_with_retry(sw['serial'], preserved, retries=5, delay=3)
        except Exception:
            logging.exception(f"Failed to assign profile / apply overrides to {sw['serial']}")
  
    # --- wireless MR33 pre-check (prompts for non-MR33 removals) ---
    safe_to_claim, mr_removed_serials, mr_claimed_serials = run_wireless_precheck_and_filter_claims(
        org_id, network_id, prevalidated_serials
    )

    # Claim remaining devices (wireless included), excluding those already claimed in helper
    claimed = claim_devices(org_id, network_id, prevalidated_serials=safe_to_claim)
    step_status['devices_claimed'] = bool(claimed)

    # Order / primary selection
    primary_mx_serial = select_primary_mx(org_id, claimed)
    ensure_primary_mx(network_id, primary_mx_serial)
    mr_order = select_device_order(org_id, claimed, 'MR')
    ms_order = select_device_order(org_id, claimed, 'MS')

    # Compute deltas for rollback (after all device changes)
    post_change_devices = dashboard.networks.getNetworkDevices(network_id)
    post_change_serials = {d['serial'] for d in post_change_devices}
    claimed_serials = list(post_change_serials - pre_change_serials)
    removed_serials = list(pre_change_serials - post_change_serials)

    if claimed:
        # Unified cleanup/config
        cleanup_after_claims(
            org_id,
            network_id,
            network_name,
            claimed=claimed,
            old_mx_devices=old_mx,      # from pre-bind snapshot
            old_mr_devices=old_mr,      # from pre-bind snapshot
            tpl_profile_map=tpl_profile_map,
            primary_mx_serial=primary_mx_serial,
            mr_order=mr_order,
            ms_order=ms_order,
            step_status=step_status,
        )

        # Export POST snapshot
        export_post_change_snapshot(org_id, network_id, network_name)
    else:
        step_status.setdefault('mx_removed', "NA")
        step_status.setdefault('mr33_removed', "NA")
        step_status.setdefault('configured', "NA")
        step_status.setdefault('old_mx', "NA")
        step_status.setdefault('old_mr33', "NA")

    print_summary(step_status)

    rollback_choice = prompt_rollback_big()

    if rollback_choice in {'yes', 'y'}:
        print("\nRolling back all changes...")
        log_change('rollback_start', 'User requested rollback')
        rollback_all_changes(
            network_id,
            pre_change_devices,
            pre_change_vlans,
            pre_change_template,
            org_id,
            claimed_serials=claimed_serials,
            removed_serials=removed_serials,
            ms_list=ms,
            network_name=network_name,
        )
        print("✅ Rollback complete.")
        log_change('rollback_end', 'Rollback completed')

    elif rollback_choice in {'no', 'n'}:
        print("\nProceeding without rollback. Rollback option will no longer be available.")
        log_change('workflow_end', 'Script finished (no rollback)')

    else:
        print("\n❌ No rollback selected (Enter pressed).")
        print("⚠️  Rollback is no longer available. Please ensure the network is functional and all required checks have been carried out.")
        log_change('workflow_end', 'Script finished (rollback skipped with Enter)')


