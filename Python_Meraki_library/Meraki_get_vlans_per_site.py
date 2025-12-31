# Created on 2025-09-04 by JS
    # Script goes through entire org and outputs VLAN associated with each site/ network 
# 2025-09-05 Updated to increase parallelism 
# 2025-09-05 updated to include a few more colums 
# 2025-09-05 updated to check access ports only (no need for trunk ports)
# please ensure the below modules are installed 
    # pip install meraki
    # pip install openpyxl



from __future__ import annotations

import logging
import re
import sys
import time
import threading
from collections import deque
from datetime import datetime
from getpass import getpass
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple, TypedDict, cast
from concurrent.futures import ThreadPoolExecutor, as_completed
import csv

import meraki
from meraki.exceptions import APIError

# ---------- Optional Excel dependency ----------
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.worksheet.filters import AutoFilter
    HAS_OPENPYXL: bool = True
except Exception:
    HAS_OPENPYXL = False
    Workbook = object        # type: ignore[assignment]
    Worksheet = object       # type: ignore[assignment]
    AutoFilter = object      # type: ignore[assignment]
    get_column_letter = None # type: ignore[assignment]

# ---------------- Logging ----------------
timestamp: str = datetime.now().strftime("%Y%m%d_%H%M%S")
logging.basicConfig(
    filename=f"meraki_script_{timestamp}.log",
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)
# (no console handler -> keep console clean)

# ---------------- Constants ----------------
MAX_RETRIES: int = 5
MAX_WORKERS: int = 10          # good balance for ~1000 sites; limiter governs RPS
MERAKI_RPS: int = 5            # conservative org limit
RATE_LIMIT_BURST: int = 5      # allow brief bursts up to 5
RATE_LIMITER_VERBOSE: bool = False
RATE_LIMIT_LOG_THRESHOLD: float = 0.25  # log only if wait >= 250ms
EXCLUDED_VLANS: Set[int] = {100, 110, 210, 220, 230, 235, 240}
CLIENTS_TIMESPAN_DAYS: int = 7
CLIENTS_TIMESPAN_SECS: int = CLIENTS_TIMESPAN_DAYS * 24 * 3600

# ---------------- Types ----------------
class VLANRow(TypedDict):
    organizationName: str
    networkId: str
    networkName: str
    networkProductTypes: str
    configTemplateId: str
    configTemplateName: str
    vlanId: str
    vlanName: str
    subnet: str
    applianceIp: str
    dhcpHandling: str
    clientsOnVlan: str        # "yes"/"no"
    AccessPortsAssignedOnSwitch: str    # "yes"/"no"

class SwitchVlanInfo(TypedDict):
    # explicit_vlans: Set[int]          # access VLANs + explicit trunk tokens
    # native_vlans: Set[int]            # trunk native VLANs
    # ranges: List[Tuple[int, int]]     # parsed ranges from allowedVlans
    # all_allowed: bool                 # any trunk allows all (or 1-4094)
    access_vlans: Set[int]   # VLANs from enabled access ports only

# ---------------- Rate limiter ----------------
class RateLimiter:
    """
    Token-bucket limiter with Condition wait (no busy loop, minimal logging).
    rate = tokens per second, burst = bucket capacity.
    """
    def __init__(self, rate: int, burst: int) -> None:
        self.rate = max(1, rate)
        self.capacity = max(1, burst)
        self.tokens: float = float(self.capacity)
        self.last_refill: float = time.monotonic()
        self._lock = threading.Lock()
        self._cv = threading.Condition(self._lock)

    def acquire(self) -> None:
        with self._cv:
            start = time.monotonic()
            while True:
                now = time.monotonic()
                elapsed = now - self.last_refill
                if elapsed > 0:
                    # Refill based on elapsed time
                    self.tokens = min(self.capacity, self.tokens + elapsed * self.rate)
                    self.last_refill = now

                if self.tokens >= 1.0:
                    self.tokens -= 1.0
                    waited = now - start
                    if RATE_LIMITER_VERBOSE and waited >= RATE_LIMIT_LOG_THRESHOLD:
                        logger.debug("RateLimiter waited %.3fs", waited)
                    return

                # compute time until next token is available
                needed = 1.0 - self.tokens
                wait_for = needed / self.rate  # seconds
                # Wait without spamming logs
                self._cv.wait(timeout=wait_for)

# initialize
limiter = RateLimiter(MERAKI_RPS, RATE_LIMIT_BURST)

# ---------------- Auth + Dashboard ----------------
API_KEY: str = getpass("Enter your API key (input hidden): ")
dashboard: meraki.DashboardAPI = meraki.DashboardAPI(
    API_KEY,
    suppress_logging=True,
    wait_on_rate_limit=True,   # SDK honors Retry-After on 429
    retry_4xx_error=True,
    maximum_retries=10,
    single_request_timeout=60,
)

def select_org() -> Tuple[str, str]:
    orgs: List[Dict[str, Any]] = dashboard.organizations.getOrganizations()
    if not orgs:
        logger.error("No organisations returned from API")
        print("No organisations available for this API key.")
        sys.exit(1)

    print("Organizations:")
    for idx, org in enumerate(orgs, 1):
        name = cast(str, org.get("name", ""))
        oid = cast(str, org.get("id", ""))
        print(f"{idx}. {name} (ID: {oid})")

    raw: str = input("Select organization by number (or press Enter to cancel): ").strip()
    if not raw:
        logger.error("No organisation selected by user")
        print("No organisation selected.")
        sys.exit(1)

    try:
        org_idx = int(raw)
        if org_idx < 1 or org_idx > len(orgs):
            raise ValueError("out of range")
    except Exception:
        logger.error("Invalid organisation selection: %s", raw)
        print("Invalid selection.")
        sys.exit(1)

    chosen: Dict[str, Any] = orgs[org_idx - 1]
    org_id = cast(str, chosen.get("id", ""))
    org_name = cast(str, chosen.get("name", ""))
    logger.debug("Selected org %s (%s)", org_name, org_id)
    return org_id, org_name

org_id, org_name = select_org()
name_filter: str = input("Filter networks by name (partial, optional): ").strip().lower()
only_with_appliance: bool = True  # set False to include all networks

# ---------------- Templates cache ----------------
TEMPLATE_NAME_BY_ID: Dict[str, str] = {}

def load_templates(org_id_param: str) -> None:
    """Populate TEMPLATE_NAME_BY_ID using getOrganizationConfigTemplates (no pagination args)."""
    global TEMPLATE_NAME_BY_ID
    try:
        templates: List[Dict[str, Any]] = api_request_with_retries(  # type: ignore[name-defined]
            dashboard.organizations.getOrganizationConfigTemplates,
            org_id_param
        )
        mapping: Dict[str, str] = {}
        for t in templates or []:
            tid = cast(str, t.get("id", ""))
            tname = cast(str, t.get("name", ""))
            if tid:
                mapping[tid] = tname
        TEMPLATE_NAME_BY_ID = mapping
        logger.debug("Loaded %d config templates", len(TEMPLATE_NAME_BY_ID))
    except Exception:
        logger.exception("Failed to load config templates for org %s", org_id_param)
        TEMPLATE_NAME_BY_ID = {}

# ---------------- Helpers ----------------
def try_int(value: Any) -> Optional[int]:
    try:
        return int(value)
    except Exception:
        return None

def api_request_with_retries(func: Any, *args: Any, **kwargs: Any) -> Any:
    fname: str = getattr(func, "__name__", str(func))  # ensure bound for except paths
    for i in range(MAX_RETRIES):
        limiter.acquire()
        try:
            logger.debug("API call %s attempt %d args=%s kwargs=%s", fname, i + 1, args, kwargs)
            return func(*args, **kwargs)
        except APIError as e:
            status: Optional[int] = getattr(e, "status", None)
            msg: str = getattr(e, "message", str(e))
            retry_after: Optional[float] = None
            ra_any: Any = getattr(e, "retry_after", None)
            try:
                if ra_any is not None:
                    retry_after = float(ra_any)
            except Exception:
                retry_after = None

            if status == 429:
                wait = retry_after if (retry_after and retry_after > 0) else float(2 ** i)
                logger.warning("429 on %s. Sleeping %.2fs (retry %d/%d)", fname, wait, i + 1, MAX_RETRIES)
                time.sleep(wait)
                continue

            logger.error("APIError in %s: status=%s msg=%s", fname, status, msg)
            raise
        except Exception:
            logger.exception("Unexpected error in %s", fname)
            raise
    raise RuntimeError(f"Max retries exceeded for {fname}")

def list_org_networks(org_id_param: str) -> List[Dict[str, Any]]:
    nets: List[Dict[str, Any]] = api_request_with_retries(
        dashboard.organizations.getOrganizationNetworks,
        org_id_param,
        total_pages="all",
        perPage=1000,  # reduce pagination overhead for large orgs
    )
    if name_filter:
        nets = [n for n in nets if name_filter in cast(str, n.get("name", "")).lower()]
    if only_with_appliance:
        nets = [n for n in nets if "appliance" in cast(List[str], n.get("productTypes") or [])]
    logger.debug("Networks after filter: %d", len(nets))
    return nets

def get_network_template_info_from_network_obj(network: Dict[str, Any]) -> Tuple[str, str]:
    """
    Fast path: use configTemplateId already present on the network object, if any.
    Fallback to getNetwork only when absent.
    """
    tmpl_id: str = cast(str, network.get("configTemplateId", "")) or ""
    if tmpl_id:
        return tmpl_id, TEMPLATE_NAME_BY_ID.get(tmpl_id, "")
    # fallback (rare)
    return get_network_template_info_api(cast(str, network.get("id", "")))

def get_network_template_info_api(net_id: str) -> Tuple[str, str]:
    """Fallback: call getNetwork to read configTemplateId, then map to name."""
    if not net_id:
        return "", ""
    try:
        ninfo: Dict[str, Any] = api_request_with_retries(dashboard.networks.getNetwork, net_id)
        tmpl_id: str = cast(str, ninfo.get("configTemplateId", "")) or ""
        tmpl_name: str = TEMPLATE_NAME_BY_ID.get(tmpl_id, "")
        logger.debug("Network %s template id=%s name=%s", net_id, tmpl_id, tmpl_name)
        return tmpl_id, tmpl_name
    except APIError as e:
        status = getattr(e, "status", None)
        if status in (400, 404):
            logger.debug("getNetwork not available for %s (status %s)", net_id, status)
            return "", ""
        logger.exception("APIError reading network %s", net_id)
        return "", ""
    except Exception:
        logger.exception("Unexpected error reading network %s", net_id)
        return "", ""

def parse_allowed_to_spec(allowed: str) -> Tuple[bool, Set[int], List[Tuple[int, int]]]:
    """
    Pre-parse 'allowedVlans' to a fast membership spec:
      returns (all_allowed, explicit_numbers, ranges)
    """
    s = allowed.strip().lower()
    if not s or s == "none":
        return (False, set(), [])
    if s in ("all", "1-4094"):
        return (True, set(), [])
    explicit: Set[int] = set()
    ranges: List[Tuple[int, int]] = []
    for token in s.split(","):
        token = token.strip()
        if not token:
            continue
        if "-" in token:
            a, b = token.split("-", 1)
            ai = try_int(a)
            bi = try_int(b)
            if ai is not None and bi is not None and ai <= bi:
                ranges.append((ai, bi))
        else:
            ti = try_int(token)
            if ti is not None:
                explicit.add(ti)
    return (False, explicit, ranges)

# def get_switch_vlan_info(net_id: str) -> SwitchVlanInfo:
#     """
#     One devices call, then ports per MS device.
#     Pre-parses 'allowedVlans' to avoid repeated string parsing per VLAN.
#     Short-circuits if any trunk port allows 'all'.
#     """
#     info: SwitchVlanInfo = {
#         "explicit_vlans": set(),
#         "native_vlans": set(),
#         "ranges": [],
#         "all_allowed": False,
#     }
#     try:
#         devices: List[Dict[str, Any]] = api_request_with_retries(
#             dashboard.networks.getNetworkDevices, net_id
#         )
#         ms_serials: List[str] = [
#             cast(str, d.get("serial", "")) for d in devices or []
#             if cast(str, d.get("model", "")).startswith("MS")
#         ]
#         for serial in ms_serials:
#             if info["all_allowed"]:
#                 break  # no need to fetch more ports
#             if not serial:
#                 continue
#             ports: List[Dict[str, Any]] = api_request_with_retries(
#                 dashboard.switch.getDeviceSwitchPorts, serial
#             )
#             for p in ports or []:
#                 if info["all_allowed"]:
#                     break
#                 enabled = cast(Optional[bool], p.get("enabled"))
#                 if enabled is False:
#                     continue
#                 ptype = cast(str, p.get("type", ""))  # "access" or "trunk"
#                 if ptype == "access":
#                     ai = try_int(p.get("vlan"))
#                     if ai is not None:
#                         info["explicit_vlans"].add(ai)
#                 elif ptype == "trunk":
#                     ni = try_int(p.get("nativeVlan"))
#                     if ni is not None:
#                         info["native_vlans"].add(ni)
#                     allowed = cast(str, p.get("allowedVlans", ""))
#                     if allowed:
#                         all_allowed, explicit, ranges = parse_allowed_to_spec(allowed)
#                         if all_allowed:
#                             info["all_allowed"] = True
#                             break
#                         info["explicit_vlans"].update(explicit)
#                         info["ranges"].extend(ranges)

#         logger.debug(
#             "Switch VLAN info for %s: explicit=%s native=%s ranges=%d all=%s",
#             net_id,
#             sorted(info["explicit_vlans"]),
#             sorted(info["native_vlans"]),
#             len(info["ranges"]),
#             info["all_allowed"],
#         )
#     except APIError as e:
#         status = getattr(e, "status", None)
#         if status in (400, 404):
#             logger.debug("No switch ports for %s (status %s)", net_id, status)
#             return info
#         logger.exception("APIError fetching switch ports for %s", net_id)
#     except Exception:
#         logger.exception("Unexpected error fetching switch ports for %s", net_id)
#     return info
def get_switch_vlan_info(net_id: str) -> SwitchVlanInfo:
    """
    Collect VLAN IDs configured on ENABLED ACCESS PORTS only.
    Ignores all trunk settings (native/allowed).
    """
    info: SwitchVlanInfo = {"access_vlans": set()}
    try:
        devices: List[Dict[str, Any]] = api_request_with_retries(
            dashboard.networks.getNetworkDevices, net_id
        )
        ms_serials: List[str] = [
            cast(str, d.get("serial", "")) for d in devices or []
            if cast(str, d.get("model", "")).startswith("MS")
        ]
        for serial in ms_serials:
            if not serial:
                continue
            ports: List[Dict[str, Any]] = api_request_with_retries(
                dashboard.switch.getDeviceSwitchPorts, serial
            )
            for p in ports or []:
                enabled = cast(Optional[bool], p.get("enabled"))
                if enabled is False:
                    continue
                if cast(str, p.get("type", "")) != "access":
                    continue
                v = try_int(p.get("vlan"))
                if v is not None:
                    info["access_vlans"].add(v)

        logger.debug(
            "Access-port VLANs for %s: %s",
            net_id, sorted(info["access_vlans"])
        )
    except APIError as e:
        status = getattr(e, "status", None)
        if status in (400, 404):
            logger.debug("No switch ports for %s (status %s)", net_id, status)
            return info
        logger.exception("APIError fetching switch ports for %s", net_id)
    except Exception:
        logger.exception("Unexpected error fetching switch ports for %s", net_id)
    return info


def vlan_in_ranges(ranges: List[Tuple[int, int]], target: int) -> bool:
    for a, b in ranges:
        if a <= target <= b:
            return True
    return False

def get_clients_vlan_set(net_id: str) -> Set[int]:
    """Return VLAN IDs seen among clients in the last N days. Robust to odd API returns."""
    seen: Set[int] = set()
    try:
        result = api_request_with_retries(
            dashboard.networks.getNetworkClients,
            net_id,
            timespan=CLIENTS_TIMESPAN_SECS,
            perPage=1000,
            total_pages="all",
        )

        # Defensive normalization
        if isinstance(result, list):
            items = result
        else:
            # Some SDK/error paths can surface a string/dict; log and bail gracefully
            logger.warning("Unexpected clients payload type for %s: %s", net_id, type(result).__name__)
            return seen

        for c in items:
            if not isinstance(c, dict):
                logger.debug("Skipping non-dict client item of type %s", type(c).__name__)
                continue
            vi = try_int(c.get("vlan"))
            if vi is not None:
                seen.add(vi)

        logger.debug("Clients VLANs for %s: %s", net_id, sorted(seen))
        return seen

    except APIError as e:
        status = getattr(e, "status", None)
        if status in (400, 404):
            logger.debug("No clients data for %s (status %s)", net_id, status)
            return seen
        logger.exception("APIError fetching clients for %s", net_id)
        return seen
    except Exception:
        logger.exception("Unexpected error fetching clients for %s", net_id)
        return seen


def fetch_network_vlan_rows(network: Dict[str, Any]) -> List[VLANRow]:
    """
    Two-step VLAN path, plus enrichment:
      1) getNetworkApplianceVlansSettings -> early skip if vlansEnabled=False
      2) getNetworkApplianceVlans       -> enumerate VLANs (excluding EXCLUDED_VLANS)
      3) configTemplateId + name (fast from network; fallback to getNetwork)
      4) getNetworkClients              -> set of client VLANs (once per network)
      5) getNetworkDevices + getDeviceSwitchPorts -> switch VLAN presence (once per network)
    """
    rows: List[VLANRow] = []
    net_id: str = cast(str, network.get("id", ""))
    net_name: str = cast(str, network.get("name", ""))
    product_types_list: List[str] = cast(List[str], network.get("productTypes") or [])
    net_types: str = ",".join(product_types_list)

    if not net_id:
        return rows

    # Step 1: VLANs enabled?
    try:
        settings: Dict[str, Any] = api_request_with_retries(
            dashboard.appliance.getNetworkApplianceVlansSettings, net_id
        )
        if not (settings and settings.get("vlansEnabled", False)):
            logger.debug("VLANs disabled for %s (%s)", net_name, net_id)
            return rows
    except APIError as e:
        status = getattr(e, "status", None)
        if status in (400, 404):
            logger.debug("No settings for %s (%s) [status %s]", net_name, net_id, status)
            return rows
        logger.exception("Settings error for %s (%s)", net_name, net_id)
        return rows
    except Exception:
        logger.exception("Unexpected settings error for %s", net_name)
        return rows

    # Step 3: Network template info (fast path from network)
    tmpl_id, tmpl_name = get_network_template_info_from_network_obj(network)

    # Step 2: Fetch VLANs for this network (and select non-excluded)
    try:
        vlans: List[Dict[str, Any]] = api_request_with_retries(
            dashboard.appliance.getNetworkApplianceVlans, net_id
        )
    except APIError as e:
        status = getattr(e, "status", None)
        if status in (400, 404):
            logger.debug("No VLANs for %s (%s) [status %s]", net_name, net_id, status)
            return rows
        logger.exception("VLANs error for %s (%s)", net_name, net_id)
        return rows
    except Exception:
        logger.exception("Unexpected VLANs error for %s", net_name)
        return rows

    vlan_objs: List[Dict[str, Any]] = []
    included_vlan_ids_int: Set[int] = set()
    for v in vlans or []:
        vi = try_int(v.get("id", ""))
        if vi is not None and vi in EXCLUDED_VLANS:
            continue
        vlan_objs.append(v)
        if vi is not None:
            included_vlan_ids_int.add(vi)
    if not vlan_objs:
        return rows

    # Step 4 & 5: fetch client VLANs and switch info (sequential here; outer pool provides concurrency)
    clients_vlans: Set[int] = get_clients_vlan_set(net_id)
    svi: SwitchVlanInfo = get_switch_vlan_info(net_id)

    # Build rows
    for v in vlan_objs:
        raw_id: Any = v.get("id", "")
        vi = try_int(raw_id)
        vlan_name = cast(str, v.get("name", ""))
        subnet = cast(str, v.get("subnet", ""))
        appliance_ip = cast(str, v.get("applianceIp", ""))
        dhcp = cast(str, v.get("dhcpHandling", ""))

        clients_yes = "yes" if (vi is not None and vi in clients_vlans) else "no"

        # switch_yes = "no"
        # if vi is not None:
        #     if svi["all_allowed"]:
        #         switch_yes = "yes"
        #     elif vi in svi["explicit_vlans"] or vi in svi["native_vlans"]:
        #         switch_yes = "yes"
        #     elif vlan_in_ranges(svi["ranges"], vi):
        #         switch_yes = "yes"

        switch_yes = "yes" if (vi is not None and vi in svi["access_vlans"]) else "no"

        row: VLANRow = VLANRow(
            organizationName=org_name,
            networkId=net_id,
            networkName=net_name,
            networkProductTypes=net_types,
            configTemplateId=tmpl_id,
            configTemplateName=tmpl_name,
            vlanId=str(raw_id),
            vlanName=vlan_name,
            subnet=subnet,
            applianceIp=appliance_ip,
            dhcpHandling=dhcp,
            clientsOnVlan=clients_yes,
            AccessPortsAssignedOnSwitch=switch_yes,
        )
        rows.append(row)

    logger.debug("Built %d VLAN rows (client/switch flags precompiled) for %s", len(rows), net_name)
    return rows

def sanitize_filename_part(text: str) -> str:
    cleaned = re.sub(r"[^\w\s\-\._]", "", text).strip()
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned or "org"

# ---------- Tabular writers (Excel preferred, CSV/TSV fallback) ----------
def data_has_commas(rows: List[VLANRow], keys: List[str]) -> bool:
    for r in rows:
        for k in keys:
            val = str(r.get(k, ""))
            if "," in val:
                return True
    return False

def write_csv_or_tsv(filepath_no_ext: str, rows: List[VLANRow], fieldnames: List[str]) -> str:
    use_tsv: bool = data_has_commas(rows, fieldnames)
    delimiter: str = "\t" if use_tsv else ","
    ext: str = "tsv" if use_tsv else "csv"
    out_path: str = f"{filepath_no_ext}.{ext}"

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        for r in rows:
            writer.writerow({k: r.get(k, "") for k in fieldnames})

    logger.info("Wrote %s (%d rows)", out_path, len(rows))
    return out_path

def write_excel(out_path: str, rows: List[VLANRow], fieldnames: List[str]) -> None:
    wb: Workbook = Workbook()  # type: ignore[call-arg]
    ws_opt = getattr(wb, "active", None)
    if ws_opt is None:
        ws_opt = wb.create_sheet(title="VLANs")  # type: ignore[operator]
    ws: Worksheet = cast(Worksheet, ws_opt)      # type: ignore[assignment]
    ws.title = "VLANs"

    ws.append(fieldnames)
    for r in rows:
        ws.append([r.get(k, "") for k in fieldnames])

    ws.freeze_panes = "A2"
    if get_column_letter is not None:
        last_col_letter = get_column_letter(len(fieldnames))  # type: ignore[operator]
        last_row = ws.max_row
        ws.auto_filter = AutoFilter(ref=f"A1:{last_col_letter}{last_row}")  # type: ignore[call-arg]

        col_widths: List[int] = [len(h) for h in fieldnames]
        for row_vals in ws.iter_rows(min_row=2, max_row=last_row, max_col=len(fieldnames), values_only=True):
            row_iter: Iterable[Optional[Any]] = cast(Iterable[Optional[Any]], row_vals)
            for i, val in enumerate(row_iter):
                l = len(str(val)) if val is not None else 0
                if l > col_widths[i]:
                    col_widths[i] = l
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 60)  # type: ignore[operator]

    wb.save(out_path)
    logger.info("Excel written: %s", out_path)

def write_tabular(org_name_val: str, rows: List[VLANRow], fieldnames: List[str]) -> str:
    org_part: str = sanitize_filename_part(org_name_val)
    base_path: str = f"org_vlans_{org_part}_{timestamp}"

    if HAS_OPENPYXL:
        out_xlsx = f"{base_path}.xlsx"
        try:
            write_excel(out_xlsx, rows, fieldnames)
            return out_xlsx
        except Exception:
            logger.exception("Excel write failed; falling back to CSV/TSV")
            return write_csv_or_tsv(base_path, rows, fieldnames)
    else:
        logger.debug("openpyxl not available; writing CSV/TSV")
        return write_csv_or_tsv(base_path, rows, fieldnames)

def sort_key_vlan(row: VLANRow) -> Tuple[str, Tuple[int, str]]:
    vlan_int = try_int(row["vlanId"])
    return (row["networkName"], (vlan_int if vlan_int is not None else 10**9, row["vlanId"]))

# ---------------- Main ----------------
def main() -> None:
    # Load templates once for the org (ID -> Name map)
    load_templates(org_id)

    networks: List[Dict[str, Any]] = list_org_networks(org_id)
    if not networks:
        print("No matching networks found in this organization.")
        return

    print(f"Scanning {len(networks)} network(s)...")  # minimal console
    logger.debug("ThreadPool workers: %d", MAX_WORKERS)

    all_rows: List[VLANRow] = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_map = {executor.submit(fetch_network_vlan_rows, n): n for n in networks}
        for fut in as_completed(future_map):
            n = future_map[fut]
            try:
                rows = fut.result()
                all_rows.extend(rows)
            except Exception:
                logger.exception("Worker failed for %s", n.get("name", "<unknown>"))

    all_rows.sort(key=sort_key_vlan)

    fieldnames: List[str] = [
        "organizationName",
        "networkId",
        "networkName",
        "networkProductTypes",
        "configTemplateId",
        "configTemplateName",
        "vlanId",
        "vlanName",
        "subnet",
        "applianceIp",
        "dhcpHandling",
        "clientsOnVlan",
        "AccessPortsAssignedOnSwitch",
    ]

    out_file: str = write_tabular(org_name, all_rows, fieldnames)

    nets_with_vlans: Set[str] = {r["networkId"] for r in all_rows}
    print(
        f"\n✅ Completed. Networks scanned: {len(networks)} | "
        f"Networks with VLANs (after excludes): {len(nets_with_vlans)} | Rows: {len(all_rows)}"
    )
    print(f"Output: {out_file}")
    logger.info("Completed run: networks=%d rows=%d file=%s", len(networks), len(all_rows), out_file)

if __name__ == "__main__":
    main()
